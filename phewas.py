import math
import re
import pandas
import numpy
import seaborn as sns
import statsmodels.formula.api as smf


import phewas_preprocess
from phewas_preprocess import self_report_circadian_variables
import phewas_tests
from phewas_tests import covariates, quantitative_tests, survival_covariates, OLS
import phewas_plots as plots
import day_plots

import util
from util import BH_FDR, legend_of_pointscale, legend_from_colormap, truncate, wrap

def summary():
    # Summarize the phecode test results
    num_nonnull = len(phecode_tests) - phecode_tests.p.sum()*2
    bonferonni_cutoff = 0.05 / len(phecode_tests)
    FDR_cutoff = phecode_tests[phecode_tests.q < 0.05].p.max()
    print(f"Of {len(phecode_tests)} tested, approx {int(num_nonnull)} expected non-null")
    print(f"and {(phecode_tests.p <= bonferonni_cutoff).sum()} exceed the Bonferonni significance threshold")
    print(f"and {(phecode_tests.p <= FDR_cutoff).sum()} exceed the FDR < 0.05 significance threshold")

    ### Create summary plots
    fig, ax = pylab.subplots(figsize=(8,6))
    color = phecode_tests.phecode_category.map(color_by_phecode_cat)
    ax.scatter(phecode_tests.N_cases, -numpy.log10(phecode_tests.p), marker="+", c=color)
    ax.set_xlabel("Number cases")
    ax.set_ylabel("-log10 (p-value)")
    ax.axhline( -numpy.log10(bonferonni_cutoff), c="k", zorder = -1 )
    ax.axhline( -numpy.log10(FDR_cutoff), c="k", linestyle="--", zorder = -1 )
    ax.set_title("PheCode - Activity associations")
    fig.savefig(OUTDIR+"phewas_summary.png")


    fig, ax = pylab.subplots(figsize=(8,6))
    pt = phecode_tests.sample(frac=1) # Randomly reorder for the plot
    color = pt.phecode_category.map(color_by_phecode_cat)
    ax.scatter(pt.std_effect, -numpy.log10(pt.p), marker="+", c=color)
    ax.set_xlabel("Effect size")
    ax.set_ylabel("-log10(p-value)")
    ax.axhline( -numpy.log10(bonferonni_cutoff), c="k", zorder = -1 )
    ax.axhline( -numpy.log10(FDR_cutoff), c="k", linestyle="--", zorder = -1 )
    ax.set_title("PheCode - Activity associations")
    legend_from_colormap(ax, color_by_phecode_cat, ncol=2, fontsize="small")
    fig.savefig(OUTDIR+"phewas.volcano_plot.png")

    def manhattan_plot(tests_df, group_on="phecode", group_by=None, color_on=None, color_by=None, y_break=20):
        # "Manhattan" plot of the PheWAS
        fig, axes = pylab.subplots(nrows=2, figsize=(12,7), sharex=True)
        x = 0
        x_ticks = []
        x_minorticks = [-0.5]
        x_ticklabels = []
        for key, pt in tests_df.sample(frac=1).groupby(group_by):
            x_ticks.append(x+pt[group_on].nunique()/2-0.5)
            x_ticklabels.append(util.capitalize(key))
            for variable, tests in pt.groupby(group_on):
                for ax in axes:
                    ax.scatter([x for i in range(len(tests))], -numpy.log10(tests.p), color=tests[color_on].map(color_by))
                x += 1
            x_minorticks.append(x-0.5)
        ax_top, ax_bottom = axes
        ax_bottom.set_ylabel("-log10(p-value)")
        ax_bottom.axhline(-numpy.log10(bonferonni_cutoff), c="k", zorder = 2)
        ax_bottom.axhline(-numpy.log10(FDR_cutoff), c="k", linestyle="--", zorder = 2)
        ax_bottom.set_xticks(x_ticks)
        ax_bottom.set_xticks(x_minorticks, minor=True)
        ax_bottom.set_xticklabels(x_ticklabels, rotation=90)
        ax_bottom.xaxis.set_tick_params(which="major", bottom=False, top=False)
        ax_bottom.xaxis.set_tick_params(which="minor", length=10)
        ax_bottom.set_ylim(0,y_break)
        ax_top.set_xmargin(0.01)
        ax_top.set_ylim(y_break)
        ax_top.tick_params(bottom=False, which='both')
        legend_from_colormap(fig, color_by, ncol=2, fontsize="small")
        fig.tight_layout(h_pad=0.01)
        return fig

    # Manhattan plot by phecode
    cats = phecode_tests.phecode_category.fillna("N/A").astype('category')
    cats.cat.reorder_categories([k if k == k else 'N/A' for k in color_by_phecode_cat.keys()], inplace=True)
    fig = manhattan_plot(
        phecode_tests,
        group_on="phecode",
        group_by=cats,
        color_on="Activity Subcategory",
        color_by=color_by_actigraphy_subcat)
    fig.savefig(OUTDIR+"manhattan_plot.png")

    # Manhattan plot by activity variable
    cats = phecode_tests['Activity Subcategory'].fillna("N/A").astype('category')
    cats.cat.reorder_categories([k if k == k else 'N/A' for k in color_by_actigraphy_subcat.keys()], inplace=True)
    fig = manhattan_plot(
        phecode_tests,
        group_on="activity_var",
        group_by=cats,
        color_on="phecode_category",
        color_by=color_by_phecode_cat)
    fig.savefig(OUTDIR+"manhattan_plot.by_activity_var.png")

    # TODO: bonferroni and FDR cutoffs change for quantitative
    # "Manhattan" plot of the PheWAS in QUANTITATIVE
    cats = quantitative_tests['Functional Category'].astype('category')
    cats.cat.reorder_categories(color_by_quantitative_function.keys(), inplace=True)
    fig = manhattan_plot(
        quantitative_tests,
        group_on="phenotype",
        group_by=cats,
        color_on="Activity Subcategory",
        color_by=color_by_actigraphy_subcat,
        y_break=40)
    fig.savefig(OUTDIR+"manhattan_plot.quantitative.png")

    cats = quantitative_tests['Activity Subcategory'].astype('category')
    cats.cat.reorder_categories(color_by_actigraphy_subcat.keys(), inplace=True)
    fig = manhattan_plot(
        quantitative_tests,
        group_on="activity_var",
        group_by=cats,
        color_on="Functional Category",
        color_by=color_by_quantitative_function,
        y_break=40)
    fig.savefig(OUTDIR+"manhattan_plot.quantitative.by_activity_var.png")


    ### TIMELINE
    # Make a timeline of the study design timing so that readers can easily see when data was collected
    ACTIGRAPHY_COLOR = "#1b998b"
    REPEAT_COLOR = "#c5d86d"
    DIAGNOSIS_COLOR = "#f46036"
    ASSESSMENT_COLOR = "#aaaaaa"
    DEATH_COLOR = "#333333"
    fig, (ax1, ax2, ax3) = pylab.subplots(figsize=(8,6), nrows=3)
    #ax2.yaxis.set_inverted(True)
    ax1.yaxis.set_label_text("Participants / month")
    ax2.yaxis.set_label_text("Diagnoses / month")
    ax3.yaxis.set_label_text("Deaths / month")
    #ax2.xaxis.tick_top()
    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)
    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_visible(False)
    ax3.spines["top"].set_visible(False)
    ax3.spines["right"].set_visible(False)
    bins = pandas.date_range("2000-1-1", "2020-6-1", freq="1M")
    def date_hist(ax, values, bins, **kwargs):
        # Default histogram for dates doesn't cooperate with being given a list of bins
        # since the list of bins doesn't get converted to the same numerical values as the values themselves
        counts, edges = numpy.histogram(values, bins)
        # Fudge factor fills in odd gaps between boxes
        ax.bar(edges[:-1], counts, width=(edges[1:]-edges[:-1])*1.05, **kwargs)
    assessment_time = pandas.to_datetime(data.blood_sample_time_collected_V0)
    actigraphy_time = pandas.to_datetime(activity_summary.loc[data.index, 'file-startTime'])
    actigraphy_seasonal_time = pandas.to_datetime(activity_summary_seasonal.loc[activity_summary_seasonal.ID.isin(data.index), 'file-startTime'], cache=False)
    death_time = pandas.to_datetime(data[~data.date_of_death.isna()].date_of_death)
    diagnosis_time = pandas.to_datetime(icd10_entries[icd10_entries.ID.isin(data.index)].first_date)
    date_hist(ax1, assessment_time, color=ASSESSMENT_COLOR, label="assessment", bins=bins)
    date_hist(ax1, actigraphy_time, color=ACTIGRAPHY_COLOR, label="actigraphy", bins=bins)
    date_hist(ax1, actigraphy_seasonal_time, color=REPEAT_COLOR, label="repeat actigraphy", bins=bins)
    date_hist(ax2, diagnosis_time, color=DIAGNOSIS_COLOR, label="Diagnoses", bins=bins)
    date_hist(ax3, death_time, color=DEATH_COLOR, label="Diagnoses", bins=bins)
    ax1.annotate("Assessment", (assessment_time.mean(), 0), xytext=(0,75), textcoords="offset points", ha="center")
    ax1.annotate("Actigraphy", (actigraphy_time.mean(), 0), xytext=(0,75), textcoords="offset points", ha="center")
    ax1.annotate("Repeat\nActigraphy", (actigraphy_seasonal_time.mean(), 0), xytext=(0,70), textcoords="offset points", ha="center")
    ax2.annotate("Medical Record\nDiagnoses", (diagnosis_time.mean(), 0), xytext=(0,60), textcoords="offset points", ha="center")
    ax3.annotate("Deaths", (death_time.mean(), 0), xytext=(0,70), textcoords="offset points", ha="center")
    fig.savefig(OUTDIR+"summary_timeline.png")

    time_difference = (actigraphy_time - assessment_time).mean()
    print(f"Mean difference between actigraphy time and initial assessment time: {time_difference/pandas.to_timedelta('1Y')} years")


    ### Diagnosis summary
    num_diagnoses = icd10_entries.groupby(pandas.Categorical(icd10_entries.ID, categories=data.index)).size()
    icd10_entries_at_actigraphy = icd10_entries[pandas.to_datetime(icd10_entries.first_date) < pandas.to_datetime(icd10_entries.ID.map(activity_summary['file-startTime']))]
    num_diagnoses_at_actigraphy = icd10_entries_at_actigraphy.groupby(pandas.Categorical(icd10_entries_at_actigraphy.ID, categories=data.index)).size()
    icd10_entries_at_assessment = icd10_entries[pandas.to_datetime(icd10_entries.first_date) < pandas.to_datetime(icd10_entries.ID.map(data['blood_sample_time_collected_V0']))]
    num_diagnoses_at_assessment = icd10_entries_at_assessment.groupby(pandas.Categorical(icd10_entries_at_assessment.ID, categories=data.index)).size()
    ID_without_actigraphy = ukbb.index[ukbb.actigraphy_file.isna()]
    icd10_entries_without_actigraphy = icd10_entries_all[icd10_entries_all.ID.isin(ID_without_actigraphy)]
    num_diagnoses_no_actigraphy = icd10_entries_without_actigraphy.groupby(pandas.Categorical(icd10_entries_without_actigraphy.ID, categories=ID_without_actigraphy)).size()
    icd10_entries_at_assessment_without_actigraphy = icd10_entries_without_actigraphy[pandas.to_datetime(icd10_entries_without_actigraphy.first_date) < pandas.to_datetime(icd10_entries_without_actigraphy.ID.map(ukbb['blood_sample_time_collected_V0']))]
    num_diagnoses_at_assessment_without_actigraphy = icd10_entries_at_assessment_without_actigraphy.groupby(pandas.Categorical(icd10_entries_at_assessment_without_actigraphy.ID, categories=ID_without_actigraphy)).size()
    fig,ax = pylab.subplots()
    color_by_actigraphy = {"With Actigraphy": "#CCC", "Without Actigraphy": "#C33"}
    actigraphy_boxes = ax.boxplot([num_diagnoses_at_assessment, num_diagnoses_at_actigraphy, num_diagnoses],
                                   showfliers=False, positions=[0,1.2,2], widths=0.4, patch_artist=True,
                                   boxprops={"color": "k", "facecolor": color_by_actigraphy['With Actigraphy']},
                                   medianprops={"color":"k"})
    no_actigraphy_boxes = ax.boxplot([num_diagnoses_at_assessment_without_actigraphy, num_diagnoses_no_actigraphy],
                                        showfliers=False, positions=[0.4,2.4], widths=0.4, patch_artist=True,
                                        boxprops={"color": "k", "facecolor": color_by_actigraphy["Without Actigraphy"]},
                                        medianprops={"color":"k"})
    ax.set_xticks([0.2,1.2,2.2])
    ax.set_xticklabels(["At Assessment", "At Actigraphy", "At Study End"])
    ax.set_ylabel("Medical Record Diagnoses per Participant")
    ax.set_title("Disease Burden")
    legend_from_colormap(ax, color_by_actigraphy)
    fig.savefig(OUTDIR+"summary_disease_burden.png")

    print(f"Median number of diagnoses by category:")
    print("At assessment:", num_diagnoses_at_assessment.describe())
    print("At actigraphy:", num_diagnoses_at_actigraphy.describe())
    print("Final:", num_diagnoses.describe())
    print("Final without actigraphy:", num_diagnoses_no_actigraphy.describe())

    ### Overall disease burden (number of phecodes) versus RA
    fig, ax = pylab.subplots()
    num_phecodes = data[phecode_groups].sum(axis=1)
    phecode_ranges = pandas.cut(num_phecodes, [0,1,6,11,16,21,26,31,num_phecodes.max()+1])
    xticklabels = []
    for i, (phecode_range, group) in enumerate(data.groupby(phecode_ranges)):
        ax.boxplot(group.acceleration_RA.values, positions=[i], showfliers=False, widths=0.8)
        if phecode_range.right == phecode_range.left+1:
            xticklabels.append(f"{int(phecode_range.left)}")
        else:
            xticklabels.append(f"{int(phecode_range.left)}-{int(phecode_range.right-1)}")
    ax.set_xticks(range(len(xticklabels)))
    ax.set_xticklabels(xticklabels)
    ax.set_xlabel("Number of unique diagnoses")
    ax.set_ylabel("RA")
    fig.savefig(OUTDIR+"num_phecodes.RA.png")

def sex_difference_plots():
    N_CUTOFF = {1: 300, 2:750}[COHORT] # Get cutoff by cohort number
    d = phecode_tests_by_sex[True #(phecode_tests_by_sex.q < 0.05 )
                            & (phecode_tests_by_sex.N_male > N_CUTOFF)
                            & (phecode_tests_by_sex.N_female > N_CUTOFF)]
    ## Create the phewas_plots.sex_difference_plots
    fig, ax = phewas_plots.sex_difference_plot(d, cmap=color_by_phecode_cat)
    fig.savefig(f"{OUTDIR}/sex_differences.all_phenotypes.png")

    fig, ax = phewas_plots.sex_difference_plot(d[d.phecode_category == 'circulatory system'], color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/sex_differences.circulatory.png")
    fig, ax = phewas_plots.sex_difference_plot(d[d.phecode_category == 'mental disorders'], color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/sex_differences.mental_disorders.png")
    fig, ax = phewas_plots.sex_difference_plot(d[d.phecode_category == 'endocrine/metabolic'], color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/sex_differences.endocrine.png")
    fig, ax = phewas_plots.sex_difference_plot(d[d.phecode_category == 'infectious diseases'], color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/sex_differences.infections.png")
    fig, ax = phewas_plots.sex_difference_plot(d[d.phecode_category == 'respiratory'], color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/sex_differences.respiratory.png")

    #Make 2x2 grid of sex differences
    fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
    ij = [[0,0], [0,1], [1,0], [1,1]]
    SUBCATEGORIES = ["circulatory system", "mental disorders", "endocrine/metabolic", "respiratory"]
    for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
        tests = d[d.phecode_category == cat]
        phewas_plots.sex_difference_plot(tests.sample(frac=1), color_by="phecode_meaning", ax=ax, legend=True, labels=False, cmap="tab20_r")
        ax.set_title(cat)
        if j == 0:
            ax.set_ylabel("Effect size in females")
        if i == 1:
            ax.set_xlabel("Effect size in males")
    fig.tight_layout()
    fig.savefig(OUTDIR+"sex_differences.2x2.png")

    # Run sex-difference and age-difference plots on the quantitative tests
    qt = quantitative_sex_tests.sample(frac=1)
    #TODO: I believe the above needs to rename columns std_male_effect -> std_male_coeff and same for female to work below
    fig, ax = phewas_plots.sex_difference_plot(qt, color_by="Functional Category", cmap=color_by_quantitative_function, lim=0.25)
    fig.savefig(OUTDIR+"sex_differences.quantitative.png")
    fig, ax = phewas_plots.sex_difference_plot(qt, color_by="Activity Subcategory", cmap=color_by_actigraphy_subcat, lim=0.25)
    fig.savefig(OUTDIR+"sex_differences.quantitative.by_activity_var.png")

    #Make 2x2 grid of quantitative sex differences
    fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
    ij = [[0,0], [0,1], [1,0], [1,1]]
    SUBCATEGORIES = ["Metabolism", "Lipoprotein Profile", "Cardiovascular Function", "Renal Function"]
    for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
        tests = quantitative_sex_tests[quantitative_sex_tests['Functional Category'] == cat]
        phewas_plots.sex_difference_plot(tests.sample(frac=1), color_by="phenotype", lim=0.25, ax=ax, legend=True, labels=False, cmap="tab20_r", names=quantitative_variable_descriptions.Name)
        ax.set_title(cat)
        if j == 0:
            ax.set_ylabel("Effect size in females")
        if i == 1:
            ax.set_xlabel("Effect size in males")
    fig.tight_layout()
    fig.savefig(OUTDIR+"sex_differences.quantitative.2x2.png")

def age_difference_plots():
    ## Plot summary of age tests
    N_CUTOFF = {1: 500, 2: 1250}[COHORT] # Get cutoff by cohort number
    d = pandas.merge(
            age_tests,
            phecode_tests[['phecode', 'activity_var', 'std_effect', 'p', 'q']],
            suffixes=["_age", "_overall"],
            on=["activity_var", "phecode"]).reset_index()
    d = d[d.N_cases > N_CUTOFF]

    fig, ax = phewas_plots.age_effect_plot(d)
    fig.savefig(f"{OUTDIR}/age_effects.png")

    fig, ax = phewas_plots.age_effect_plot(d[d.phecode_category == 'mental disorders'], labels=False, color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/age_effects.mental_disorders.png")
    fig, ax = phewas_plots.age_effect_plot(d[d.phecode_category == 'circulatory system'], labels=False, color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/age_effects.circulatory.png")
    fig, ax = phewas_plots.age_effect_plot(d[d.phecode_category == 'endocrine/metabolic'], labels=False, color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/age_effects.endorcine.png")
    fig, ax = phewas_plots.age_effect_plot(d[d.phecode_category == 'genitourinary'], labels=False, color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/age_effects.genitourinary.png")
    fig, ax = phewas_plots.age_effect_plot(d[d.phecode_category == 'respiratory'], labels=False, color_by="phecode_meaning")
    fig.savefig(f"{OUTDIR}/age_effects.respiratory.png")

    #Make 2x2 grid of age effect plots
    fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
    ij = [[0,0], [0,1], [1,0], [1,1]]
    SUBCATEGORIES = ["circulatory system", "mental disorders", "endocrine/metabolic", "respiratory"]
    for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
        tests = d[d.phecode_category == cat]
        phewas_plots.age_effect_plot(tests.sample(frac=1), color_by="phecode_meaning", ax=ax, legend=True, labels=False, cmap="tab20_r")
        ax.set_title(cat)
        if j == 0:
            ax.set_ylabel("Effect size at 70")
        if i == 1:
            ax.set_xlabel("Effect size at 55")
    fig.tight_layout()
    fig.savefig(OUTDIR+"age_effects.2x2.png")


    ## age effect for quantitative traits
    dage = quantitative_age_tests.copy()
    dage['p_overall'] = pandas.merge(
        dage,
        quantitative_tests,
        on=["phenotype", "activity_var"],
    ).p
    dage['p_age'] = dage.age_difference_p
    fig, ax = phewas_plots.age_effect_plot(dage.sample(frac=1), color_by="Functional Category", cmap=color_by_quantitative_function, lim=0.3)
    fig.savefig(f"{OUTDIR}/age_effects.quantitative.png")

    #Make 2x2 grid of quantitative age differences
    fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
    ij = [[0,0], [0,1], [1,0], [1,1]]
    SUBCATEGORIES = ["Metabolism", "Lipoprotein Profile", "Cardiovascular Function", "Renal Function"]
    for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
        tests = dage[dage['Functional Category'] == cat]
        phewas_plots.age_effect_plot(tests.sample(frac=1), color_by="phenotype", lim=0.25, ax=ax, legend=True, labels=False, cmap="tab20_r", names=quantitative_variable_descriptions.Name)
        ax.set_title(cat)
        if j == 0:
            ax.set_ylabel("Effect size at 70")
        if i == 1:
            ax.set_xlabel("Effect size at 55")
    fig.tight_layout()
    fig.savefig(OUTDIR+"age_effects.quantitative.2x2.png")

def fancy_plots():
    # Hypertension
    fig = phewas_plots.fancy_case_control_plot(data, 401, normalize=False, confidence_interval=True, annotate=True)
    fig.savefig(OUTDIR+"phenotypes.hypertension.png")

    fig = phewas_plots.incidence_rate_by_category(data, 401, categories="birth_year_category", confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.hypertension.by_age.png")

    sns.lmplot(x="acceleration_RA", y="cholesterol", data=data, hue="birth_year_category", markers='.')
    pylab.gcf().savefig(OUTDIR+"phenotypes.LDL.png")

    sns.lmplot(x="acceleration_RA", y="hdl_cholesterol", data=data, hue="birth_year_category", markers='.')
    pylab.gcf().savefig(OUTDIR+"phenotypes.HDL.png")

    sns.lmplot(x="acceleration_RA", y="systolic_blood_pressure_V0", data=data, hue="birth_year_category", markers='.')
    pylab.gcf().savefig(OUTDIR+"phenotypes.systolic_blood_pressure.png")

    sns.lmplot(x="acceleration_RA", y="diastolic_blood_pressure_V0", data=data, hue="birth_year_category", markers='.')
    pylab.gcf().savefig(OUTDIR+"phenotypes.diastolic_blood_pressure.png")

    # Diabetes
    fig = phewas_plots.fancy_case_control_plot(data, 250, normalize=False, confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.diabetes.png")

    fig = phewas_plots.incidence_rate_by_category(data, 250, categories="birth_year_category", confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.diabetes.by_age.png")

    sns.lmplot(x="acceleration_RA", y="glycated_heamoglobin", data=data, hue="birth_year_category", markers='.')
    pylab.gcf().savefig(OUTDIR+"phenotypes.glycated_heamoglobin.png")

    percent_diabetes_with_hypertension = (data[401].astype(bool) & data[250].astype(bool)).mean() / data[250].mean()
    print(f"Percentage of participants with diabetes that also have hypertension: {percent_diabetes_with_hypertension:0.2%}")

    # Mood disorders
    #TODO!!
    #fig = phewas_plots.fancy_case_control_plot(data, 250, normalize=False, confidence_interval=True)
    #fig.savefig(OUTDIR+"phenotypes.diabetes.png")

    # Some more phenotypes:
    fig = phewas_plots.fancy_case_control_plot(data, 585, normalize=True, confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.renal_failure.png")
    fig = phewas_plots.fancy_case_control_plot(data, 276, normalize=True, confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.disorders_fuild_electrolyte_etc.png")
    fig = phewas_plots.fancy_case_control_plot(data, 290, normalize=True, confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.delirium_dementia_alzheimers.png")
    fig = phewas_plots.fancy_case_control_plot(data, 332, normalize=True, confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.parkinsons.png")
    fig = phewas_plots.fancy_case_control_plot(data, 480, normalize=True, confidence_interval=True)
    fig.savefig(OUTDIR+"phenotypes.pneumonia.png")

def survival_curves():
    # Survival by RA
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "acceleration_RA", "RA")
    fig.savefig(OUTDIR+"survival.RA.png")
    death_counts.to_csv(OUTDIR+"deaths.RA.by_year.txt", sep="\t")

    # Survival by main_sleep_offset_avg
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "main_sleep_offset_mean", "Sleep Offset")
    fig.savefig(OUTDIR+"survival.main_sleep_offset_mean.png")

    # Survival by MVPA_overall_avg
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "MVPA_overall", "MVPA Mean")
    fig.savefig(OUTDIR+"survival.MVPA_overall.png")

    # Survival by MVPA_overall_avg
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "MVPA_hourly_SD", "MVPA hourly SD")
    fig.savefig(OUTDIR+"survival.MVPA_hourly_SD.png")

    # Survival by acceleration_hourly_SD
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "acceleration_hourly_SD", "Acceleration Hourly SD")
    fig.savefig(OUTDIR+"survival.acceleration_hourly_SD.png")
    death_counts.to_csv(OUTDIR+"deaths.acceleration_hourly_SD.by_year.txt", sep="\t")

    # Survival by main_sleep_ratio_mean
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "main_sleep_ratio_mean", "Sleep Ratio")
    fig.savefig(OUTDIR+"survival.main_sleep_ratio_mean.png")
    death_counts.to_csv(OUTDIR+"deaths.main_sleep_ratio_mean.by_year.txt", sep="\t")

    # Survival by phase
    fig, ax = pylab.subplots()
    data['phase_adjusted'] = (data.phase - 8) % 24 + 8
    fig, death_counts = phewas_plots.quintile_survival_plot(data, "phase_adjusted", "phase")
    fig.savefig(OUTDIR+"survival.phase.png")

    # Survival by sex
    fig, death_counts = phewas_plots.quintile_survival_plot_by_cat(data, "acceleration_RA", "sex", "RA")
    fig.savefig(OUTDIR+"survival.RA.by_sex.png")
    fig, death_counts = phewas_plots.quintile_survival_plot_by_cat(data, "main_sleep_ratio_mean", "sex", "Sleep Ratio")
    fig.savefig(OUTDIR+"survival.main_sleep_ratio.by_sex.png")
    fig, death_counts = phewas_plots.quintile_survival_plot_by_cat(data, "acceleration_hourly_SD", "sex", "Acceleration Hourly SD")
    fig.savefig(OUTDIR+"survival.acceleration_hourly_SD.by_sex.png")
    fig, death_counts = phewas_plots.quintile_survival_plot_by_cat(data, "moderate_between_day_SD", "sex", "Moderate between-day SD")
    fig.savefig(OUTDIR+"survival.moderate_between_day_SD.by_sex.png")
    fig, death_counts = phewas_plots.quintile_survival_plot_by_cat(data, "MET_overall", "sex", "MET")
    fig.savefig(OUTDIR+"survival.MET_overall.by_sex.png")

def survival_plots():
    ### Plot survival assocations versus inter/intra personal variance for validation
    fig, ax = pylab.subplots(figsize=(8,6))
    colorby = survival_tests.activity_var.map(activity_variable_descriptions.Category)
    color = colorby.map(color_by_actigraphy_cat)
    def get_variance_ratio(var):
        if var.endswith("_abs_dev"):
            var = var[:-8]
        try:
            return activity_variance.corrected_intra_personal_normalized.loc[var]
        except KeyError:
            print(var)
            return float("NaN")
    variance_ratio = survival_tests.activity_var.apply(get_variance_ratio)
    variance_ratio.index = survival_tests.activity_var
    ax.scatter(survival_tests['standardized log Hazard Ratio'],
                variance_ratio,
                s=plots.p_to_size(survival_tests.p),
                c=color)
    ax.set_xlabel("Standardized log Hazard Ratio")
    ax.set_ylabel("Within-person variation / Between-person variation")
    ax.set_ylim(0,1)
    ax.axvline(0, c='k')
    for index, row in survival_tests.sort_values(by="p").head(20).iterrows():
        # Label the top points
        ax.annotate(
            row.activity_var,
            (
            row['standardized log Hazard Ratio'],
            variance_ratio.loc[row.activity_var]),
            xytext=(0,15),
            textcoords="offset pixels",
            arrowprops={'arrowstyle':"->"})
    util.legend_from_colormap(fig, color_by_actigraphy_cat, ncol=2, fontsize="small")
    #fig.tight_layout()
    fig.savefig(OUTDIR+"survival_versus_variation.svg")
    fig.savefig(OUTDIR+"survival_versus_variation.eps")

    ### Sex-difference survival plot
    def sex_difference_survival_plot(d, color_by="activity_var_category"):
        colormap = {cat:color for cat, color in
                            zip(d[color_by].unique(),
                                [pylab.get_cmap("Set3")(i) for i in range(20)])}
        color = [colormap[c] for c in d[color_by]]
        fig, ax = pylab.subplots(figsize=(9,9))
        # The points
        ax.scatter(
            d.std_male_logHR,
            d.std_female_logHR,
            label="phenotypes",
            s=plots.p_to_size(d.p),
            c=color)
        ax.set_title("Survival associations by sex")
        ax.spines['bottom'].set_color(None)
        ax.spines['top'].set_color(None)
        ax.spines['left'].set_color(None)
        ax.spines['right'].set_color(None)
        ax.axvline(c="k", lw=1)
        ax.axhline(c="k", lw=1)
        ax.set_xlabel("log HR in males")
        ax.set_ylabel("log HR in females")
        ax.set_aspect("equal")
        #ax.set_xlim(-0.5,0.5)
        #ax.set_ylim(-0.5,0.5)
        # Diagonal y=x line
        bound = max(abs(numpy.min([ax.get_xlim(), ax.get_ylim()])),
                    numpy.max([ax.get_xlim(), ax.get_ylim()]))
        diag = numpy.array([-bound, bound])
        ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
        ax.plot(diag, -diag, linestyle="--", c='k', zorder=-1, label="diagonal")
        bbox = {'facecolor': (1,1,1,0.8), 'edgecolor':(0,0,0,0)}
        #ax.annotate("Male Effect Larger", xy=(0.4,0), ha="center", bbox=bbox, zorder=3)
        #ax.annotate("Male Effect Larger", xy=(-0.4,0), ha="center", bbox=bbox, zorder=3)
        #ax.annotate("Female Effect Larger", xy=(0,0.4), ha="center", bbox=bbox, zorder=3)
        #ax.annotate("Female Effect Larger", xy=(0,-0.25), ha="center", bbox=bbox, zorder=3)
        util.legend_from_colormap(ax, colormap, loc="upper left")
        return fig, ax
    d = survival_tests.copy()
    activity_var_stds = data[activity_variables].std() #TODO: should we separate male/female stds?
    d['std_male_logHR'] = d.activity_var.map(activity_var_stds) * d['male_logHR']
    d['std_female_logHR'] = d.activity_var.map(activity_var_stds) * d['female_logHR']
    d['activity_var_category'] = d.activity_var.map(activity_variable_descriptions.Subcategory)
    fig, ax = sex_difference_survival_plot(d)
    fig.savefig(OUTDIR+"survival.by_sex.png")

def circadian_component_plots():
    ## Plot the amount RA goes "beyond" other variables
    fig, (ax1, ax2) = pylab.subplots(ncols=2, figsize=(10,6))
    c = beyond_RA_tests.Subcategory.map(color_by_actigraphy_subcat)
    ax1.scatter(
        beyond_RA_tests['standardized log Hazard Ratio'].abs(),
        beyond_RA_tests['standardized log Hazard Ratio RA'].abs(),
        c=c)
    ax1.set_ylim(0, ax1.get_ylim()[1])
    ax1.set_xlabel("log Hazard Ratio / SD of alternative variable")
    ax1.set_ylabel("log Hazard Ratio / SD of RA")
    ax1.axhline(survival_tests.loc[survival_tests.activity_var == "acceleration_RA", "standardized log Hazard Ratio"].abs().values,
                linestyle="--", c="k")
    ax2.scatter(
        -numpy.log10(beyond_RA_tests.p),
        -numpy.log10(beyond_RA_tests.p_RA),
        c=c)
    ax2.set_ylim(0, ax2.get_ylim()[1])
    ax2.set_xlabel("-log10 p-value of alternative variable")
    ax2.set_ylabel("-log10 p-value of RA")
    ax2.axhline(-numpy.log10(survival_tests.loc[survival_tests.activity_var == "acceleration_RA", "p"].values),
                linestyle="--", c="k")
    legend_from_colormap(fig, color_by_actigraphy_subcat, loc="upper left", fontsize="small", ncol=2)
    fig.savefig(OUTDIR+"additive_benefit_RA.png")

    top_phenotypes = phecode_tests[(~phecode_tests.activity_var.str.startswith('self_report')) & (phecode_tests.N_cases > 1000)].sort_values(by='p').phecode.unique()
    fig, axes = phewas_plots.circadian_component_plot(phecode_three_component_tests, top_phenotypes[:20], quantitative=False)
    fig.savefig(OUTDIR+"circadian_vs_other_vars.png")

    top_phenotypes = quantitative_tests[(~quantitative_tests.activity_var.str.startswith('self_report'))].sort_values(by='p').phenotype.unique()
    fig, axes = phewas_plots.circadian_component_plot(quantitative_three_component_tests, top_phenotypes[:20], quantitative=True)
    fig.savefig(OUTDIR+"circadian_vs_other_vars.quantitative.png")

    # Plot the components in a triangular plot
    def trianglize(coords):
        # Take values summing to 1 and convert to the equilateral triangle of length 1 in the plane
        coords = numpy.maximum(coords,0)
        norm_factors = coords.sum(axis=len(coords.shape)-1)
        norm_effs = (coords.T/norm_factors).T
        return norm_effs @ numpy.array([[0,1,0.5], [0,0,numpy.sqrt(3/4)]]).T
    def triangle_frame(ax):
        # Draw axis frame in a triangle
        ax.set_aspect('equal')
        ax.plot([0,1,0.5,0], [0,0,numpy.sqrt(3/4),0], c="k")
        ax.annotate("Circadian", (0,0), horizontalalignment="right")
        ax.annotate("PA", (1,0))
        ax.annotate("Sleep", (0.5,numpy.sqrt(3/4)))
        ax.axis('off')
    def plot_triangular_ses(ax, effs, seses, colors):
        #color = "#999"
        for eff, ses, color in zip(effs.values, seses.values, colors):
            base_point = trianglize(eff)
            circ_point = trianglize(eff + [ses[0],0,0])
            physical_point = trianglize(eff + [0,ses[1],0])
            sleep_point = trianglize(eff + [0,0,ses[2]])
            #ax.plot(*numpy.array([base_point, circ_point]).T, c=color, zorder=-1)
            #ax.plot(*numpy.array([base_point, physical_point]).T, c=color, zorder=-1)
            #ax.plot(*numpy.array([base_point, sleep_point]).T, c=color, zorder=-1)
            #The negative directions
            neg_circ_point = trianglize(eff + [-ses[0],0,0])
            neg_physical_point = trianglize(eff + [0,-ses[1],0])
            neg_sleep_point = trianglize(eff + [0,0,-ses[2]])
            ax.plot(*numpy.array([base_point, neg_circ_point]).T, c=color, zorder=-1)
            ax.plot(*numpy.array([base_point, neg_physical_point]).T, c=color, zorder=-1)
            ax.plot(*numpy.array([base_point, neg_sleep_point]).T, c=color, zorder=-1)
            #poly_x, poly_y= numpy.array([
            #    #base_point,
            #    circ_point,
            #    neg_physical_point,
            #    sleep_point,
            #    neg_circ_point,
            #    physical_point,
            #    neg_sleep_point,
            #]).T
            #print(poly_x, poly_y)
            #ax.fill(
            #    poly_x,
            #    poly_y,
            #    color=color,
            #    alpha=0.2,
            #    zorder=-1)

    effs_cols = ['circ_eff', 'physical_eff', 'sleep_eff']
    ses_cols = ['circ_ses', 'physical_ses', 'sleep_ses']
    effs = phecode_three_component_tests[effs_cols].abs()
    effs.set_index(phecode_three_component_tests.phenotype, inplace=True)
    ses = phecode_three_component_tests[ses_cols].abs()
    ses.set_index(phecode_three_component_tests.phenotype, inplace=True)

    coords = trianglize(effs.values)
    fig, ax = pylab.subplots()
    c = phecode_three_component_tests.phenotype.map(phecode_info.category.map(color_by_phecode_cat))
    s = -numpy.log10(phecode_three_component_tests.overall_p.astype(float)) + 8
    ax.scatter(*coords.T, c=c, s=s)
    plot_triangular_ses(ax, effs, ses, c)
    triangle_frame(ax)
    legend_from_colormap(fig, color_by_phecode_cat)

    # Plot the sex differences
    fig, axes = pylab.subplots(ncols=3, figsize=(14,5))
    vars = ['circ', 'physical', 'sleep']
    phecode = phecode_three_component_tests_by_sex.phenotype
    #colormap = {cat:color for cat, color in
    #                    zip(phecode.unique(),
    #                        [pylab.get_cmap("Set3")(i) for i in range(20)])}
    c = phecode.map(phecode_info.set_index('phenotype').category).map(color_by_phecode_cat)
    for ax, var in zip(axes, vars):
        males = phecode_three_component_tests_by_sex[f"male_{var}_eff"]
        females = phecode_three_component_tests_by_sex[f"female_{var}_eff"]
        ax.scatter(
            males, females,
            c=c
        )
        m = min(numpy.min(males), numpy.min(females))
        M = max(numpy.max(males), numpy.max(females))
        ax.plot([m,M], [m,M], color='k')
        ax.set_title(var)
        ax.set_aspect('equal')
        ax.set_xlabel("Male effect size")
    axes[0].set_ylabel("Female effect size")
    legend_from_colormap(fig, color_by_phecode_cat)
    fig.tight_layout()
    fig.subplots_adjust(right=0.8) # Make room for legend
    fig.savefig(OUTDIR+"three_components.phecodes.by_sex.png")

    # Plot the sex differences in quantitative
    fig, axes = pylab.subplots(ncols=3, figsize=(14,5))
    vars = ['circ', 'physical', 'sleep']
    phenotype = quantitative_three_component_tests_by_sex.phenotype
    c = phenotype.map(quantitative_variable_descriptions['Functional Categories']).map(color_by_quantitative_function)
    for ax, var in zip(axes, vars):
        males = quantitative_three_component_tests_by_sex[f"male_{var}_eff"]
        females = quantitative_three_component_tests_by_sex[f"female_{var}_eff"]
        ax.scatter(
            males, females,
            c=c
        )
        m = min(numpy.min(males), numpy.min(females))
        M = max(numpy.max(males), numpy.max(females))
        ax.plot([m,M], [m,M], color='k')
        ax.set_title(var)
        ax.set_xlabel("Male effect size")
        ax.set_aspect('equal')
    axes[0].set_ylabel("Female effect size")
    legend_from_colormap(fig, color_by_quantitative_function)
    fig.tight_layout()
    fig.subplots_adjust(right=0.8) # Make room for legend
    fig.savefig(OUTDIR+"three_components.quantitative.by_sex.png")

    # Plot the age differences
    fig, axes = pylab.subplots(ncols=3, figsize=(14,5))
    vars = ['circ', 'physical', 'sleep']
    phecode = phecode_three_component_tests_by_age.phenotype
    c = phecode.map(phecode_info.set_index('phenotype').category).map(color_by_phecode_cat)
    for ax, var in zip(axes, vars):
        age55 = phecode_three_component_tests_by_age[f"age55_{var}_eff"]
        age70 = phecode_three_component_tests_by_age[f"age70_{var}_eff"]
        ax.scatter(
            age55, age70,
            c=c
        )
        m = min(numpy.min(age55), numpy.min(age70))
        M = max(numpy.max(age55), numpy.max(age70))
        ax.plot([m,M], [m,M], color='k')
        ax.set_title(var)
        ax.set_aspect('equal')
        ax.set_xlabel("Age 55 effect size")
    axes[0].set_ylabel("Age 70 effect size")
    legend_from_colormap(fig, color_by_phecode_cat)
    fig.tight_layout()
    fig.subplots_adjust(right=0.8) # Make room for legend
    fig.savefig(OUTDIR+"three_components.phecodes.by_age.png")

    # Plot the age differences in quantitative
    fig, axes = pylab.subplots(ncols=3, figsize=(14,5))
    vars = ['circ', 'physical', 'sleep']
    phenotype = quantitative_three_component_tests_by_age.phenotype
    c = phenotype.map(quantitative_variable_descriptions['Functional Categories']).map(color_by_quantitative_function)
    for ax, var in zip(axes, vars):
        age55 = quantitative_three_component_tests_by_age[f"age55_{var}_eff"]
        age70 = quantitative_three_component_tests_by_age[f"age70_{var}_eff"]
        ax.scatter(
            age55, age70,
            c=c
        )
        m = min(numpy.min(age55), numpy.min(age70))
        M = max(numpy.max(age55), numpy.max(age70))
        ax.plot([m,M], [m,M], color='k')
        ax.set_title(var)
        ax.set_aspect('equal')
        ax.set_xlabel("Age 55 effect size")
    axes[0].set_ylabel("Age 70 effect size")
    legend_from_colormap(fig, color_by_quantitative_function)
    fig.tight_layout()
    fig.subplots_adjust(right=0.8) # Make room for legend
    fig.savefig(OUTDIR+"three_components.quantitative.by_age.png")


def objective_subjective_plots():
    ### Comparisons of self-reported versus objectively derived variables
    # Some self-reported variables are closely matched by objectively derived variables
    # so which one is best associated with phecodes?

    # self_report_sleep_duration versus main_sleep_duration_mean or total_sleep_mean
    # Sleep durations have a fairly U-shaped curve so we use the abs_dev versions here
    # and the total_sleep_mean variable includes napping time as well as sleep time
    # so we don't use that here since napping is investigated in the self_report_nap_during_day variable
    variable_pairs = { "sleep duration": ("sleep_duration", "main_sleep_duration_mean"),
                    "sleep duration abs. dev.": ("sleep_duration_abs_dev", "main_sleep_duration_mean_abs_dev"),
                    "phase": ("morning_evening_person", "phase"),
                    #("morning_evening_person", "phase_abs_dev"),
                    "napping": ("nap_during_day", "other_sleep_mean"),
                    "sleeplessness": ("sleeplessness", "WASO_mean")}
    #fig, (ax1, ax2) = pylab.subplots(ncols=2, sharey=True)
    fig, ax2 = pylab.subplots(figsize=(4,4))
    for i, (name, (self_report, activity_var)) in enumerate(variable_pairs.items()):
        if self_report == "sleep_duration_abs_dev":
            self_report_var = "self_report_sleep_duration_abs_dev"
        else:
            self_report_var = "self_report_" + self_report_circadian_variables[self_report]['name']
        print(f"Comparing: {self_report} to {activity_var}")
        self_report_survival = survival_tests.query(f"activity_var == '{self_report_var}'").iloc[0]
        self_report_objective = survival_tests.query(f"activity_var == '{activity_var}'").iloc[0]
        # Downsample the objective ot have the same population as the self_report
        downsampled_data = data[~data[self_report_var].isna()]
        downsampled_uncensored = data.uncensored[~data[self_report_var].isna()]
        downsampled_entry_age = data.entry_age[~data[self_report_var].isna()]
        covariate_formula = "BMI + smoking_ever"
        survival_test = smf.phreg(formula = f"age_at_death_censored ~ {activity_var} + sex + {covariate_formula}",
                                        data=downsampled_data,
                                        status=downsampled_uncensored,
                                        entry=downsampled_entry_age,
                                        ).fit()
        pvalues = pandas.Series(survival_test.pvalues, index=survival_test.model.exog_names)
        params = pandas.Series(survival_test.params, index=survival_test.model.exog_names)
        if self_report in data and data[self_report].dtype.name == "category":
            cat_counts = data[self_report].value_counts()
            most_common = cat_counts.idxmax()
            dropped_cats = cat_counts.index[cat_counts < 400].to_list() # Want at least this many in the category
            cats = '"' + '", "'.join(cat_counts.index[cat_counts >= 400].to_list()) + '"'
            subset = ~data[self_report].isin(dropped_cats)
            self_report_survival_test = smf.phreg(
                formula = f'age_at_death_censored ~ C({self_report}, levels=[{cats}]) + sex + {covariate_formula}',
                #formula = f'age_at_death_censored ~ {self_report} + sex + {covariate_formula}',
                data=data[subset],
                status=data.uncensored[subset],
                entry=data.entry_age[subset],
            ).fit()
            self_report_pvalue = self_report_survival_test.f_test(
                [[1 if j == i else 0 for j in range(len(self_report_survival_test.model.exog_names))]
                    for i, var in enumerate(self_report_survival_test.model.exog_names)
                    if self_report in var]
            ).pvalue
            self_report_pvalue = self_report_pvalue.flatten()[0]
        else:
            # Non-categorical values don't have an alternative, use the same as before
            self_report_pvalue = self_report_survival.p
        df = pandas.DataFrame.from_dict({
            "self_report": {
                "survival_p": self_report_survival.p,
                "survival_logHR": self_report_survival['standardized log Hazard Ratio'],
            },
            "objective": {
                "survival_p": self_report_objective.p,
                "survival_logHR": self_report_objective['standardized log Hazard Ratio'],
            },
            "downsampled_objective": {
                "survival_p": pvalues[activity_var],
                "survival_logHR": params[activity_var] * downsampled_data[activity_var].std(),
            },
            "self_report_cat": {
                "survival_p": self_report_pvalue,
            },
        })
        print(df.T)
        # Plot the points
        #ax1.scatter([i], [-numpy.log10(df.loc["survival_p", "self_report"])], c="k")
        #ax1.scatter([i], [-numpy.log10(df.loc["survival_p", "downsampled_objective"])], c="r")
        #ax2.scatter([i], [-numpy.log10(df.loc["survival_p", "self_report_cat"])], c="k")
        #ax2.scatter([i], [-numpy.log10(df.loc["survival_p", "objective"])], c="r")
        BAR_WIDTH = 0.6
        ax2.bar(2*i-BAR_WIDTH/2, -numpy.log10(df.loc['survival_p', 'self_report_cat']), color="k", width=BAR_WIDTH)
        ax2.bar(2*i+BAR_WIDTH/2, -numpy.log10(df.loc['survival_p', 'objective']), color="r", width=BAR_WIDTH)
    #ax1.set_ylabel("-log10 p-value")
    #ax1.set_xticks(range(len(variable_pairs)))
    #ax1.set_xticklabels([f"{self_report}\n{activity_var}" for self_report, activity_var in variable_pairs])
    #ax2.set_xticks(range(len(variable_pairs)))
    #ax2.set_xticklabels([f"{self_report}\n{activity_var}" for self_report, activity_var in variable_pairs])
    #ax1.xaxis.set_tick_params(rotation=90)
    ax2.set_xticks(2*numpy.arange(len(variable_pairs)))
    ax2.set_xticklabels([name for name in variable_pairs.keys()])
    ax2.set_ylabel("-log10 p-value")
    ax2.xaxis.set_tick_params(rotation=90)
    legend_elts = [matplotlib.lines.Line2D(
                            [0],[0],
                            marker="o", markerfacecolor=c, markersize=10,
                            label=l,
                            c=c, lw=0)
                        for c, l in zip(["k", "r"], ["Subjective", "Objective"])]
    ax2.legend(handles=legend_elts)
    fig.tight_layout()
    fig.savefig(OUTDIR+"subjective_objective_comparison.survival.png")


    # Compare the objective/subjective variables again on the phecodes
    fig, axes = pylab.subplots(ncols=len(variable_pairs), figsize=(15,7), sharex=True, sharey=True)
    Q_CUTOFF = 0.05
    for ax, (name, (self_report, activity_var)) in zip(axes.flatten(), variable_pairs.items()):
        if self_report is not "sleep_duration_abs_dev":
            self_report_var = "self_report_" + self_report_circadian_variables[self_report]['name']
        else:
            self_report_var = "self_report_sleep_duration_abs_dev"
        either_significant = phecode_tests[
            ((phecode_tests.activity_var == self_report_var) |
            (phecode_tests.activity_var == activity_var)) &
            (phecode_tests.q < Q_CUTOFF)].phecode.unique()
        downsampled_data = data[~data[self_report_var].isna()]
        objective_tests_list = []
        for phecode in either_significant:
            covariate_formula = ' + '.join(c for c in covariates if c != 'sex')
            N = data[phecode].sum()
            fit = OLS(f"{activity_var} ~ Q({phecode}) + sex * ({covariate_formula})",
                        data=downsampled_data)
            p = fit.pvalues[f"Q({phecode})"]
            coeff = fit.params[f"Q({phecode})"]
            std_effect = coeff / data[activity_var].std()
            N_cases = data.loc[~data[activity_var].isna(), phecode].sum()
            objective_tests_list.append({"phecode": phecode,
                                    "activity_var": activity_var,
                                    "p": p,
                                    "coeff": coeff,
                                    "std_effect": std_effect,
                                    "N_cases": N_cases,
                                })
        objective_tests = pandas.DataFrame(objective_tests_list)
        subjective_ps = objective_tests.phecode.map(phecode_tests[(phecode_tests.activity_var == self_report_var)].set_index("phecode").p)
        colors = objective_tests.phecode.map(phecode_info.category).map(color_by_phecode_cat)
        ax.scatter(-numpy.log10(objective_tests.p), -numpy.log10(subjective_ps), c=colors)
        ax.set_aspect("equal")
        ax.set_xlabel(f"-log10 p objective variable")
        if ax == axes.flatten()[0]:
            ax.set_ylabel(f"-log10 p subjective variable")
        ax.set_title(name)
    for ax in axes.flatten():
        bound = max(abs(numpy.min([ax.get_xlim(), ax.get_ylim()])),
                    numpy.max([ax.get_xlim(), ax.get_ylim()]))
        diag = numpy.array([0, bound])
        ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
        ax.set_xlim(0,20)
        ax.set_ylim(0,20)
    fig.tight_layout()
    legend_from_colormap(fig, color_by_phecode_cat, ncol=2)
    fig.savefig(OUTDIR+"objective_subjective_comparison.png")


def by_date_plots():
    fig, ax, hypertension_by_date = phewas_plots.plot_by_diagnosis_date(data, ["I10"], 401, "Hypertension", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.hypertension.png")

    fig, ax, diabetes_by_date = phewas_plots.plot_by_diagnosis_date(data, ["E09", "E10", "E11", "E13"], 250, "Diabetes", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.diabetes.png")

    fig, ax, IHD_by_date = phewas_plots.plot_by_diagnosis_date(data, ["I20", "I21", "I22", "I23", "I25"], 411, "Ischemic Heart Disease", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.IHD.png")

    fig, ax, pneumonia_by_date = phewas_plots.plot_by_diagnosis_date(data, ["J09", "J10", "J11", "J12", "J13", "J14", "J15", "J16", "J17", "J18"], 480, "Pneumonia", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.pneumonia.png")

    fig, ax, mood_disorders_by_date = phewas_plots.plot_by_diagnosis_date(data, ["F30", "F31", "F32", "F33", "F34", "F39"], 296, "Mood Disorders", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.mood_disorders.png")

    fig, ax, anxiety_by_date = phewas_plots.plot_by_diagnosis_date(data, ["F40", "F41"], 300, "Anxiety Disorders", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.anxiety_disorders.png")

    fig, ax, renal_failure_by_date = phewas_plots.plot_by_diagnosis_date(data, ["N17", "N18", "N19", "Y60", "Y84", "Z49"], 585, "Renal Failure", icd10_entries, phecode_tests)
    fig.savefig(OUTDIR+"by_date.renal_failure.png")

def generate_results_table():
    #### Combine all tables into the summary with the header file
    print(f"Writing out the complete results table to {OUTDIR+'results.xlsx'}")
    import openpyxl
    workbook = openpyxl.load_workbook("../table_header.xlsx")
    descriptions = workbook['Variables']
    start_column = len(list(descriptions.tables.values())[0].tableColumns) + 1
    var_stats = data[activity_variable_descriptions.index].describe(percentiles=[0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]).T
    for i, col in enumerate(var_stats.columns):
        descriptions.cell(1, start_column + i, col)
        for j, value in enumerate(var_stats[col].values):
            descriptions.cell(j+2, start_column + i, value)
    workbook.save(OUTDIR+"results.xlsx")
    with pandas.ExcelWriter(OUTDIR+"results.xlsx", mode="a") as writer:
        survival_tests.sort_values(by="p").to_excel(writer, sheet_name="Survival Associations", index=False)
        phecode_tests_raw.sort_values(by="p").to_excel(writer, sheet_name="Phecode Associations", index=False)
        quantitative_tests_raw.sort_values(by="p").to_excel(writer, sheet_name="Quantitative Associations", index=False)
        phecode_tests_by_sex.sort_values(by="p_diff").to_excel(writer, sheet_name="Sex-specific Associations", index=False)
        quantitative_sex_tests.sort_values(by="sex_difference_p").to_excel(writer, sheet_name="Quantitative Sex Differences", index=False)
        age_tests.sort_values(by="p").to_excel(writer, sheet_name="Age-dependence", index=False)
        quantitative_age_tests.sort_values(by="age_difference_p").to_excel(writer, sheet_name="Quantitative Age Differences", index=False)
        phecode_details.to_excel(writer, sheet_name="PheCODEs")

def temperature_trace_plots(N_IDS=500):
    ids = day_plots.get_ids_of_traces_available()

    def temp_to_C(temp):
        return temp
        #return (500*temp - 2550)/256 # Convert "temp" units to Celsius
        #TODO: should we be doing this at another step??

    ## Overall temperature cycle
    _, _, ax = day_plots.plot_average_trace(numpy.random.choice(ids, size=N_IDS, replace=False),
                    var="temp",
                    transform = temp_to_C,
                    normalize_mean = True)
    ax.set_ylabel("Temperature (C)")

    ## By categories
    def temp_trace_by_cat(cats, colors=None, show_variance=True, show_confidence_intervals=False):
        temp_mean = temp_to_C(full_activity.temp_mean_mean.mean())
        fig, ax = pylab.subplots()
        for cat in cats.cat.categories:
            selected_ids = data[(cats == cat) & (data.index.isin(ids))].index
            selected_ids = numpy.random.choice(selected_ids, size=min(len(selected_ids), N_IDS), replace=False)
            day_plots.plot_average_trace(selected_ids,
                        var="temp",
                        transform = temp_to_C,
                        normalize_mean = True,
                        set_mean = 0,
                        ax=ax,
                        color=colors[cat] if colors is not None else None,
                        label=cat,
                        show_variance=show_variance,
                        show_confidence_intervals=show_confidence_intervals)
        ax.set_ylabel("Temperature (C)")
        fig.legend()
        fig.tight_layout()
        return fig

    def case_control(phecode):
        cats = data[phecode].astype("category").cat.rename_categories({0:"Control", 1:"Case"})
        fig = temp_trace_by_cat(cats,
                                 colors = {"Case": "orange", "Control": "teal"})
        fig.gca().set_title(phecode_info.loc[phecode].phenotype)
        fig.tight_layout()
        return fig

    fig = case_control(250)
    fig.savefig(OUTDIR+"temperature.diabetes.png")
    fig = case_control(401)
    fig.savefig(OUTDIR+"temperature.hypertension.png")
    fig = case_control(496)
    fig.savefig(OUTDIR+"temperature.chronic_airway_obstruction.png")
    fig = case_control(443)
    fig.savefig(OUTDIR+"temperature.peripheral_vascular_disease.png")
    fig = case_control(495)
    fig.savefig(OUTDIR+"temperature.asthma.png")
    fig = case_control(480)
    fig.savefig(OUTDIR+"temperature.pneumonia.png")

    morning_evening = data.morning_evening_person.cat.remove_categories(["Prefer not to answer", "Do not know"])
    fig = temp_trace_by_cat(morning_evening, show_variance=False)
    fig.gca().set_title("Chronotype")
    fig.tight_layout()
    fig.savefig(OUTDIR+"temperature.chronotype.png")

    age_cutoffs = numpy.arange(45,75,5) # every 5 years from 40 to 75
    age_categories = pandas.cut(data.age_at_actigraphy, age_cutoffs)
    fig = temp_trace_by_cat(age_categories, show_variance=False)
    fig.gca().set_title("Age")
    fig.tight_layout()
    fig.savefig(OUTDIR+"temperature.age.png")

    fig = temp_trace_by_cat(data.sex, colors={'Male': '#1f77b4', 'Female': '#ff7f0e'})
    fig.gca().set_title("Sex")
    fig.tight_layout()
    fig.savefig(OUTDIR+"temperature.sex.png")

    napping = data.nap_during_day.cat.remove_categories(["Prefer not to answer"])
    fig = temp_trace_by_cat(napping, show_variance=False)
    fig.gca().set_title("Nap During Day")
    fig.tight_layout()
    fig.savefig(OUTDIR+"temperature.nap.png")

    ## Asthma
    obese = data.BMI > 30
    normal = (data.BMI < 25) & (data.BMI > 18.5)
    cats = data[495].map({1: "Asthma", 0: "Control"}) + obese.map({True: " Obese", False: " Normal"})
    cats[(~normal) & (~obese)] = float("NaN") # Remove 'overweight-but-not-obese' middle category
    cats = cats.astype("category")
    fig = temp_trace_by_cat(cats, show_variance=False, show_confidence_intervals=True)
    fig.gca().set_title("Asthma by Weight")
    fig.tight_layout()
    fig.savefig(OUTDIR+"temperature.asthma.by_bmi.png")

def chronotype_plots():
    # Chronotype by sex+age plot
    age_cutoffs = numpy.arange(40,80,5) # every 5 years from 40 to 75
    responses = ["Definitely a 'morning' person", "More a 'morning' than 'evening' person", 'Do not know', "More an 'evening' than a 'morning' person", "Definitely an 'evening' person", 'Prefer not to answer']
    cmap = pylab.get_cmap("viridis")
    colors = [cmap(0), cmap(0.25), cmap(0.5), cmap(0.75), cmap(1.0), "k"]
    age_at_assessment = (pandas.to_datetime(data.blood_sample_time_collected_V0) - data.birth_year_dt) / pandas.to_timedelta("1Y")
    fig, axes = pylab.subplots(nrows=2, figsize=(8,6))
    for ax, sex in zip(axes.flatten(), ['Male', 'Female']):
        d = data[data.sex == sex]
        age_buckets = pandas.cut(age_at_assessment[data.sex == sex], age_cutoffs, right=True)
        for i, (bucket, age_data) in enumerate(d.groupby(age_buckets).morning_evening_person):
            counts = age_data.value_counts()
            counts /= counts.sum()
            base = 0
            for response, color in zip(responses, colors):
                ax.bar(i, counts[response], bottom=base, color = color)
                base += counts[response]
        ax.set_xticks(numpy.arange(len(age_buckets.cat.categories)))
        ax.set_xticklabels([f"{c.left}-{c.right}" for c in age_buckets.cat.categories])
    util.legend_from_colormap(fig, dict(list(zip(responses, colors))[::-1]))
    fig.subplots_adjust(right=0.65)

def activity_var_comparisons():

    def facet_chart(base_vars, other_vars, other_label="other", kind="pvalue"):
        fig, axes = pylab.subplots(ncols=4, nrows=math.ceil(len(base_vars)/4), figsize=(10,8), sharex=True, sharey=True, squeeze=False)
        for base_var, other_var, ax in zip(base_vars, other_vars, axes.flatten()):
            pt = phecode_tests[phecode_tests.activity_var == base_var].join(
                phecode_tests[phecode_tests.activity_var == other_var].set_index("phecode"),
                on="phecode",
                lsuffix="_base",
                rsuffix="_other")
            if kind == "pvalue":
                ax.scatter(-numpy.log10(pt.p_base), -numpy.log10(pt.p_other), alpha=0.3)
                if ax in axes[-1,:]:
                    ax.set_xlabel("-log10 p-value\nbase variable")
                if ax in axes[:,0]:
                    ax.set_ylabel(f"-log10 p-value\n{other_label} variable")
            else:
                ax.scatter(pt.std_effect_base, pt.std_effect_other, alpha=0.3)
                if ax in axes[-1,:]:
                    ax.set_xlabel("Std Effect Size\nbase variable")
                if ax in axes[:,0]:
                    ax.set_ylabel(f"Std Effect Size\n{other_label} variable")
            ax.set_title(base_var)
            #ax.set_aspect('equal')
        for base_var, ax in zip(base_vars, axes.flatten()):
            diag = numpy.array([numpy.min([ax.get_xlim(), ax.get_ylim()]), numpy.max([ax.get_xlim(), ax.get_ylim()])])
            ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
            ax.set_xlim(diag[0], diag[1])
            ax.set_ylim(diag[0], diag[1])
        fig.tight_layout()
        return fig

    # Comparison plots for M10 versus whole day
    base_vars = [v for v in activity_variables if v + "_M10" in activity_variables]
    other_vars = [v+"_M10" for v in base_vars]
    fig = facet_chart(base_vars, other_vars, other_label="M10")
    fig.savefig(OUTDIR+"activity_var_types.M10_vars.pvalues.png")
    fig = facet_chart(base_vars, other_vars, other_label="M10", kind="std_effect")
    fig.savefig(OUTDIR+"activity_var_types.M10_vars.std_effect.png")

    # Comparison plots for L5 versus whole day
    base_vars = [v for v in activity_variables if v + "_L5" in activity_variables]
    other_vars = [v+"_L5" for v in base_vars]
    fig = facet_chart(base_vars, other_vars, other_label="L5")
    fig.savefig(OUTDIR+"activity_var_types.L5_vars.pvalues.png")
    fig = facet_chart(base_vars, other_vars, other_label="L5", kind="std_effect")
    fig.savefig(OUTDIR+"activity_var_types.L5_vars.std_effect.png")

    # Comparison plots of overall versus hourly_SD
    other_vars = [v for v in activity_variables if "hourly_SD" in v]
    base_vars = [re.sub("hourly_SD", "overall", v) for v in other_vars]
    fig = facet_chart(base_vars, other_vars, other_label="hourly_SD")
    fig.savefig(OUTDIR+"activity_var_types.hourly_SD_vars.pvalues.png")
    fig = facet_chart(base_vars, other_vars, other_label="hourly_SD", kind="std_effect")
    fig.savefig(OUTDIR+"activity_var_types.hourly_SD_vars.std_effect.png")

    # Comparison plots of overall versus between_day_SD
    other_vars = [v for v in activity_variables if "between_day_SD" in v]
    base_vars = [re.sub("between_day_SD", "overall", v) for v in other_vars]
    fig = facet_chart(base_vars, other_vars, other_label="hourly_SD")
    #fig.savefig(OUTDIR+"activity_var_types.hourly_SD_vars.pvalues.png")
    fig = facet_chart(base_vars, other_vars, other_label="hourly_SD", kind="std_effect")
    #fig.savefig(OUTDIR+"activity_var_types.hourly_SD_vars.std_effect.png")


    # Stability comparison plots of variables
    def stability_comparison(base_vars, other_vars, other_label):
        fig, ax = pylab.subplots(figsize=(8,8))
        ax.scatter(
            activity_variance.loc[base_vars, "corrected_intra_personal_normalized"],
            activity_variance.loc[other_vars, "corrected_intra_personal_normalized"],
        )
        diag = numpy.array([0, numpy.max([ax.get_xlim(), ax.get_ylim()])])
        ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
        ax.set_xlim(diag[0], diag[1])
        ax.set_ylim(diag[0], diag[1])
        ax.axhline(1, linestyle="--", c='k', zorder=-1)
        ax.axvline(1, linestyle="--", c='k', zorder=-1)
        ax.set_xlabel(f"Variance ratio\nbase variable")
        ax.set_ylabel(f"Variance ratio\n{other_label} variable")
        return fig

    base_vars = [v for v in activity_variance.index if v + "_M10" in activity_variance.index]
    other_vars = [v+"_M10" for v in base_vars]
    fig = stability_comparison(base_vars, other_vars, other_label="M10")
    fig.savefig(OUTDIR+"activity_var_types.M10.stability.png")

    base_vars = [v for v in activity_variance.index if v + "_L5" in activity_variance.index]
    other_vars = [v+"_L5" for v in base_vars]
    fig = stability_comparison(base_vars, other_vars, other_label="L5")
    fig.savefig(OUTDIR+"activity_var_types.L5.stability.png")

    other_vars = [v for v in activity_variance.index if "hourly_SD" in v]
    base_vars = [re.sub("hourly_SD", "overall", v) for v in other_vars]
    fig = stability_comparison(base_vars, other_vars, other_label="hourly SD")
    fig.savefig(OUTDIR+"activity_var_types.hourly_SD.stability.png")

    other_vars = [v for v in activity_variance.index if "between_day_SD" in v]
    base_vars = [re.sub("between_day_SD", "overall", v) for v in other_vars]
    fig = stability_comparison(base_vars, other_vars, other_label="between-day SD")
    fig.savefig(OUTDIR+"activity_var_types.between_day_SD.stability.png")


def plot_case_control_by_age(phecode, activity_variable):
    age_cutoffs = numpy.arange(45,81,5) # every 5 years from 40 to 75
    fig, ax = pylab.subplots(figsize=(8,6))
    age_buckets = pandas.cut(data.age_at_actigraphy, age_cutoffs, right=True)
    colors = {"Control": "k", "Case": "r"}
    for status in ['Control', 'Case']:
        in_status = data[phecode] == (1 if status == 'Case' else 0)
        means = data[in_status].groupby(age_buckets[in_status])[activity_variable].mean()
        sems = data[in_status].groupby(age_buckets[in_status])[activity_variable].sem()
        ax.plot(numpy.arange(len(means)), means,
            c=colors[status])
        ax.errorbar(numpy.arange(len(means)), means,
            yerr=sems,
            c=colors[status])
    ax.set_xticks(numpy.arange(len(age_buckets.cat.categories)))
    ax.set_xticklabels([f"{c.left}-{c.right}" for c in age_buckets.cat.categories])
    phenotype_name = phecode_info.loc[phecode].phenotype
    ax.set_ylabel(activity_variable)
    ax.set_xlabel("Age")
    util.legend_from_colormap(fig, colors)
    ax.set_title(f"{phenotype_name} - {activity_variable}")
    return fig


def plot_quantitative_by_age(phenotype, activity_var, df=None, name=None, categorical=False, ax=None):
    if df is None:
        df = data
    df = df.copy()
    age_cutoffs = numpy.arange(45,81,5) # every 5 years from 45 to 80
    df['age_bucket'] = pandas.cut(df.age_at_actigraphy, age_cutoffs, right=True)
    if  categorical:
        group_by = df[activity_var].astype("category")
    else:
        activity_var_quintiles = pandas.qcut(df[activity_var], 5)
        group_by = activity_var_quintiles
    cmap = pylab.get_cmap("viridis")
    colors = {group: cmap(t) for group, t in zip(group_by.cat.categories, numpy.linspace(0,1,len(group_by.cat.categories)))}
    if ax is None:
        fig, ax = pylab.subplots(figsize=(8,6))
    else:
        fig = None
    for (group_name, group_data) in df.groupby(group_by):
        means = group_data.groupby("age_bucket")[phenotype].mean()
        sems = group_data.groupby("age_bucket")[phenotype].sem()
        ax.plot(numpy.arange(len(means)), means,
            c=colors[group_name])
        ax.errorbar(numpy.arange(len(means)), means,
            yerr=sems,
            c=colors[group_name])
    ax.set_xticks(numpy.arange(len(df.age_bucket.cat.categories)))
    phenotype_name = name if name is not None else quantitative_variable_descriptions.loc[phenotype].Name
    ax.set_ylabel(phenotype_name)
    ax.set_xlabel("Age at Actigraphy")
    ax.set_xticklabels([f"{c.left} - {c.right}" for c in df.age_bucket.cat.categories])
    if categorical:
        util.legend_from_colormap(ax, colors)
    else:
        util.legend_from_colormap(ax,
            {label:color for label, color in zip(["0-20%", "20-40%", "40-60%", "60-80%", "80-100%"], colors.values())},
            title=f"{activity_var} Quintile")
    ax.set_title(f"{phenotype_name} by {activity_var}")
    return fig, ax

def age_by_categorical_question(phecode, question, responses):
    age_cutoffs = numpy.arange(45,85,5) # every 5 years from 40 to 80
    cmap = pylab.get_cmap("viridis")
    colors = [cmap(t) for t in numpy.linspace(0,1,len(responses))]
    fig, ax = pylab.subplots(figsize=(8,6))
    for status in ['Control', 'Case']:
        in_status = data[phecode] == (1 if status == 'Case' else 0)
        d = data[in_status]
        age_buckets = pandas.cut(data.age_at_actigraphy[in_status], age_cutoffs, right=True)
        for i, (bucket, age_data) in enumerate(d.groupby(age_buckets)[question]):
            counts = age_data.value_counts()
            counts /= counts[responses].sum()
            base = 0
            for response, color in zip(responses, colors):
                ax.bar(i + (-0.2 if status == "Control" else 0.2),
                        counts[response],
                        bottom=base,
                        color = color,
                        alpha = 0.5 if status == "Control" else 1,
                        width=0.35)
                base += counts[response]
        xticks = numpy.arange(len(age_buckets.cat.categories))
        ax.set_xticks(xticks)
        ax.set_xticklabels([f"{c.left}-{c.right}" for c in age_buckets.cat.categories])
        ax.set_xlabel("Age")
        ax.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter(xmax=1))
    legend =  dict(list(zip(responses, colors))[::-1])
    legend.update(dict(Control= [0.5,0.5,0.5,0.5], Case=[0.5,0.5,0.5,1]))
    util.legend_from_colormap(fig, legend)
    fig.subplots_adjust(right=0.9)
    return fig

def age_interaction_plots():
    lipoprotein_vars = ["cholesterol", "hdl_cholesterol", "ldl_direct", "triglycerides", "apolipoprotein_A", "apolipoprotein_B", "lipoprotein_A"]
    for var in lipoprotein_vars:
        fig = plot_quantitative_by_age(var, "acceleration_overall")
        fig.savefig(OUTDIR+f"by_age.{var}.vs.acceleration_overall.png")
        fig = plot_quantitative_by_age(var, "IPAQ_activity_group", categorical=True)
        fig.savefig(OUTDIR+f"by_age_.{var}.vs.IPAQ_activity_group.png")
    df = data.copy()
    df['cholesterol_ratio'] = data.cholesterol / data.hdl_cholesterol
    fig = plot_quantitative_by_age("cholesterol_ratio", "acceleration_overall", df=df, name="Cholesterol Ratio")
    fig.savefig(OUTDIR+"by_age.cholesterol_ratio.vs.acceleration_overall.png")

    # By sex
    no_cholesterol_meds = ((data.medication_cholesterol_bp_diabetes_or_exog_hormones_Cholesterol_lowering_medication == 0) # For males
                            | (data.medication_cholesterol_bp_diabetes_Cholesterol_lowering_medication == 0)) # For females
    for meds_status in ['no_meds', 'with_meds', 'all']:
        if meds_status == 'no_meds':
            df = data[no_cholesterol_meds]
        elif meds_status == 'with_meds':
            df = data[~no_cholesterol_meds]
        else:
            df = data
        for var in lipoprotein_vars:
            fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(10,10))
            for i, sex in enumerate(["Male", "Female"]):
                _, ax = plot_quantitative_by_age(var, "acceleration_overall", df=df[df.sex==sex], ax=axes[i, 0])
                ax.set_title(f"{sex} - by acceleration_overall")
                _, ax = plot_quantitative_by_age(var, "IPAQ_activity_group", df=df[df.sex==sex], ax=axes[i, 1], categorical=True)
                ax.set_title(f"{sex} - by IPAQ_activity_group")
            fig.savefig(OUTDIR+f"by_age.by_sex.{var}.combined.{meds_status}.png")


    renal_function_vars = quantitative_variable_descriptions.index[quantitative_variable_descriptions['Functional Categories'] == 'Renal Function']
    for var in renal_function_vars:
        fig = plot_quantitative_by_age(var, "acceleration_overall")
        fig.savefig(OUTDIR+f"by_age.{var}.vs.acceleration_overall.png")

    # Plot multiple response questions by age and case-control
    responses = ['Not at all easy', 'Very easy']
    fig = age_by_categorical_question(296, 'getting_up_in_morning', responses)
    fig.savefig(OUTDIR+"by_age.mood_disorder.vs.getting_up_in_morning.png")
    responses = ['Not at all easy', 'Not very easy', 'Fairly easy', 'Very easy']
    fig = age_by_categorical_question(296, 'getting_up_in_morning', responses)
    fig.savefig(OUTDIR+"by_age.mood_disorder.vs.getting_up_in_morning.full.png")

    # Plot case-control plots
    fig = plot_case_control_by_age(401, "walking_overall_M10")
    fig.savefig(OUTDIR+"by_age.hypertension.vs.walking_overall_M10.png")
    fig = plot_case_control_by_age(401, "acceleration_overall")
    fig.savefig(OUTDIR+"by_age.hypertension.vs.acceleration_overall.png")
    fig = plot_case_control_by_age(296, "main_sleep_duration_mean")
    fig.savefig(OUTDIR+"by_age.mood_disorders.vs.main_sleep_duration_mean.png")
    fig = plot_case_control_by_age(296, "main_sleep_offset_mean")
    fig.savefig(OUTDIR+"by_age.mood_disorders.vs.main_sleep_offset_mean.png")
    fig = plot_case_control_by_age(300, "main_sleep_offset_mean")
    fig.savefig(OUTDIR+"by_age.anxiety_disorders.vs.main_sleep_offset_mean.png")


def temperature_calibration_plots():
    # Plot the (mis)-calibration of the tempreature variables
    device = activity_summary.loc[data.index, 'file-deviceID']
    random_device = pandas.Series(device.sample(frac=1.0).values, index=device.index)
    temp_mean = full_activity[full_activity.run == 0].set_index("id").loc[data.index, 'temp_mean_mean']
    temp_RA = data['temp_RA']

    fig, axes = pylab.subplots(figsize=(8,5), ncols=2)
    for (name, measure), ax in zip({"mean": temp_mean, "RA": temp_RA}.items(), axes):
        device_mean = measure.groupby(device).mean()
        random_mean = measure.groupby(random_device).mean()
        m = device_mean.quantile(0.01)
        M = device_mean.quantile(0.99)
        bins = numpy.linspace(m,M,21)
        ax.hist(random_mean, bins=bins, color='k', alpha=0.5, label="Randomized" if ax == axes[0] else None)
        ax.hist(device_mean, bins=bins, color='r', alpha=0.5, label="True" if ax == axes[0] else None)
        ax.set_xlabel(f"Temperature {name}")
    fig.legend()
    fig.savefig(OUTDIR+"temperature_calibration.png")


def demographics_table():
    # Create a table of demographics of the population studied, compared to the overall UK Biobank
    demographics = {}
    ukbb_without_actigraphy = ukbb[ukbb.actigraphy_file.isna()]
    for name, d in zip(["Actigraphy", "Without Actigraphy"], [data, ukbb_without_actigraphy]):
        demographics[name] = {
            "Male": (d.sex == "Male").mean(),
            "Female": (d.sex == "Female").mean(),

            "White": (d.ethnicity.isin(["British", "Any other white background", "Irish", "White"])).mean(),
            "Nonwhite": (~d.ethnicity.isin(["British", "Any other white background", "Irish", "White"])).mean(),

            "Birth Year": (d.birth_year).mean(),
            "Birth Year (SD)": (d.birth_year).std(),

            "BMI": (d.BMI).mean(),
            "BMI (SD)": (d.BMI).std(),
        }
    demographics = pandas.DataFrame(demographics)
    demographics.to_csv(OUTDIR+"demographics.txt", sep="\t")
    return

def associations_by_meds():
    med_list = ["metformin", "aspirin", "simvastatin"]
    med_status = {med: data.index.isin(medications.ID[medications.medication ==  med]) for med in med_list}
    med_status['combined'] = numpy.any([status for status in med_status.values()], axis=0)
    activity_var = "acceleration_overall"
    phenotype = "cholesterol"
    phenotype_std = data[phenotype].std()
    covariate_formula = ' + '.join(c for c in covariates if c != 'sex')
    d = data.copy()

    results = []
    for med, status in med_status.items():
        d['med_status'] = status
        N = (~d[[activity_var, phenotype]].isna().any(axis=1)).sum()
        fit = OLS(f"{phenotype} ~ {activity_var} + med_status + sex * ({covariate_formula})",
                        data=d)
        p = fit.pvalues[activity_var]
        medication_p = fit.pvalues["med_status[T.True]"]
        coeff = fit.params[activity_var]
        activity_var_std = data[activity_var].std()
        std_effect = coeff * activity_var_std / phenotype_std
        results.append({
                        "phenotype": phenotype,
                        "activity_var": activity_var,
                        "medication": med,
                        "medication_p": medication_p,
                        "p": p,
                        "coeff": coeff,
                        "std_effect": std_effect,
                        "med_stat"
                        "N": N,
                    })
    results = pandas.DataFrame(results)

def quantitative_traits_with_medications():

    categories = {
        "cholesterol" : {
            "category": ["Lipoprotein Profile"],
            "med": ['medication_cholesterol_bp_diabetes_Cholesterol_lowering_medication', 'medication_cholesterol_bp_diabetes_or_exog_hormones_Cholesterol_lowering_medication'],
        },
        "BP": {
            "category": ["Cardiovascular Function"],
            "med": ['medication_cholesterol_bp_diabetes_Blood_pressure_medication', 'medication_cholesterol_bp_diabetes_or_exog_hormones_Blood_pressure_medication'],
        },
        "insulin": {
            "category": ["Glucose Metabolism"],
            "med": ['medication_cholesterol_bp_diabetes_Insulin', 'medication_cholesterol_bp_diabetes_or_exog_hormones_Insulin'],
        },
    }


    recompute = RECOMPUTE
    if not recompute:
        try:
            with_meds = pandas.read_csv(OUTDIR+"quantitative_traits.with_meds.txt", sep="\t", index_col=0)
            with_meds_by_sex = pandas.read_csv(OUTDIR+"quantitative_traits.by_sex.with_meds.txt", sep="\t", index_col=0)
            with_meds_by_age = pandas.read_csv(OUTDIR+"quantitative_traits.by_age.with_meds.txt", sep="\t", index_col=0)
        except FileNotFoundError:
            recompute = True
    if recompute:
        with_meds = []
        with_meds_by_age = []
        with_meds_by_sex = []
        for name, values in categories.items():
            quant_vars = quantitative_variable_descriptions[quantitative_variable_descriptions['Functional Categories'].isin(values['category'])].index
            med_covariate = data[values['med']].any(axis=1)
            d = data.copy()
            d['med_covariate'] = med_covariate
            results, age_results, sex_results = phewas_tests.quantitative_tests(
                d,
                quantitative_variables=quant_vars,
                activity_variables = activity_variables,
                activity_variable_descriptions=activity_variable_descriptions,
                quantitative_variable_descriptions=quantitative_variable_descriptions,
                extra_covariates = ['med_covariate'],
                OUTDIR = OUTDIR + f"{name}.",
                RECOMPUTE=RECOMPUTE,
                )
            with_meds.append(results)
            with_meds_by_age.append(age_results)
            with_meds_by_sex.append(sex_results)
        with_meds = pandas.concat(with_meds)
        with_meds['original_p'] = pandas.merge(with_meds[['phenotype', 'activity_var']], quantitative_tests, on=["phenotype", "activity_var"], how='inner').p
        with_meds['original_std_effect'] = pandas.merge(with_meds[['phenotype', 'activity_var']], quantitative_tests, on=["phenotype", "activity_var"], how='inner').std_effect
        with_meds.to_csv(OUTDIR+"quantitative_traits.with_meds.txt", sep="\t")

        with_meds_by_sex = pandas.concat(with_meds_by_sex)
        with_meds_by_sex['original_p'] = pandas.merge(with_meds_by_sex[['phenotype', 'activity_var']], quantitative_sex_tests, on=["phenotype", "activity_var"], how='inner').sex_difference_p
        with_meds_by_sex.to_csv(OUTDIR+"quantitative_traits.by_sex.with_meds.txt", sep="\t")

        with_meds_by_age = pandas.concat(with_meds_by_age)
        with_meds_by_age['original_p'] = pandas.merge(with_meds_by_age[['phenotype', 'activity_var']], quantitative_age_tests, on=["phenotype", "activity_var"], how='inner').age_difference_p
        with_meds_by_age.to_csv(OUTDIR+"quantitative_traits.by_age.with_meds.txt", sep="\t")

    # Display results in figures
    # with/without medication controls for quant variables
    fig, ax = pylab.subplots(figsize=(8,8))
    ax.scatter(-numpy.log10(with_meds.original_p), -numpy.log10(with_meds.p), c = with_meds['Functional Category'].map(color_by_quantitative_function))
    legend_from_colormap(ax, color_by_quantitative_function)
    ax.set_xlabel("-log10 p without medication controls")
    ax.set_ylabel("-log10 p with medication controls")
    # by effect size
    fig, ax = pylab.subplots(figsize=(8,8))
    ax.scatter((with_meds.original_std_effect), (with_meds.std_effect), c = with_meds['Functional Category'].map(color_by_quantitative_function))
    legend_from_colormap(ax, color_by_quantitative_function)
    ax.set_xlabel("std effect size without medication controls")
    ax.set_ylabel("std effect size with medication controls")

    # by sex
    fig, ax = pylab.subplots(figsize=(8,8))
    ax.scatter(-numpy.log10(with_meds_by_sex.original_p), -numpy.log10(with_meds_by_sex.sex_difference_p), c = with_meds_by_sex['Functional Category'].map(color_by_quantitative_function))
    legend_from_colormap(ax, color_by_quantitative_function)
    ax.set_xlabel("-log10 p without medication controls")
    ax.set_ylabel("-log10 p with medication controls")
    ax.set_title("Sex-differences")

    # by age
    fig, ax = pylab.subplots(figsize=(8,8))
    ax.scatter(-numpy.log10(with_meds_by_age.original_p), -numpy.log10(with_meds_by_age.age_difference_p), c = with_meds_by_age['Functional Category'].map(color_by_quantitative_function))
    legend_from_colormap(ax, color_by_quantitative_function)
    ax.set_xlabel("-log10 p without medication controls")
    ax.set_ylabel("-log10 p with medication controls")
    ax.set_title("Age-differences")

    ## Plot the by-sex plot
    qt = with_meds_by_sex.rename(columns={"std_male_effect": "std_male_coeff", "std_female_effect":"std_female_coeff"}).sample(frac=1)
    fig, ax = phewas_plots.sex_difference_plot(qt, color_by="Functional Category", cmap=color_by_quantitative_function, lim=0.25)
    fig.savefig(OUTDIR+"sex_differences.with_meds.quantitative.png")

    # Just lipoproteins
    # With medications
    qt_restricted = qt[qt['Functional Category'] == 'Lipoprotein Profile']
    color_by_lipoproteins = dict(zip(qt_restricted.phenotype.unique(),
                                    [pylab.get_cmap("tab20")(i) for i in range(20)]))
    fig, ax = phewas_plots.sex_difference_plot(qt_restricted, color_by="phenotype", cmap=color_by_lipoproteins, lim=0.25)
    fig.savefig(f"{OUTDIR}/sex_differences.with_meds.quantitative.lipoproteins.png")
    # without medications
    qt2_restricted = pandas.merge(
        qt_restricted[['phenotype', 'activity_var']],
        quantitative_sex_tests,
        on=['phenotype', 'activity_var'],
        how='left',
    ).rename(columns={"std_male_effect": "std_male_coeff", "std_female_effect":"std_female_coeff"})
    fig, ax = phewas_plots.sex_difference_plot(qt2_restricted, color_by="phenotype", cmap=color_by_lipoproteins, lim=0.25)
    fig.savefig(f"{OUTDIR}/sex_differences.quantitative.lipoproteins.png")

    ## Plot the by-age plot
    dage = with_meds_by_age.copy().sample(frac=1)
    dage['p_overall'] = pandas.merge(
        dage,
        with_meds,
        on=["phenotype", "activity_var"],
    ).p
    dage['p_age'] = dage.age_difference_p
    fig, ax = phewas_plots.age_effect_plot(dage.sample(frac=1), color_by="Functional Category", cmap=color_by_quantitative_function, lim=0.3)
    fig.savefig(f"{OUTDIR}/age_effects.with_meds.quantitative.png")

    # Just lipoproteins
    dage_restricted = dage[dage['Functional Category'] == 'Lipoprotein Profile'].sample(frac=1)
    fig, ax = phewas_plots.age_effect_plot(dage_restricted, color_by="phenotype", cmap=color_by_lipoproteins, lim=0.3)
    fig.savefig(f"{OUTDIR}/age_effects.with_meds.quantitative.lipoproteins.png")
    # without medications
    dage2_restricted = pandas.merge(
        dage_restricted[['phenotype', 'activity_var']],
        quantitative_age_tests,
        on=['phenotype', 'activity_var'],
        how='left',
    )
    dage2_restricted['p_overall'] = pandas.merge(
        dage2_restricted,
        quantitative_tests,
        on=['phenotype', 'activity_var']
    ).p
    dage2_restricted['p_age'] = dage2_restricted.age_difference_p
    fig, ax = phewas_plots.age_effect_plot(dage2_restricted, color_by="phenotype", cmap=color_by_lipoproteins, lim=0.3)
    fig.savefig(f"{OUTDIR}/age_effects.quantitative.lipoproteins.png")



def med_differences_plots():
    d = -numpy.log10(med_differences.pivot_table(columns=["medication"], index=["phenotype"], values="p") + 1e-20)

    import scipy.cluster.hierarchy
    Z = scipy.cluster.hierarchy.linkage(d.values)
    #leaves = scipy.cluster.hierarchy.optimal_leaf_ordering(Z, d.values)
    leaves_list = scipy.cluster.hierarchy.leaves_list(Z)
    d = d.iloc[leaves_list,:]
    ZT = scipy.cluster.hierarchy.linkage(d.values.T)
    #leavesT = scipy.cluster.hierarchy.optimal_leaf_ordering(ZT, d.values.T)
    leaves_list_T = scipy.cluster.hierarchy.leaves_list(ZT)
    d = d.iloc[:, leaves_list_T]

    fig, ax = pylab.subplots(figsize=(10,10))
    h = ax.imshow(d.values, vmax=20)
    ax.set_xticks(range(len(d.columns)))
    ax.set_xticklabels(d.columns, rotation=90)
    ax.set_yticks(range(len(d.index)))
    ax.set_yticklabels(d.index)
    fig.colorbar(h)
    fig.tight_layout()

medication_groups = {
    "beta_agonist": ["ventolin 100micrograms inhaler", "seretide 50 evohaler",],
    "beta_blocker": ["atenolol"],
}

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description="run phewas pipeline on actigraphy\nOutputs to ../global_phewas/cohort#/")
    parser.add_argument("--cohort", help="Cohort number to load data for", type = int)
    parser.add_argument("--force_recompute", help="Whether to force a rerun of the statistical tests, even if already computed", default=False, action="store_const", const=True)
    parser.add_argument("--all", help="Whether to run all analyses. Warning: slow.", default=False, action="store_const", const=True)
    parser.add_argument("--noplots", help="Disable running plots, useful for just loading the data", default=False, action="store_const", const=True)
    parser.add_argument("--no_display", help="Disable visual output, uses non-graphical backend such as when running on a server", default=False, action="store_const", const=True)

    args = parser.parse_args()

    if args.no_display:
        # Use the non-graphical backend Agg
        import matplotlib
        matplotlib.use("Agg")
    import pylab

    COHORT = args.cohort
    RECOMPUTE = args.force_recompute
    OUTDIR = f"../global_phewas/cohort{COHORT}/"
    FDR_CUTOFF_VALUE = 0.05

    #### Load and preprocess the underlying data
    data, ukbb, activity, activity_summary, activity_summary_seasonal, activity_variables, activity_variance, full_activity, phecode_data, phecode_groups, phecode_info, phecode_map, icd10_entries, icd10_entries_all, phecode_details = phewas_preprocess.load_data(COHORT)

    medications = phewas_preprocess.load_medications(data.index)

    # Load descriptions + categorization of activity variables and quantitative variables
    activity_variable_descriptions = pandas.read_excel("../table_header.xlsx", index_col="Activity Variable", sheet_name="Variables", engine="openpyxl")
    quantitative_variable_descriptions = pandas.read_excel("../quantitative_variables.xlsx", index_col=0, engine="openpyxl")

    # Gather list of the
    import fields_of_interest
    quantitative_blocks = [
        fields_of_interest.blood_fields,
        fields_of_interest.urine,
        fields_of_interest.arterial_stiffness,
        fields_of_interest.physical_measures,
    ]
    def find_var(var):
        for v in [var, var+"_V0"]:
            if v in data.columns:
                if pandas.api.types.is_numeric_dtype(data[v].dtype):
                    return v
        return None # can't find it
    quantitative_variables = [find_var(c) for block in quantitative_blocks
                            for c in block
                            if find_var(c) is not None]



    #### Run (or load from disk if they already exist) 
    #### the statistical tests
    phecode_tests, phecode_tests_by_sex = phewas_tests.phecode_tests(data, phecode_groups, activity_variables, activity_variable_descriptions, phecode_info, OUTDIR, RECOMPUTE)
    phecode_tests_raw = phecode_tests.copy()
    phecode_tests['activity_var_category'] = phecode_tests['activity_var'].map(activity_variable_descriptions.Category)
    phecode_tests['q_significant'] = (phecode_tests.q < FDR_CUTOFF_VALUE).astype(int)

    quantitative_tests, quantitative_age_tests, quantitative_sex_tests = phewas_tests.quantitative_tests(data, quantitative_variables, activity_variables, activity_variable_descriptions, quantitative_variable_descriptions, OUTDIR, RECOMPUTE)
    quantitative_tests_raw  = quantitative_tests.copy()

    age_tests = phewas_tests.age_tests(data, phecode_groups, activity_variables, activity_variable_descriptions, phecode_info, OUTDIR, RECOMPUTE)
    age_tests_raw = age_tests.copy()
    age_tests['activity_var_category'] = age_tests['activity_var'].map(activity_variable_descriptions.Category)

    age_sex_tests = phewas_tests.age_sex_interaction_tests(data, phecode_groups, activity_variables, phecode_info, OUTDIR, RECOMPUTE)

    survival_tests = phewas_tests.survival_tests(data, activity_variables, activity_variable_descriptions, OUTDIR, RECOMPUTE)

    beyond_RA_tests = phewas_tests.beyond_RA_tests(data, activity_variables, activity_variable_descriptions, OUTDIR, RECOMPUTE)

    med_differences = phewas_tests.assess_medications(data, quantitative_variables, medications, OUTDIR, RECOMPUTE)

    phecode_three_component_tests, phecode_three_component_tests_by_sex, phecode_three_component_tests_by_age, quantitative_three_component_tests, quantitative_three_component_tests_by_sex, quantitative_three_component_tests_by_age = phewas_tests.three_components_tests(data, phecode_groups, quantitative_variables, phecode_info, OUTDIR, RECOMPUTE)


    #### Prepare color maps for the plots
    color_by_phecode_cat = {cat:color for cat, color in
                                zip(phecode_tests.phecode_category.unique(),
                                    [pylab.get_cmap("tab20")(i) for i in range(20)])}
    color_by_actigraphy_cat = {cat:color for cat, color in
                                    zip(["Sleep", "Circadianness", "Physical activity"],
                                        [pylab.get_cmap("Dark2")(i) for i in range(20)])}
    color_by_actigraphy_subcat = {cat:color for cat, color in
                                    zip(activity_variable_descriptions.Subcategory.unique(),
                                        [pylab.get_cmap("Set3")(i) for i in range(20)])}
    color_by_quantitative_function = {cat:color for cat, color in
                                        zip(quantitative_variable_descriptions['Functional Categories'].unique(),
                                            [pylab.get_cmap("tab20b")(i) for i in range(20)])}
    colormaps = {
        "phecode_cat": color_by_phecode_cat,
        "actigraphy_cat": color_by_actigraphy_cat,
        "actigraphy_subcat": color_by_actigraphy_subcat,
        "quantitative_function": color_by_quantitative_function,
    }

    ## Create the plotter object
    # for common plot types
    phewas_plots = plots.Plotter(phecode_info, colormaps, activity_variables, activity_variable_descriptions, quantitative_variable_descriptions)

    ## Make the plots
    if not args.noplots:
        summary()
        sex_difference_plots()
        age_difference_plots()
        fancy_plots()
        survival_curves()
        survival_plots()
        objective_subjective_plots()
        circadian_component_plots()
        activity_var_comparisons()
        age_interaction_plots()
        temperature_trace_plots()
        temperature_calibration_plots()
        chronotype_plots()
        quantitative_traits_with_medications()
        if args.all:
            # Note: slow to run: performs many regressions
            by_date_plots()

        ## Summarize everything
        generate_results_table()

    demographics_table()