# # PheWAS analysis
# 
# Check if there are associations of phenotypes with circadian problems,
# particularly for unusual timing or for lack of consistent rhythm.

import re

import scipy
import scipy.stats
import numpy
import statsmodels.api as sm
import statsmodels.formula.api as smf
import seaborn as sns
import matplotlib
import matplotlib.patches as mpatches
from scipy.cluster import hierarchy
import pylab
import pandas
import networkx

pandas.plotting.register_matplotlib_converters()

COHORT = 1
OUTDIR = f"../global_phewas/cohort{COHORT}/"

### Whether to run all the big computations or attempt to load from disk since already computed
RECOMPUTE = False

# Point at which to call FDR q values significant
FDR_CUTOFF_VALUE = 0.05

# OLS method is either 'svd' or 'qr'
# Using 'qr' since intermittent problems wtih SVD convergence
OLS_METHOD = 'qr'
def OLS(*args, **kwargs):
    for i in range(100):
        try:
            return smf.ols(*args, **kwargs).fit()#method=OLS_METHOD)
        except numpy.linalg.LinAlgError:
            print("Failed regression:")
            print(args)
            print(kwargs)
            print("Attempt number {i}")
            continue
    raise numpy.linalg.LinAlgError

full_activity = pandas.read_csv("../processed/activity_features_aggregate_seasonal.txt", sep="\t", dtype={'Unnamed: 0': str})
activity_summary = pandas.read_csv("../processed/activity_summary_aggregate.txt", index_col=0, sep="\t")
activity_summary_seasonal = pandas.read_csv("../processed/activity_summary_aggregate_seasonal.txt", index_col=0, sep="\t")
activity_summary_seasonal["ID"] = activity_summary_seasonal.index.astype(int)
ukbb = pandas.read_hdf("../processed/ukbb_data_table.h5")

ukbb.columns = ukbb.columns.str.replace("[,:/]","_") # Can't use special characters easily

# Separate out the user ID from the run number (0 = original, 1-4 are seasonal repeats)
full_activity.rename(columns={"Unnamed: 0": "run_id"}, inplace=True)
full_activity['id'] = full_activity.run_id.apply(lambda x: int(x.split('.')[0]))
full_activity['run'] = full_activity.run_id.apply(lambda x: int(x.split('.')[1]))
activity = full_activity[full_activity.run == 0]
activity.set_index('id', inplace=True)


## Select the activity variables that have between-person variance greater than their within-person variance
# and for the summary variables, use only those that are overall summary variables
activity_variance = pandas.read_csv("../processed/inter_intra_personal_variance.seasonal_correction.txt", sep="\t", index_col=0)
activity_variance['summary_var'] = activity_variance.index.isin(activity_summary.columns)
activity_variance['use'] = (~activity_variance.summary_var) | activity_variance.index.str.contains("overall-")
good_variance = (activity_variance.corrected_intra_personal_normalized < 1)
activity_variables = activity_variance.index[good_variance & activity_variance.use]
activity_variables = activity_variables.intersection(activity.columns)

print(f"Started with {len(activity.columns.intersection(activity_variance[activity_variance.use].index))} activity variables")
activity = activity[activity.columns[activity.columns.isin(activity_variables)]]
print(f"Selected {len(activity.columns)} after discarding those with poor intra-personal variance")

# Load descriptions + categorization of activity variables
activity_variable_descriptions = pandas.read_excel("../table_header.xlsx", index_col="Activity Variable", sheet_name="Variables")

# Load descriptions of the quantitative variables
quantitative_variable_descriptions = pandas.read_excel("../quantitative_variables.xlsx", index_col=0)

# drop activity for people who fail basic QC
okay = (activity_summary['quality-goodCalibration'].astype(bool)
            & (~activity_summary['quality-daylightSavingsCrossover'].astype(bool))
            & (activity_summary['quality-goodWearTime'].astype(bool))
       )
activity = activity[okay]
activity.columns = activity.columns.str.replace("-","_") # Can't use special characters easily
print(f"Dropping {(~okay).sum()} entries out of {len(okay)} due to bad quality or wear-time")

activity_variance.index = activity_variance.index.str.replace("-","_") # Can't use special characters easily

## Process activity variables that need cleaning
activity.phase = activity.phase % 24

## Correct activity measures based off of seasonality
# compute the 'fraction of year' value (0 = January 1st, 1 = December 31st)
actigraphy_start_date = activity.index.map(pandas.to_datetime(activity_summary['file-startTime']))
year_start = pandas.to_datetime(actigraphy_start_date.year.astype(str) + "-01-01")
year_fraction = (actigraphy_start_date - year_start) / (pandas.to_timedelta("1Y"))
cos_year_fraction = numpy.cos(year_fraction*2*numpy.pi)
sin_year_fraction = numpy.sin(year_fraction*2*numpy.pi)

for var in activity_variables:
    if var.startswith("self_report"):
        continue
    activity[var] = activity[var] - cos_year_fraction * activity_variance.loc[var].cos - sin_year_fraction * activity_variance.loc[var].sin


## Add self-reported variables to activity document
# they need to be converted to 0,1 and NaNs
self_report_circadian_variables = {
    "daytime_dozing": {
        "name": "daytime_dozing",
        "zeros": ["Never/rarely"],
        "ones": ["All of the time", "Often"],
    },
    "getting_up_in_morning": {
        "name": "getting_up_in_morning",
        "zeros": ["Very easy"],
        "ones": ["Not at all easy"],
    },
    "morning_evening_person": {
        "name": "chronotype",
        "zeros": ["Definitely a 'morning' person"],
        "ones": ["Definitely an 'evening' person"],
    },
    "nap_during_day": {
        "name": "nap_during_day",
        "zeros": ["Never/rarely"],
        "ones": ["Usually"],
    },
    "sleep_duration": {
        "name": "sleep_duration",
        "zeros": None, # Sleep duration is quantative
        "ones": None,
    },
    "sleeplessness": {
        "name": "sleeplessness",
        "zeros": ["Never/rarely"],
        "ones": ["Usually"],
    },
    "snoring": {
        "name": "snoring",
        "zeros": ["No"],
        "ones": ["Yes"],
    },
    "IPAQ_activity_group": {
        "name": "IPAQ_activity_group",
        "zeros": ['low'],
        "ones": ['high'],
    },
}

self_report_data = {}
for variable, var_data in self_report_circadian_variables.items():
    if var_data['zeros'] is not None:
        # Binarize
        zeros = ukbb[variable].isin(var_data['zeros'])
        ones = ukbb[variable].isin(var_data['ones'])
        values = ones.astype(int)
        values[(~zeros) & (~ones)] = float("NaN")
    else:
        values = ukbb[variable]
    self_report_data["self_report_"+var_data['name']] = values
self_report_data = pandas.DataFrame(self_report_data)
activity = activity.join(self_report_data)

# Create absolute deviation variables from phase variables
# since both extreme low and high phase are expected to be correlated with outcomes
# and we do the same for sleep duration measures
for var in activity_variable_descriptions.index:
    if var.endswith('_abs_dev'):
        base_var = var[:-8]
        activity[var] = (activity[base_var] - activity[base_var].mean(axis=0)).abs()

# List the activity variables
activity_variables = activity.columns

# Gather all the data
data_full = activity.join(ukbb, how="inner")
print(f"Data starting size: {data_full.shape}")


# Down sample for testing
numpy.random.seed(0)
# Note: total 92331, half is 46164
cohort_id_ranges = {1: slice(0, 25000),
           2: slice(25000,50000)}
selected_ids = numpy.random.choice(data_full.index, size=data_full.shape[0], replace=False)[cohort_id_ranges[COHORT]]
data = data_full.loc[selected_ids].copy()
print(f"Data size after selecting test set: {data.shape}")

# Age/birth year processing
data['birth_year_category'] = pandas.cut(data.birth_year, bins=[1930, 1940, 1950, 1960, 1970])
data['actigraphy_start_date'] = data.index.map(pandas.to_datetime(activity_summary['file-startTime']))
def year_to_jan_first(year):
    if year != year:
        return float("NaN")
    else:
        return str(int(year)) + "-01-01"
birth_year = pandas.to_datetime(data.birth_year.apply(year_to_jan_first)) # As datetime
data['age_at_actigraphy'] = (data.actigraphy_start_date - birth_year) / pandas.to_timedelta("1Y")

# Create simplified versions of the categorical covarites
# This is necessary for convergence of the logistic models
data['ethnicity_white'] = data.ethnicity.isin(["British", "any other white background", "Irish", "White"])
data['overall_health_good'] = data.overall_health.isin(["Good", "Excellent"])
data.loc[data.overall_health.isin(['Do not know', 'Prefer not to answer']), 'overall_health_good'] = float("NaN")
data['smoking_ever'] = data.smoking.isin(['Previous', 'Current'])
data.loc[data.smoking == 'Prefer not to answer', 'smoking_ever'] = float("NaN")
data['high_income'] = data.household_income.isin(['52,000 to 100,000', 'Greater than 100,000'])
data.loc[data.high_income == 'Do not know', 'high_income'] = float("NaN")
data['college_education'] = data['education_College_or_University_degree']

# List of covariates we will controll for in the linear model
covariates = [
              "sex", "ethnicity_white", "overall_health_good", "high_income", "smoking_ever", "age_at_actigraphy", "BMI", "college_education"
                ]

# Q-value utility
def BH_FDR(ps):
    ''' Benjamini-Hochberg FDR control

    Converts p values to q values'''

    # For the purposes of comparison, an implementation of Benjamini Hochberg correction
    sort_order = numpy.argsort(ps)

    adjusted = numpy.zeros(ps.shape)
    adjusted[sort_order] = numpy.array(ps)[sort_order]*len(ps)/numpy.arange(1,len(ps)+1)

    # Make monotone, skipping NaNs
    m = 1
    for i, r in enumerate(sort_order[::-1]):
        if numpy.isfinite(adjusted[r]):
            m = min(adjusted[r], m)
            adjusted[r] = m

    return adjusted # the q-values

# Load the PheCode mappings
# Downloaded from https://phewascatalog.org/phecodes_icd10
# Has columns:
# ICD10 | PHECODE | Exl. Phecodes | Excl. Phenotypes
phecode_info = pandas.read_csv("../phecode_definitions1.2.csv", index_col=0)
phecode_map = pandas.read_csv("../Phecode_map_v1_2_icd10_beta.csv")
phecode_map.set_index(phecode_map.ICD10.str.replace(".",""), inplace=True) # Remove '.' to match UKBB-style ICD10 codes

# and convert to phecodes
# v1.2 Downloaded from https://phewascatalog.org/phecodes
phecode_map_icd9 = pandas.read_csv("../phecode_icd9_map_unrolled.csv")
phecode_map_icd9.rename(columns={"icd9":"ICD9", "phecode":"PHECODE"}, inplace=True)
phecode_map_icd9.set_index( phecode_map_icd9['ICD9'].str.replace(".",""), inplace=True) # Remove dots to match UKBB-style ICD9s

# ## Load the ICD10/9 code data
icd10_entries_all = pandas.read_csv("../processed/ukbb_icd10_entries.txt", sep="\t")
# Select our cohort from all the entries
icd10_entries = icd10_entries_all[icd10_entries_all.ID.isin(selected_ids)].copy()
icd10_entries.rename(columns={"ICD10_code": "ICD10"}, inplace=True)
icd10_entries = icd10_entries.join(phecode_map.PHECODE, on="ICD10")

### and the ICD9 data
icd9_entries_all = pandas.read_csv("../processed/ukbb_icd9_entries.txt", sep="\t")
# Select our cohort from all the entries
icd9_entries = icd9_entries_all[icd9_entries_all.ID.isin(selected_ids)].copy()
icd9_entries.rename(columns={"ICD9_code": "ICD9"}, inplace=True)
icd9_entries = icd9_entries.join(phecode_map_icd9.PHECODE, on="ICD9")

# Self-reported conditions from the interview stage of the UK Biobank
self_reported_all = pandas.read_csv("../processed/ukbb_self_reported_conditions.txt", sep="\t", dtype={"condition_code":int})
self_reported = self_reported_all[self_reported_all.ID.isin(selected_ids)].copy()
data_fields = pandas.read_csv("../Data_Dictionary_Showcase.csv", index_col="FieldID")
codings = pandas.read_csv("../Codings_Showcase.csv", dtype={"Coding": int})
SELF_REPORTED_CONDITION_FIELD = 20002
condition_code_to_meaning = codings[codings.Coding  == data_fields.loc[20002].Coding].drop_duplicates(subset=["Value"], keep=False).set_index("Value")
self_reported["condition"] = self_reported.condition_code.astype(str).map(condition_code_to_meaning.Meaning)

# # Run a PheCode-based analysis

# Convert self-reported conditions to phecodes

# Load Manaully mapped self-reports to phecodes
self_report_phecode_map = pandas.read_csv("../self_report_conditions_meanings.txt", sep="\t", index_col=0)
self_reported["phecode"] = self_reported.condition_code.map(self_report_phecode_map['PheCODE'])


# Gather whether each person has a diagnosis from a given PheCode group

# Group phecodes together that differ only after the '.'
# i.e. if they convert to the same integer
phecode_groups = phecode_info.index.astype(int).unique()

phecode_data_icd10 = {}
phecode_data_icd9 = {}
phecode_data_self_report = {}
for group in phecode_groups:
    group_data = phecode_info[phecode_info.index.astype(int) == group]
    icd10_codes = phecode_map[phecode_map.PHECODE.isin(group_data.index)].index
    icd9_codes = phecode_map_icd9[phecode_map_icd9.PHECODE.isin(group_data.index)].index
    in_block = icd10_entries.ICD10.isin(icd10_codes)
    in_block_icd9 = icd9_entries.ICD9.isin(icd9_codes)
    
    diagnosed = in_block.groupby(icd10_entries.ID).any()
    phecode_data_icd10[group] = diagnosed
    phecode_data_icd9[group] =  in_block_icd9.groupby(icd9_entries.ID).any()
    phecode_data_self_report[group] = self_reported.phecode.isin(group_data.index)

phecode_data_icd10 = pandas.DataFrame(phecode_data_icd10)
phecode_data_icd9 = pandas.DataFrame(phecode_data_icd9)
phecode_data_self_report = pandas.DataFrame(phecode_data_self_report).set_index(self_reported.ID)
phecode_data = pandas.concat([phecode_data_icd10, phecode_data_icd9, phecode_data_self_report]).reset_index().groupby(by="ID").any()


## Tally the number of occurances of each phecode (at the lowest level, not just groups)
phecode_count_details = pandas.concat([
    icd10_entries[['ID', 'PHECODE']],
    icd9_entries[['ID', 'PHECODE']],
    self_reported[['ID', 'phecode']].rename(columns={"phecode":"PHECODE"})
]).groupby('PHECODE').ID.nunique()
phecode_count_details = pandas.DataFrame({"count": phecode_count_details})
phecode_count_details['phecode_meaning'] = phecode_count_details.index.map(phecode_info.phenotype)
phecode_count_details['phecode_category'] = phecode_count_details.index.map(phecode_info.category)
phecode_count_details.to_csv(OUTDIR+"phecode_counts.txt", sep="\t", header=True)


# ### Display which sources the cases come from for the top codes


phecode_counts = pandas.DataFrame({"counts": phecode_data.sum(axis=0)})
for name, d in {"icd10": phecode_data_icd10, "icd9": phecode_data_icd9, "self_report": phecode_data_self_report}.items():
    cases = d.reset_index().groupby(by="ID").any()
    phecode_counts[name + "_cases"] = cases.sum(axis=0)
phecode_counts["phecode_meaning"] = phecode_counts.index.map(phecode_info.phenotype)
print("Most frequent phecodes by source")
print(phecode_counts.sort_values(by="counts", ascending=False).head(20))

# Gather phecode diagnosis information for each subject
for group in phecode_groups:
    # Note that those without any ICD10 entries at all should be marked as non-case, hence the fillna()
    data[group] = data.index.map(phecode_data[group].astype(int)).fillna(0)

# Correlate each block-level code with our activity variable
# Loop through all the activity variables and phecode groups we are interested in
def compute_phecode_test(activity_variable, phecode, data):
    covariate_formula = ' + '.join(c for c in covariates if c != 'sex')
    fit = OLS(f"{activity_variable} ~ Q({phecode}) + sex * ({covariate_formula})",
                 data=data)
    p = fit.pvalues[f"Q({phecode})"]
    coeff = fit.params[f"Q({phecode})"]
    std_effect = coeff / data[activity_variable].std()
    N_cases = data.loc[~data[activity_variable].isna(), phecode].sum()
    return {"phecode": phecode,
             "activity_var": activity_variable,
             "p": p,
             "coeff": coeff,
             "std_effect": std_effect,
             "N_cases": N_cases,
    }, fit 

## Compute the phecode associations
if RECOMPUTE:
    phecode_tests_list = []
    for group in phecode_groups:
        print(group, )
        N = data[group].sum()
        if N < 50:
            print(f"Skipping {group} - only {N} cases found")
            continue
        
        for activity_variable in activity.columns:
            summary, fit = compute_phecode_tests(activity_variable, group, data)
            phecode_tests.list.append(summary)
    phecode_tests = pandas.DataFrame(phecode_tests_list)

    phecode_tests['q'] = BH_FDR(phecode_tests.p)
    phecode_tests["phecode_meaning"] = phecode_tests.phecode.map(phecode_info.phenotype)
    phecode_tests["phecode_category"] = phecode_tests.phecode.map(phecode_info.category)

    phecode_tests.to_csv(OUTDIR+f"phecodes.txt", sep="\t", index=False)
else:
    phecode_tests = pandas.read_csv(OUTDIR+"phecodes.txt", sep="\t")
phecode_tests_raw = phecode_tests.copy()

phecode_tests['activity_var_category'] = phecode_tests['activity_var'].map(activity_variable_descriptions.Category)
phecode_tests['q_significant'] = (phecode_tests.q < FDR_CUTOFF_VALUE).astype(int)

# Summarize the phecode test results
num_nonnull = len(phecode_tests) - phecode_tests.p.sum()*2
bonferonni_cutoff = 0.05 / len(phecode_tests)
FDR_cutoff = phecode_tests[phecode_tests.q < 0.05].p.max()
print(f"Of {len(phecode_tests)} tested, approx {int(num_nonnull)} expected non-null")
print(f"and {(phecode_tests.p <= bonferonni_cutoff).sum()} exceed the Bonferonni significance threshold")
print(f"and {(phecode_tests.p <= FDR_cutoff).sum()} exceed the FDR < 0.05 significance threshold")

### Prepare color maps for the plots
color_by_phecode_cat = {cat:color for cat, color in
                            zip(phecode_tests.phecode_category.unique(),
                                [pylab.get_cmap("tab20")(i) for i in range(20)])}

color_by_actigraphy_cat = {cat:color for cat, color in
                                zip(["Sleep", "Circadianness", "Physical activity"],
                                    [pylab.get_cmap("Dark2")(i) for i in range(20)])}
color_by_actigraphy_subcat = {cat:color for cat, color in
                                zip(activity_variable_descriptions.Subcategory.unique(),
                                    [pylab.get_cmap("Set3")(i) for i in range(20)])}
color_by_quantitative_function = {cat:color for cat, color in
                                    zip(quantitative_variable_descriptions['Functional Categories'].unique(),
                                        [pylab.get_cmap("tab20b")(i) for i in range(20)])}
def legend_from_colormap(fig, colormap, **kwargs):
    legend_elts = [matplotlib.lines.Line2D(
                            [0],[0],
                            marker="o", markerfacecolor=c, markersize=10,
                            label=cat if not pandas.isna(cat) else "NA",
                            c=c, lw=0)
                        for cat, c in colormap.items()]
    fig.legend(handles=legend_elts, **kwargs)
def legend_of_pointscale(fig, offset, values_to_size, values_to_show, fmt="{}", **kwargs):
    legend_elts = [matplotlib.lines.Line2D(
                            [0],[0],
                            marker="o", markerfacecolor='k',
                            markersize=numpy.sqrt(offset + values_to_size * value),
                            label=fmt.format(value),
                            c='k', lw=0)
                        for value in values_to_show]
    fig.legend(handles=legend_elts, **kwargs)

## Prepare utility functions
def truncate(string, N):
    if len(string) <= N:
        return string
    return string[:N-3] + "..."

def wrap(string, N):
    # Wrap to lines at most length N, preserving words
    words = string.split()
    lines = []
    current_line = []
    for word in words:
        if len(word) + sum(len(w) for w in current_line) > N:
            lines.append(' '.join(current_line))
            current_line = []
        current_line.append(word)
    lines.append(' '.join(current_line))
    return '\n'.join(lines)

### Create summary plots
fig, ax = pylab.subplots(figsize=(8,6))
color = phecode_tests.phecode_category.map(color_by_phecode_cat)
ax.scatter(phecode_tests.N_cases, -numpy.log10(phecode_tests.p), marker="+", c=color)
ax.set_xlabel("Number cases")
ax.set_ylabel("-log10 (p-value)")
ax.axhline( -numpy.log10(bonferonni_cutoff), c="k", zorder = -1 )
ax.axhline( -numpy.log10(FDR_cutoff), c="k", linestyle="--", zorder = -1 )
ax.set_title("PheCode - Activity associations")
fig.savefig(OUTDIR+"phewas_summary.png")


fig, ax = pylab.subplots(figsize=(8,6))
pt = phecode_tests.sample(frac=1) # Randomly reorder for the plot
color = pt.phecode_category.map(color_by_phecode_cat)
ax.scatter(pt.std_effect, -numpy.log10(pt.p), marker="+", c=color)
ax.set_xlabel("Effect size")
ax.set_ylabel("-log10(p-value)")
ax.axhline( -numpy.log10(bonferonni_cutoff), c="k", zorder = -1 )
ax.axhline( -numpy.log10(FDR_cutoff), c="k", linestyle="--", zorder = -1 )
ax.set_title("PheCode - Activity associations")
legend_from_colormap(ax, color_by_phecode_cat, ncol=2, fontsize="small")
fig.savefig(OUTDIR+"phewas.volcano_plot.png")


# # Test sex differences in actigraphy-diagnosis associations
# 
# We are interested in whether there is a difference between male and female susceptibility to
# loss of circadian rhythm and differences in the impact of loss of circadian rhythm.
# 
# We extract the most significant associations and plot their associations within each sex.

# Correlate each block-level code with our activity variable within each sex

if RECOMPUTE:
    phecode_tests_by_sex_list = []
    sex_covariate_formula = ' + '.join(c for c in covariates if c != 'sex')

    for group in phecode_groups:
        #TODO: skip any sex-specific phecodes
        N = data[group].sum()
        N_male = numpy.sum(data[group].astype(bool) & (data.sex == "Male"))
        N_female = numpy.sum(data[group].astype(bool) & (data.sex == "Female"))
        if N_male <= 50 or N_female < 50:
            print(f"Skipping {group} - only {N_male} M and  {N_female} F cases found")
            continue
            
        if False: #phecode_tests.loc[group, "q"] > 0.01:
            # Skip test, not significant
            print(f"Skipping {group} since q > 0.01")
            continue
        
        for activity_variable in activity.columns:
            fit = OLS(f"{activity_variable} ~ 0 + C(sex, Treatment(reference=-1)) : ({sex_covariate_formula} +  Q({group}))",
                             data=data)


            female_coeff = fit.params[f'C(sex, Treatment(reference=-1))[Female]:Q({group})']
            male_coeff = fit.params[f'C(sex, Treatment(reference=-1))[Male]:Q({group})']
            p_female = fit.pvalues[f'C(sex, Treatment(reference=-1))[Female]:Q({group})']
            p_male = fit.pvalues[f'C(sex, Treatment(reference=-1))[Male]:Q({group})']
            diff_test = fit.t_test(f'C(sex, Treatment(reference=-1))[Male]:Q({group}) = C(sex, Treatment(reference=-1))[Female]:Q({group})')
            p_diff = diff_test.pvalue
            conf_ints = fit.conf_int()
            male_conf_int = conf_ints.loc[f'C(sex, Treatment(reference=-1))[Male]:Q({group})']
            female_conf_int = conf_ints.loc[f'C(sex, Treatment(reference=-1))[Female]:Q({group})']

            male_std = data.loc[data.sex == "Male", activity_variable].std()
            female_std = data.loc[data.sex == "Female", activity_variable].std()
            
            phecode_tests_by_sex_list.append({
                "phecode": group,
                "activity_var": activity_variable,
                "std_male_coeff": float(male_coeff) / male_std,
                "std_female_coeff": float(female_coeff) /  female_std,
                "p_male": float(p_male),
                "p_female": float(p_female),
                "p_diff": float(p_diff),
                "N_male": N_male,
                "N_female": N_female,
                "std_male_coeff_low": float(male_conf_int[0]) / male_std,
                "std_male_coeff_high": float(male_conf_int[1]) / male_std,
                "std_female_coeff_low": float(female_conf_int[0]) / female_std,
                "std_female_coeff_high": float(female_conf_int[1]) / female_std,
            })

    phecode_tests_by_sex = pandas.DataFrame(phecode_tests_by_sex_list)

    phecode_tests_by_sex["phecode_meaning"] = phecode_tests_by_sex.phecode.map(phecode_info.phenotype)
    phecode_tests_by_sex["phecode_category"] = phecode_tests_by_sex.phecode.map(phecode_info.category)
    phecode_tests_by_sex = phecode_tests_by_sex.join(phecode_tests.q, how="left")
    phecode_tests_by_sex['q_diff'] = BH_FDR(phecode_tests_by_sex.p_diff)
    phecode_tests_by_sex['differential_std_coeff'] = phecode_tests_by_sex.std_male_coeff - phecode_tests_by_sex.std_female_coeff
    phecode_tests_by_sex.sort_values(by="p_diff", inplace=True)

    phecode_tests_by_sex.to_csv(OUTDIR+"/all_phenotypes.by_sex.txt", sep="\t", index=False)
else:
    phecode_tests_by_sex = pandas.read_csv(OUTDIR+"/all_phenotypes.by_sex.txt", sep="\t")
phecode_tests_by_sex_raw = phecode_tests_by_sex.copy()

### Generate summaries of the phecode test by-sex results

# Plot the regression coefficients for each of the phenotypes
def sex_difference_plot(d, color_by="phecode_category", cmap="Dark2", lim=0.5, ax=None, legend=True, labels=True):
    if color_by == "phecode_category":
        colormap = color_by_phecode_cat
        color = [colormap[c] for c in d[color_by]]
    elif color_by is not None:
        if type(cmap) == str:
            cats = d[color_by].unique()
            if cmap == "rainbow":
                cmap = [pylab.get_cmap("rainbow")(i) for i in numpy.arange(len(cats))/len(cats)]
            else:
                cmap = [pylab.get_cmap(cmap)(i) for i in range(len(cats))]
            colormap = {cat:color for cat, color in
                                zip(cats, cmap)}
        else:
            colormap = cmap
        color = [colormap[c] for c in d[color_by]]
    else:
        color = None
    if ax is None:
        fig, ax = pylab.subplots(figsize=(9,9))
        just_ax = False
    else:
        fig = ax.figure
        just_ax = True
    # The points
    ax.scatter(
        d.std_male_coeff,
        d.std_female_coeff,
        label="phenotypes",
        #s=-numpy.log10(d.p_diff)*10,
        s=-numpy.log10(numpy.minimum(d.p_male, d.p_female))*4,
        c=color)
    ax.spines['bottom'].set_color(None)
    ax.spines['top'].set_color(None)
    ax.spines['left'].set_color(None)
    ax.spines['right'].set_color(None)
    ax.axvline(c="k", lw=1)
    ax.axhline(c="k", lw=1)
    if labels:
        #ax.set_title("Effect sizes by sex\nAmong signifcant associations")
        ax.set_xlabel("Effect size in males")
        ax.set_ylabel("Effect size in females")
        bbox = {'facecolor': (1,1,1,0.8), 'edgecolor':(0,0,0,0)}
        ax.annotate("Male Effect Larger", xy=(0.8*lim,0), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Male Effect Larger", xy=(-0.8*lim,0), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Female Effect Larger", xy=(0,0.8*lim), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Female Effect Larger", xy=(0,-0.5*lim), ha="center", bbox=bbox, zorder=3)
    ax.set_aspect("equal")
    ax.set_xlim(-lim, lim)
    ax.set_ylim(-lim, lim)
    # Diagonal y=x line
    bound = max(abs(numpy.min([ax.get_xlim(), ax.get_ylim()])),
                numpy.max([ax.get_xlim(), ax.get_ylim()]))
    diag = numpy.array([-bound, bound])
    ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
    ax.plot(diag, -diag, linestyle="--", c='k', zorder=-1, label="diagonal")
    if color_by is not None and legend:
        legend_elts = [matplotlib.lines.Line2D(
                                [0],[0],
                                marker="o", markerfacecolor=c, markersize=10,
                                label=truncate(cat,35) if not pandas.isna(cat) else "NA",
                                c=c, lw=0)
                            for cat, c in colormap.items()]
        if just_ax:
            ax.legend(handles=legend_elts, ncol=2, fontsize="small")
        else:
            fig.legend(handles=legend_elts, ncol=2, fontsize="small")
    return fig, ax

num_male = (data.sex == "Male").sum()
num_female = (data.sex == "Female").sum()
d = phecode_tests_by_sex[True #(phecode_tests_by_sex.q < 0.05 )
                        & (phecode_tests_by_sex.N_male > 300)
                        & (phecode_tests_by_sex.N_female > 300)]
fig, ax = sex_difference_plot(d)
fig.savefig(f"{OUTDIR}/sex_differences.all_phenotypes.png")

fig, ax = sex_difference_plot(d[d.phecode_category == 'circulatory system'], color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/sex_differences.circulatory.png")
fig, ax = sex_difference_plot(d[d.phecode_category == 'mental disorders'], color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/sex_differences.mental_disorders.png")
fig, ax = sex_difference_plot(d[d.phecode_category == 'endocrine/metabolic'], color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/sex_differences.endocrine.png")
fig, ax = sex_difference_plot(d[d.phecode_category == 'infectious diseases'], color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/sex_differences.infections.png")
fig, ax = sex_difference_plot(d[d.phecode_category == 'respiratory'], color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/sex_differences.respiratory.png")

#Make 2x2 grid of quantitative sex differences
fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
ij = [[0,0], [0,1], [1,0], [1,1]]
SUBCATEGORIES = ["circulatory system", "mental disorders", "endocrine/metabolic", "respiratory"]
for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
    tests = d[d.phecode_category == cat]
    sex_difference_plot(tests.sample(frac=1), color_by="phecode_meaning", ax=ax, legend=True, labels=False, cmap="tab20_r")
    ax.set_title(cat)
    if j == 0:
        ax.set_ylabel("Effect size in females")
    if i == 1:
        ax.set_xlabel("Effect size in males")
fig.tight_layout()
fig.savefig(OUTDIR+"sex_differences.2x2.png")


#Make 2x2 grid of quantitative sex differences
fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
ij = [[0,0], [0,1], [1,0], [1,1]]
SUBCATEGORIES = ["circulatory system", "mental disorders", "endocrine/metabolic", "respiratory"]
for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
    tests = d[d.phecode_category == cat]
    sex_difference_plot(tests.sample(frac=1), color_by="phecode_meaning", ax=ax, legend=True, labels=False, cmap="tab20_r")
    ax.set_title(cat)
    if j == 0:
        ax.set_ylabel("Effect size in females")
    if i == 1:
        ax.set_xlabel("Effect size in males")
fig.tight_layout()
fig.savefig(OUTDIR+"sex_differences.2x2.png")


def local_regression(x,y, out_x, bw=0.05):
    # Preform a local regression y ~ x and evaluate it at the provided points `out_x`
    reg = sm.nonparametric.KernelReg(exog=x, endog=y, var_type='c',
                                     bw=[bw],
                                    )
    fit, mfx = reg.fit(out_x)
    return fit


# Plot disease-incidence rates versus the coefficients, in both male and female
# TODO: should be done selecting only significant associations or all?
fig, ax = pylab.subplots( figsize=(9,9) )
selected = (phecode_tests_by_sex.q < FDR_CUTOFF_VALUE) & (phecode_tests_by_sex.N_male > 100) & (phecode_tests_by_sex.N_female > 100)
d = phecode_tests_by_sex[selected]
ax.scatter(numpy.log10(d.N_male),
            d.std_male_coeff.abs(),
            c="r", label="Male", marker='.')
ax.scatter(numpy.log10(d.N_female),
            d.std_female_coeff.abs(),
            c="b", label="Female", marker='.')
for i in range(1):
    if i > 0:
        d = phecode_tests_by_sex[selected].sample(len(phecode_tests_by_sex),replace=True)
    else:
        d = phecode_tests_by_sex[selected]
    male_smooth = sm.nonparametric.lowess(
                        d.std_male_coeff.abs(),
                        numpy.log10(d.N_male),
                        return_sorted=True,
                        frac=0.4,
                        it=0)
    ax.plot(male_smooth[:,0], male_smooth[:,1], c="r", alpha=1, linewidth=5)
    female_smooth = sm.nonparametric.lowess(
                        d.std_female_coeff.abs(),
                        numpy.log10(d.N_female),
                        return_sorted=True,
                        frac=0.4,
                        it=0)
    ax.plot(female_smooth[:,0], female_smooth[:,1], c="b", alpha=1, linewidth=5)
ax.legend()
ax.set_xlabel("Number of Cases (log10)")
ax.set_ylabel("Standardized Effect Size")
ax.set_title("Phenotype-Rhythmicity association by prevalnce rate")
#ax.set_ylim(-0.04,0.00)
fig.savefig(OUTDIR+"/all_phenotypes.by_incidence_rate.png")

# ### Check the overall average of effect size by sex of the RA-phenotype associations

male_weights = 1 / (phecode_tests_by_sex.std_male_coeff_high - phecode_tests_by_sex.std_male_coeff_low)**2 * (phecode_tests_by_sex.std_male_coeff != 0.0)
female_weights = 1 / (phecode_tests_by_sex.std_female_coeff_high - phecode_tests_by_sex.std_female_coeff_low)**2 * (phecode_tests_by_sex.std_female_coeff != 0.0)
rel_male_coeff = numpy.abs(phecode_tests_by_sex.std_male_coeff  * male_weights)
rel_female_coeff = numpy.abs(phecode_tests_by_sex.std_female_coeff * female_weights)

print(f"Weighted mean male effect:   {rel_male_coeff.mean() / male_weights.mean():0.4f}")
print(f"Median male effect:          {phecode_tests_by_sex.std_male_coeff.abs().median():0.4f}")
print(f"Weighted mean female effect: {rel_female_coeff.mean() / male_weights.mean():0.4f}")
print(f"Median female effect:        {phecode_tests_by_sex.std_female_coeff.abs().median():0.4f}")
#print(f"Note: effects are the difference in mean RA values between cases and controls of the phenotype.")
#print(f"   standard deviation of RA:  {data.acceleration_RA.std():0.4f}")

### Associate with quantitiative traits
# Quantitative traits:
import fields_of_interest
quantitative_blocks = [
    fields_of_interest.blood_fields,
    fields_of_interest.urine,
    fields_of_interest.arterial_stiffness,
    fields_of_interest.physical_measures,
]
def find_var(var):
    for v in [var, var+"_V0"]:
        if v in data.columns:
            if pandas.api.types.is_numeric_dtype(data[v].dtype):
                return v
    return None # can't find it
quantitative_vars = [find_var(c) for block in quantitative_blocks
                        for c in block
                        if find_var(c) is not None]

if RECOMPUTE:
    quantitative_tests_list = []
    covariate_formula = ' + '.join(c for c in covariates if c != 'sex')
    for phenotype in quantitative_vars:
        if phenotype in covariates:
            # Can't regress a variable that is also a exogenous variable (namely, BMI)
            continue

        N = data[phenotype].count()
        if N < 50:
            print(f"Skipping {phenotype} - only {N} cases found")
            continue
        
        phenotype_std = data[phenotype].std()
        for activity_var in activity.columns:
            N = (~data[[activity_var, phenotype]].isna().any(axis=1)).sum()
            fit = OLS(f"{phenotype} ~ {activity_var} + sex * ({covariate_formula})",
                         data=data)
            p = fit.pvalues[activity_var]
            coeff = fit.params[activity_var]
            activity_var_std = data[activity_var].std()
            std_effect = coeff * activity_var_std / phenotype_std
            # Fit the by-sex fit
            sex_fit = OLS(f"{phenotype} ~ 0 + C(sex, Treatment(reference=-1)) : ({activity_var} +  {covariate_formula})",
                             data=data)
            _, sex_difference_p, _ = sex_fit.compare_f_test(fit)
            female_coeff = sex_fit.params[f'C(sex, Treatment(reference=-1))[Female]:{activity_var}']
            male_coeff = sex_fit.params[f'C(sex, Treatment(reference=-1))[Male]:{activity_var}']
            p_female = sex_fit.pvalues[f'C(sex, Treatment(reference=-1))[Female]:{activity_var}']
            p_male = sex_fit.pvalues[f'C(sex, Treatment(reference=-1))[Male]:{activity_var}']
            #diff_test = sex_fit.t_test(f'C(sex, Treatment(reference=-1))[Male]:{activity_var} = C(sex, Treatment(reference=-1))[Female]:{activity_var}')
            #p_diff = diff_test.pvalue
            male_std_ratio = data.loc[data.sex == "Male", activity_var].std() / data.loc[data.sex == "Male", phenotype].std()
            female_std_ratio = data.loc[data.sex == "Female", activity_var].std() / data.loc[data.sex == "Female", phenotype].std()
            #By-age association
            age_fit = OLS(f"{phenotype} ~ {activity_var} * age_at_actigraphy + sex * ({covariate_formula})",
                             data=data)
            age_difference_p = age_fit.pvalues[f"{activity_var}:age_at_actigraphy"]
            main_coeff = age_fit.params[activity_var]
            age_coeff = age_fit.params[f"{activity_var}:age_at_actigraphy"]
            std_age_effect = age_coeff * activity_var_std / phenotype_std
            quantitative_tests_list.append({"phenotype": phenotype,
                                    "activity_var": activity_var,
                                    "p": p,
                                    "coeff": coeff,
                                    "std_effect": std_effect,
                                    "sex_difference_p": sex_difference_p,
                                    "p_male": p_male,
                                    "p_female": p_female,
                                    "std_male_coeff": male_coeff * male_std_ratio,
                                    "std_female_coeff": female_coeff * female_std_ratio,
                                    "age_difference_p": age_difference_p,
                                    "age_main_coeff": main_coeff,
                                    "age_effect_coeff": age_coeff,
                                    "std_age_effect": std_age_effect,
                                    "N": N,
                                   })

    quantitative_tests = pandas.DataFrame(quantitative_tests_list)
    quantitative_tests['q'] = BH_FDR(quantitative_tests.p)
    def base_name(x):
        if "_V" in x:
            return x.split("_V")[0]
        return x
    base_variable_name = quantitative_tests.phenotype.apply(base_name)
    quantitative_tests['ukbb_field'] = base_variable_name.map(fields_of_interest.all_fields)
    quantitative_tests.to_csv(OUTDIR+"/quantitative_traits.txt", sep="\t", index=False)
else:
    quantitative_tests = pandas.read_csv(OUTDIR+"/quantitative_traits.txt", sep="\t")
quantitative_tests_raw = quantitative_tests.copy()
quantitative_tests['Functional Category'] = quantitative_tests.phenotype.map(quantitative_variable_descriptions['Functional Categories'])

pylab.close('all')


### Network analysis between quantitative traits and a phenotype
## Connect a quantitative trait to a phenotype if they both associate with the same activity value
quantitative_tests['q_significant'] = (quantitative_tests.q < 0.05).astype(int)
quantitative_associations = quantitative_tests.set_index(["phenotype", "activity_var"]).q_significant.unstack()
phecode_associations = phecode_tests.set_index(["phecode", "activity_var"]).q_significant.unstack()
common_associations = phecode_associations @ quantitative_associations.T
common_associations.index = common_associations.index.map(phecode_info.phenotype)

# Compute the directionality and magnitude of these associations
# I.e. for each activity variable that associates significantly with both a phenotype and a trait
# we compute the direction of the common association between the phenotype and trait
# as the product of the activity variable effects with the two
# Then we aggregate across all the significant associates to get the dominant direction of association
# if any. If all point in the same direction, get +/- 1. If directionality varies randomly, get ~0
quantitative_effects = numpy.sign(quantitative_tests.set_index(["phenotype", "activity_var"]).std_effect.unstack() * quantitative_associations)
phecode_effects = numpy.sign(phecode_tests.set_index(["phecode", "activity_var"]).std_effect.unstack() * phecode_associations)
common_direction = (phecode_effects @ quantitative_effects.T)
common_direction.index = common_direction.index.map(phecode_info.phenotype)
common_direction = (common_direction / common_associations).fillna(0)

ax = plot_heatmap(common_associations)
ax.figure.savefig(OUTDIR+"/common_assocations_heatmap.png")

# Compute the number of expected associations assuming independence
expected_associations = phecode_associations.mean(axis=1).values.reshape((-1,1)) @ quantitative_associations.mean(axis=1).values.reshape((1,-1))
expected_associations = pandas.DataFrame(expected_associations,
                            index = common_associations.index,
                            columns = common_associations.columns)
common_associations_ratio = (common_associations / (expected_associations+0.001)).fillna(0)
ax = plot_heatmap(common_associations_ratio)
ax.figure.savefig(OUTDIR+"/common_assocations_ratio_heatmap.png")

# Run sex-difference and age-difference plots on the quantitative tests
fig, ax = sex_difference_plot(quantitative_tests.sample(frac=1), color_by="Functional Category", cmap=color_by_quantitative_function, lim=0.25)
fig.savefig(OUTDIR+"sex_differences.quantitative.png")

#Make 2x2 grid of quantitative sex differences
fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
ij = [[0,0], [0,1], [1,0], [1,1]]
SUBCATEGORIES = ["Metabolism", "Lipoprotein Profile", "Cardiovascular Function", "Renal Function"]
for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
    tests = quantitative_tests[quantitative_tests['Functional Category'] == cat]
    sex_difference_plot(tests.sample(frac=1), color_by="phenotype", lim=0.25, ax=ax, legend=True, labels=False, cmap="tab20_r")
    ax.set_title(cat)
    if j == 0:
        ax.set_ylabel("Effect size in females")
    if i == 1:
        ax.set_xlabel("Effect size in males")
fig.tight_layout()
fig.savefig(OUTDIR+"sex_differences.quantitative.2x2.png")

##### Age-associations
if RECOMPUTE:
    print("Computing age associations")
    age_tests_list = []
    covariate_formula = ' + '.join(c for c in covariates if (c != 'birth_year'))
    for group in phecode_groups:
        N = data[group].sum()
        if N < 200:
            print(f"Skipping {group} - only {N} cases found")
            continue
        
        for activity_variable in activity.columns:
            #if not phecode_tests[(phecode_tests.phecode == group)
            #                     & (phecode_tests.activity_var == activity_variable)].q_significant.any():
            #    continue # Only check for age-effects in significant main-effects variables

            fit = OLS(f"{activity_variable} ~ Q({group}) * age_at_actigraphy + ({covariate_formula})",
                         data=data)
            p = fit.pvalues[f"Q({group}):age_at_actigraphy"]
            main_coeff = fit.params[f"Q({group})"]
            age_coeff = fit.params[f"Q({group}):age_at_actigraphy"]
            std_effect = age_coeff / data[activity_variable].std()
            age_tests_list.append({"phecode": group,
                                    "activity_var": activity_variable,
                                    "p": p,
                                    "main_coeff": main_coeff,
                                    "age_effect_coeff": age_coeff,
                                    "std_age_effect": std_effect,
                                    "N_cases": N,
                                   })
    age_tests = pandas.DataFrame(age_tests_list)

    age_tests['q'] = BH_FDR(age_tests.p)
    age_tests["phecode_meaning"] = age_tests.phecode.map(phecode_info.phenotype)
    age_tests["phecode_category"] = age_tests.phecode.map(phecode_info.category)

    age_tests.to_csv(OUTDIR+f"phecodes.age_effects.txt", sep="\t", index=False)
else:
    age_tests = pandas.read_csv(OUTDIR+"phecodes.age_effects.txt", sep="\t")

age_tests['activity_var_category'] = age_tests['activity_var'].map(activity_variable_descriptions.Category)


## Plot summary of age tests
mean_age = data.age_at_actigraphy.mean()
young_offset = 55 - mean_age
old_offset = 70 - mean_age
d = pandas.merge(
        age_tests,
        phecode_tests[['phecode', 'activity_var', 'std_effect', 'p', 'q']],
        suffixes=["_age", "_overall"],
        on=["activity_var", "phecode"]).reset_index()
d = d[d.N_cases > 500]
d['age_55_effect'] = d["std_effect"] + d['std_age_effect'] * young_offset
d['age_75_effect'] = d["std_effect"] + d['std_age_effect'] * old_offset

def age_effect_plot(d, legend=True, labels=True, color_by="phecode_category", cmap="Dark2", lim=0.45, ax=None):
    if ax is None:
        fig, ax = pylab.subplots(figsize=(9,9))
        just_ax = False
    else:
        fig = ax.figure
        just_ax = True
    if color_by == "phecode_category":
        colormap = color_by_phecode_cat
        color = [colormap[c] for c in d[color_by]]
    elif color_by is not None:
        if type(cmap) == str:
            cats = d[color_by].unique()
            if cmap == "rainbow":
                cmap = [pylab.get_cmap("rainbow")(i) for i in numpy.arange(len(cats))/len(cats)]
            else:
                cmap = [pylab.get_cmap(cmap)(i) for i in range(len(cats))]
            colormap = {cat:color for cat, color in
                                zip(cats, cmap)}
        else:
            colormap = cmap
        color = [colormap[c] for c in d[color_by]]
    else:
        color = None
    color = [colormap[c] for c in d[color_by]]
    # The points
    ax.scatter(
        d.age_55_effect,
        d.age_75_effect,
        s=-numpy.log10(numpy.minimum(d.p_overall, d.p_age))*3,
        c=color)
    ax.spines['bottom'].set_color(None)
    ax.spines['top'].set_color(None)
    ax.spines['left'].set_color(None)
    ax.spines['right'].set_color(None)
    ax.axvline(c="k", lw=1)
    ax.axhline(c="k", lw=1)
    ax.set_xlim(-lim, lim)
    ax.set_ylim(-lim, lim)
    #ax.set_xticks(numpy.linspace(-0.4,0.4,11))
    #ax.set_yticks(numpy.linspace(-0.4,0.4,11))
    # Diagonal y=x line
    bound = max(abs(numpy.min([ax.get_xlim(), ax.get_ylim()])),
                numpy.max([ax.get_xlim(), ax.get_ylim()]))
    diag = numpy.array([-bound, bound])
    ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal", linewidth=1)
    ax.plot(diag, -diag, linestyle="--", c='k', zorder=-1, label="diagonal", linewidth=1)
    ax.set_aspect("equal")
    if labels:
        #ax.set_title("Effect sizes by age\nAmong signifcant associations")
        ax.set_xlabel("Effect size at 55")
        ax.set_ylabel("Effect size at 70")
        bbox = {'facecolor': (1,1,1,0.8), 'edgecolor':(0,0,0,0)}
        ax.annotate("Age 55 Effect Larger", xy=(0.8*lim,0), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Age 55 Effect Larger", xy=(-0.8*lim,0), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Age 70 Effect Larger", xy=(0,0.8*lim), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Age 70 Effect Larger", xy=(0,-0.8*lim), ha="center", bbox=bbox, zorder=3)
        ax.annotate("Equal Effects", xy=(0.8*lim,0.8*lim), ha="center", va="center", bbox=bbox, zorder=3, rotation=45)
        ax.annotate("Opposite Effects", xy=(0.8*lim,-0.8*lim), ha="center", va="center", bbox=bbox, zorder=3, rotation=-45)
    if legend:
        legend_elts = [matplotlib.lines.Line2D(
                                [0],[0],
                                marker="o", markerfacecolor=c, markersize=10,
                                label=truncate(cat, 35) if not pandas.isna(cat) else "NA",
                                c=c, lw=0)
                            for cat, c in colormap.items()]
        if just_ax == True:
            ax.legend(handles=legend_elts, ncol=2, fontsize="small", loc="upper left")
        else:
            fig.legend(handles=legend_elts, ncol=2, fontsize="small", loc="upper left")
    return fig,ax

fig, ax = age_effect_plot(d)
fig.savefig(f"{OUTDIR}/age_effects.png")

fig, ax = age_effect_plot(d[d.phecode_category == 'mental disorders'], labels=False, color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/age_effects.mental_disorders.png")
fig, ax = age_effect_plot(d[d.phecode_category == 'circulatory system'], labels=False, color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/age_effects.circulatory.png")
fig, ax = age_effect_plot(d[d.phecode_category == 'endocrine/metabolic'], labels=False, color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/age_effects.endorcine.png")
fig, ax = age_effect_plot(d[d.phecode_category == 'genitourinary'], labels=False, color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/age_effects.genitourinary.png")
fig, ax = age_effect_plot(d[d.phecode_category == 'respiratory'], labels=False, color_by="phecode_meaning")
fig.savefig(f"{OUTDIR}/age_effects.respiratory.png")

#Make 2x2 grid of age effect plots
fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
ij = [[0,0], [0,1], [1,0], [1,1]]
SUBCATEGORIES = ["circulatory system", "mental disorders", "endocrine/metabolic", "respiratory"]
for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
    tests = d[d.phecode_category == cat]
    age_effect_plot(tests.sample(frac=1), color_by="phecode_meaning", ax=ax, legend=True, labels=False, cmap="tab20_r")
    ax.set_title(cat)
    if j == 0:
        ax.set_ylabel("Effect size at 70")
    if i == 1:
        ax.set_xlabel("Effect size at 55")
fig.tight_layout()
fig.savefig(OUTDIR+"age_effects.2x2.png")


## age effect for quantitative traits
dage = quantitative_tests.copy()
dage['age_55_effect'] = dage["std_effect"] + dage['std_age_effect'] * young_offset
dage['age_75_effect'] = dage["std_effect"] + dage['std_age_effect'] * old_offset
dage['p_overall'] = dage.p
dage['p_age'] = dage.age_difference_p
fig, ax = age_effect_plot(dage.sample(frac=1), color_by="Functional Category", cmap=color_by_quantitative_function, lim=0.3)
fig.savefig(f"{OUTDIR}/age_effects.quantitative.png")

#Make 2x2 grid of quantitative age differences
fig, axes = pylab.subplots(ncols=2, nrows=2, figsize=(11,11))
ij = [[0,0], [0,1], [1,0], [1,1]]
SUBCATEGORIES = ["Metabolism", "Lipoprotein Profile", "Cardiovascular Function", "Renal Function"]
for cat, ax, (i,j) in zip(SUBCATEGORIES, axes.flatten(), ij):
    tests = dage[dage['Functional Category'] == cat]
    age_effect_plot(tests.sample(frac=1), color_by="phenotype", lim=0.25, ax=ax, legend=True, labels=False, cmap="tab20_r")
    ax.set_title(cat)
    if j == 0:
        ax.set_ylabel("Effect size at 70")
    if i == 1:
        ax.set_xlabel("Effect size at 55")
fig.tight_layout()
fig.savefig(OUTDIR+"age_effects.quantitative.2x2.png")



######## PHENOTYPE-SPECIFIC PLOTS

# Plot config by variables
plot_config = {
    "acceleration_RA": {
        "xbottom": 0.6,
        "xtop": 1.0,
        "point_width": 0.01,
        "bandwidth": 0.15,
        "label": "RA",
    },
    "amplitude": {
        "xbottom": 0.1,
        "xtop": 0.9,
        "point_width": 0.01,
        "bandwidth": 0.25,
        "label": "Amplitude",
    },
}

# Fancy style plot
# Only really works for highly abundant phenotypes like hypertension (401)
def fancy_case_control_plot(data, code, var="acceleration_RA", normalize=False, confidence_interval=False, rescale=True, annotate=False):
    CONTROL_COLOR = "teal"
    CASE_COLOR = "orange"
    UNCERTAIN_COLOR = (0.8, 0.8, 0.8)

    case = data[code] == True
    config = plot_config[var]
    xbottom = config['xbottom']
    xtop = config['xtop']
    point_width = config['point_width']
    bandwidth = config['bandwidth']
    if numpy.sum(case) < 200:
        # Small numbers of cases need to be averaged across a wider band
        bandwidth *= 2
    eval_x = numpy.linspace(xbottom, xtop, int(0.5/point_width + 1))

    case_scaling = (case).sum() * point_width if rescale else 1
    control_scaling = (~case).sum() * point_width if rescale else 1
    case_avg = data[var][case].mean()
    control_avg = data[var][~case].mean()
    total_incidence = case.sum()/len(case)

    def densities_and_incidence(data):
        case_density = scipy.stats.gaussian_kde(data[var][case], bw_method=bandwidth)(eval_x) * case_scaling
        control_density = scipy.stats.gaussian_kde(data[var][~case], bw_method=bandwidth)(eval_x) * control_scaling
        incidence = case_density / (control_density  + case_density)
        return case_density, control_density, incidence
    
    case_density, control_density, incidence = densities_and_incidence(data)

    if confidence_interval:
        N = 40
        incidences = numpy.empty((len(eval_x), N))
        for i in range(N):
            sample = data.sample(len(data), replace=True)
            _, _, incidence = densities_and_incidence(sample)
            incidences[:,i] = incidence
        incidences = numpy.sort(incidences, axis=1)
        lower_bound = incidences[:,0]
        upper_bound = incidences[:,-1]
        middle = incidences[:,incidences.shape[1]//2]

    fig, (ax1,ax3,ax2) = pylab.subplots(nrows=3, sharex=True,
                                        gridspec_kw = {"hspace":0.1,
                                                       "height_ratios":[0.2,0.2,0.6]})

    # Plot the data
    ax1.fill_between(eval_x, 0, control_density, color=CONTROL_COLOR)
    ax3.fill_between(eval_x, 0, case_density, color=CASE_COLOR)
    if confidence_interval:
        ax2.fill_between(eval_x, lower_bound, middle, color='lightgray')
        ax2.fill_between(eval_x, middle, upper_bound, color='lightgray')
    ax2.plot(eval_x, middle, color='k')

    # Plot avgs
    ax1.axvline(control_avg, c='k', linestyle="--")
    ax3.axvline(case_avg, c='k', linestyle="--")
    ax2.axhline(total_incidence, c='k', linestyle="--")

    # Label plot
    ax1.set_ylabel(f"Controls\nN={(~case).sum()}")
    ax2.set_ylabel(f"Prevalence\n(overall={total_incidence:0.1%})")
    ax3.set_ylabel(f"Cases\nN={case.sum()}") 
    ax2.set_xlabel(config['label'])

    ax1.spines['left'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.tick_params(bottom=False)
    ax3.tick_params(bottom=False)
    ax1.yaxis.set_ticks([])
    #ax2.xaxis.set_ticks_position('none')
    ax2.yaxis.set_ticks_position('right')
    if not normalize:
        ax2.yaxis.set_ticks([0, 0.25, 0.5, 0.75, 1])
        ax2.yaxis.set_ticklabels(["0%", "25%", "50%", "75%","100%"])
    else:
        ax2.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter(xmax=1))
    ax3.spines['left'].set_visible(False)
    ax3.spines['right'].set_visible(False)
    ax3.spines['bottom'].set_visible(False)
    ax3.yaxis.set_ticks([])

    # Set axis limits
    ax1.set_xlim(xbottom, xtop)
    if not normalize:
        max_density = max(numpy.max(case_density), numpy.max(control_density))
        ax1.set_ylim(0, max_density)
        ax3.set_ylim(0, max_density)
        ax2.set_ylim(0, 1)
    else:
        ax1.set_ylim(0)
        ax3.set_ylim(0)
        ax2.set_ylim(0, numpy.minimum(numpy.max(middle)*1.3, 1.0))
    ax3.invert_yaxis()

    if annotate:
        ax1.annotate("Control mean",
                        xy=(control_avg, numpy.max(control_density)/2),
                        xytext=(-50,0),
                        textcoords="offset pixels",
                        ha="right",
                        va="center",
                        arrowprops={"arrowstyle": "->"})
        ax3.annotate("Case mean",
                        xy=(case_avg, numpy.max(control_density)/2),
                        xytext=(-50,0),
                        textcoords="offset pixels",
                        ha="right",
                        va="center",
                        arrowprops={"arrowstyle": "->"})
        ax2.annotate("Overall prevalence",
                        xy=((xtop*0.9 + xbottom*0.1), total_incidence),
                        xytext=(0,25),
                        textcoords="offset pixels",
                        ha="right",
                        va="center",
                        arrowprops={"arrowstyle": "->"},
                        )
        i = len(eval_x)//5
        ax2.annotate("95% confidence interval",
                        xy=(eval_x[i], upper_bound[i]),
                        xytext=(0,25),
                        textcoords="offset pixels",
                        ha="center",
                        va="bottom",
                        arrowprops={"arrowstyle": "->"},#"-[, lengthB=5.0"},
                        )

    try:
        ax1.set_title(phecode_info.loc[code].phenotype)
    except KeyError:
        ax1.set_title(code)
    return fig

def incidence_rate_by_category(data, code, categories, var="acceleration_RA", normalize=False, confidence_interval=False, rescale=False):
    # Break up individuals by categorical variable (eg: sex, age bins)
    # and plot the incidence rate of the phecode by the variable
    case = data[code] == True

    config = plot_config[var]
    xbottom = config['xbottom']
    xtop = config['xtop']
    point_width = config['point_width']
    bandwidth = config['bandwidth']
    eval_x = numpy.linspace(xbottom, xtop, int(0.5/point_width + 1))

    case_scaling = (case).sum() * point_width if rescale else  point_width
    control_scaling = (~case).sum() * point_width if rescale else point_width
    case_avg = data[var][case].median()
    control_avg = data[var][~case].median()
    total_incidence = case.sum()/len(case)

    def densities_and_incidence(data):
        case_kde = scipy.stats.gaussian_kde(data[var][case], bw_method=bandwidth)
        case_density = case_kde(eval_x) * case_scaling * case_kde.n
        control_kde = scipy.stats.gaussian_kde(data[var][~case], bw_method=bandwidth)
        control_density = control_kde(eval_x) * control_scaling * control_kde.n
        if not normalize:
            incidence = case_density / (control_density  + case_density)
        if normalize:
            incidence = case_density / total_incidence / 2 / (control_density + case_density / total_incidence / 2)
        return case_density, control_density, incidence

    lower_bounds = []
    upper_bounds = []
    middle_values = []
    all_categories = data[categories].cat.categories
    for value in all_categories:
        category_data = data[data[categories] == value]

        if confidence_interval:
            N = 40
        else:
            N = 1
        incidences = numpy.empty((len(eval_x), N))
        for i in range(N):
            sample = category_data.sample(len(category_data), replace=True)
            _, _, incidence = densities_and_incidence(sample)
            incidences[:,i] = incidence
        incidences = numpy.sort(incidences, axis=1)
        lower_bound = incidences[:,0]
        upper_bound = incidences[:,-1]
        middle = incidences[:,incidences.shape[1]//2]


        lower_bounds.append(lower_bound)
        upper_bounds.append(upper_bound)
        middle_values.append(middle)

    fig, ax = pylab.subplots()

    # Plot the data
    for lower_bound, upper_bound, middle, cat in zip(lower_bounds, upper_bounds, middle_values, all_categories):
        ax.fill_between(eval_x, lower_bound, upper_bound, color='lightgrey', alpha=0.3, label=None)
        ax.plot(eval_x, middle, label=cat)

    # Plot avgs
    ax.axhline(total_incidence, c='k', linestyle="--")

    # Label plot
    ax.set_ylabel(f"Prevalence\n(overall={total_incidence:0.1%})")
    ax.set_xlabel(config['label'])

    ax.yaxis.set_ticks_position('right')
    ax.yaxis.set_ticks([0, 0.25, 0.5, 0.75, 1])
    ax.yaxis.set_ticklabels(["0%", "25%", "50%", "75%","100%"])

    # Set axis limits
    ax.set_xlim(xbottom, xtop)
    ax.set_ylim(0, 1)
    try:
        ax.set_title(phecode_info.loc[code].phenotype + ("\n(normalized)" if normalize else ""))
    except KeyError:
        ax.set_title(code)

    fig.legend()
    return fig

# By-age plots
def age_plot(data, var, phecode, difference=False):
    CONTROL_COLOR = "teal"
    CASE_COLOR = "orange"
    fig, ax = pylab.subplots()
    age_at_actigraphy = (data.actigraphy_start_date - birth_year) / pandas.to_timedelta("1Y")
    eval_x = numpy.arange(numpy.floor(numpy.min(age_at_actigraphy)), numpy.ceil(numpy.max(age_at_actigraphy))+1, 3)
    cases = (data[phecode] == 1)
    controls = (data[phecode] == 0)
    def reg_with_conf_interval(subset):
        main = local_regression(age_at_actigraphy.iloc[subset], data.iloc[subset][var], eval_x, bw=3.0)
        samples = []
        for i in range(40):
            s = numpy.random.choice(subset, size=len(subset))
            samples.append(local_regression(age_at_actigraphy.iloc[s], data.iloc[s][var], eval_x, bw=3.0))
        samples = numpy.array(samples)
        samples = numpy.sort(samples, axis=0)
        bottom = samples[0,:]
        top = samples[-1,:]
        return main, bottom, top
    case_mid, case_bottom, case_top = reg_with_conf_interval(numpy.where(cases)[0])
    control_mid, control_bottom, control_top = reg_with_conf_interval(numpy.where(controls)[0])
    if difference == False:
        ax.plot(eval_x, case_mid, label="cases", c=CASE_COLOR)
        ax.plot(eval_x, control_mid, label="controls", c=CONTROL_COLOR)
        ax.fill_between(eval_x, case_bottom, case_top, color=CASE_COLOR, alpha=0.5)
        ax.fill_between(eval_x, control_bottom, control_top, color=CONTROL_COLOR, alpha=0.5)
    else:
        ax.plot(eval_x, case_mid - control_mid, c="k")
        ax.fill_between(eval_x, case_bottom - control_mid, case_top - case_bottom, color="k", alpha=0.5)
    ax.set_xlabel("Age")
    ax.set_ylabel(var)
    ax.set_title(phecode_info.loc[phecode].phenotype)
    fig.legend()
    return fig, ax

# Hypertension
fig = fancy_case_control_plot(data, 401, normalize=False, confidence_interval=True, annotate=True)
fig.savefig(OUTDIR+"phenotypes.hypertension.png")

fig = incidence_rate_by_category(data, 401, categories="birth_year_category", confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.hypertension.by_age.png")

sns.lmplot("acceleration_RA", "cholesterol", data=data, hue="birth_year_category", markers='.')
pylab.gcf().savefig(OUTDIR+"phenotypes.LDL.png")

sns.lmplot("acceleration_RA", "hdl_cholesterol", data=data, hue="birth_year_category", markers='.')
pylab.gcf().savefig(OUTDIR+"phenotypes.HDL.png")

sns.lmplot("acceleration_RA", "systolic_blood_pressure_V0", data=data, hue="birth_year_category", markers='.')
pylab.gcf().savefig(OUTDIR+"phenotypes.systolic_blood_pressure.png")

sns.lmplot("acceleration_RA", "diastolic_blood_pressure_V0", data=data, hue="birth_year_category", markers='.')
pylab.gcf().savefig(OUTDIR+"phenotypes.diastolic_blood_pressure.png")

fig, ax = age_plot(data, "amplitude", 401)
fig.savefig(OUTDIR+"phenotypes.hypertension.amplitude_age_effect.png")

fig, ax = age_plot(data, "cosinor_rsquared", 401)
fig.savefig(OUTDIR+"phenotypes.hypertension.cosinor_rsquared_age_effect.png")

# Diabetes
fig = fancy_case_control_plot(data, 250, normalize=False, confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.diabetes.png")

fig = incidence_rate_by_category(data, 250, categories="birth_year_category", confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.diabetes.by_age.png")

sns.lmplot("acceleration_RA", "glycated_heamoglobin", data=data, hue="birth_year_category", markers='.')
pylab.gcf().savefig(OUTDIR+"phenotypes.glycated_heamoglobin.png")

percent_diabetes_with_hypertension = (data[401].astype(bool) & data[250].astype(bool)).mean() / data[250].mean()
print(f"Percentage of participants with diabetes that also have hypertension: {percent_diabetes_with_hypertension:0.2%}")

# Mood disorders
#TODO!!
#fig = fancy_case_control_plot(data, 250, normalize=False, confidence_interval=True)
#fig.savefig(OUTDIR+"phenotypes.diabetes.png")

# Some more phenotypes:
fig = fancy_case_control_plot(data, 585, normalize=True, confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.renal_failure.png")
fig = fancy_case_control_plot(data, 276, normalize=True, confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.disorders_fuild_electrolyte_etc.png")
fig = fancy_case_control_plot(data, 290, normalize=True, confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.delirium_dementia_alzheimers.png")
fig = fancy_case_control_plot(data, 332, normalize=True, confidence_interval=True)
fig.savefig(OUTDIR+"phenotypes.parkinsons.png")

#### Age trajectories plot
#TODO: finalize these plots - are they of any value?
fig, (ax1, ax2) = pylab.subplots(nrows=2)
age_at_actigraphy = (data.actigraphy_start_date - birth_year) / pandas.to_timedelta("1Y")
eval_x = numpy.arange(numpy.floor(numpy.min(age_at_actigraphy)), numpy.ceil(numpy.max(age_at_actigraphy))+1)
for _, row in phecode_tests.sort_values(by="p").head(300).iterrows():
    if row.N_cases < 1500:
        continue
    cases = (data[row.phecode] == 1)
    controls = (data[row.phecode] == 0)
    bandwidth = 2
    case_value = local_regression(age_at_actigraphy[cases], data[cases][row.activity_var], eval_x, bw=bandwidth)
    control_value = local_regression(age_at_actigraphy[controls], data[controls][row.activity_var], eval_x, bw=bandwidth)
    midpoint = len(eval_x)//2
    if case_value[midpoint] - control_value[midpoint] > 0:
        ax = ax1
    else:
        ax = ax2
    ax.plot(eval_x, (case_value - control_value) / data[row.activity_var].std(), c="k", alpha=0.2)
ax.set_xlabel("Age")
ax.set_ylabel("Case-control difference in standard deviations")
fig.savefig(OUTDIR+"age_trajectories.png")

### TIMELINE
# Make a timeline of the study design timing so that readers can easily see when data was collected
ACTIGRAPHY_COLOR = "#1b998b"
REPEAT_COLOR = "#c5d86d"
DIAGNOSIS_COLOR = "#f46036"
ASSESSMENT_COLOR = "#aaaaaa"
DEATH_COLOR = "#333333"
fig, (ax1, ax2, ax3) = pylab.subplots(figsize=(8,6), nrows=3)
#ax2.yaxis.set_inverted(True)
ax1.yaxis.set_label_text("Participants per month")
ax2.yaxis.set_label_text("Diagnoses per month")
ax3.yaxis.set_label_text("Deaths per month")
#ax2.xaxis.tick_top()
ax1.spines["top"].set_visible(False)
ax1.spines["right"].set_visible(False)
ax2.spines["top"].set_visible(False)
ax2.spines["right"].set_visible(False)
ax3.spines["top"].set_visible(False)
ax3.spines["right"].set_visible(False)
bins = pandas.date_range("2000-1-1", "2020-6-1", freq="1M")
def date_hist(ax, values, bins, **kwargs):
    # Default histogram for dates doesn't cooperate with being given a list of bins
    # since the list of bins doesn't get converted to the same numerical values as the values themselves
    counts, edges = numpy.histogram(values, bins)
    # Fudge factor fills in odd gaps between boxes
    ax.bar(edges[:-1], counts, width=(edges[1:]-edges[:-1])*1.05, **kwargs)
assessment_time = pandas.to_datetime(data.blood_sample_time_collected_V0)
actigraphy_time = pandas.to_datetime(activity_summary.loc[data.index, 'file-startTime'])
actigraphy_seasonal_time = pandas.to_datetime(activity_summary_seasonal.loc[activity_summary_seasonal.ID.isin(data.index), 'file-startTime'], cache=False)
death_time = pandas.to_datetime(data[~data.date_of_death.isna()].date_of_death)
diagnosis_time = pandas.to_datetime(icd10_entries[icd10_entries.ID.isin(data.index)].first_date)
date_hist(ax1, assessment_time, color=ASSESSMENT_COLOR, label="assessment", bins=bins)
date_hist(ax1, actigraphy_time, color=ACTIGRAPHY_COLOR, label="actigraphy", bins=bins)
date_hist(ax1, actigraphy_seasonal_time, color=REPEAT_COLOR, label="repeat actigraphy", bins=bins)
date_hist(ax2, diagnosis_time, color=DIAGNOSIS_COLOR, label="Diagnoses", bins=bins)
date_hist(ax3, death_time, color=DEATH_COLOR, label="Diagnoses", bins=bins)
ax1.annotate("Assessment", (assessment_time.mean(), 1250), ha="center")
ax1.annotate("Actigraphy", (actigraphy_time.mean(), 1250), ha="center")
ax1.annotate("Repeat\nActigraphy", (actigraphy_seasonal_time.mean(), 1250), ha="center")
ax2.annotate("Medical Record\nDiagnoses", (diagnosis_time.mean(), 1500), ha="center")
ax3.annotate("Deaths", (death_time.mean(), 20), ha="center")
fig.savefig(OUTDIR+"summary_timeline.png")

time_difference = (actigraphy_time - assessment_time).mean()
print(f"Mean difference between actigraphy time and initial assessment time: {time_difference/pandas.to_timedelta('1Y')} years")


### Diagnosis summary
num_diagnoses = icd10_entries.groupby(pandas.Categorical(icd10_entries.ID, categories=data.index)).size()
icd10_entries_at_actigraphy = icd10_entries[pandas.to_datetime(icd10_entries.first_date) < pandas.to_datetime(icd10_entries.ID.map(activity_summary['file-startTime']))]
num_diagnoses_at_actigraphy = icd10_entries_at_actigraphy.groupby(pandas.Categorical(icd10_entries_at_actigraphy.ID, categories=data.index)).size()
icd10_entries_at_assessment = icd10_entries[pandas.to_datetime(icd10_entries.first_date) < pandas.to_datetime(icd10_entries.ID.map(data['blood_sample_time_collected_V0']))]
num_diagnoses_at_assessment = icd10_entries_at_assessment.groupby(pandas.Categorical(icd10_entries_at_assessment.ID, categories=data.index)).size()
ID_without_actigraphy = ukbb.index[ukbb.actigraphy_file.isna()]
icd10_entries_without_actigraphy = icd10_entries_all[icd10_entries_all.ID.isin(ID_without_actigraphy)]
num_diagnoses_no_actigraphy = icd10_entries_without_actigraphy.groupby(pandas.Categorical(icd10_entries_without_actigraphy.ID, categories=ID_without_actigraphy)).size()
fig,ax = pylab.subplots()
ax.boxplot([num_diagnoses_at_assessment, num_diagnoses_at_actigraphy, num_diagnoses, num_diagnoses_no_actigraphy], showfliers=False)
ax.set_xticklabels(["At Assessment", "At Actigraphy", "Final", "Without Actigraphy\nFinal"])
ax.set_ylabel("Medical Record Diagnoses per Participant")
ax.set_title("Disease Burden")
fig.savefig(OUTDIR+"summary_disease_burden.png")

print(f"Median number of diagnoses by category:")
print("At assessment:", num_diagnoses_at_assessment.describe())
print("At actigraphy:", num_diagnoses_at_actigraphy.describe())
print("Final:", num_diagnoses.describe())
print("Final without actigraphy:", num_diagnoses_no_actigraphy.describe())


### Death data
# Survival curves
start_date = pandas.to_datetime(data.date_of_death).min()
def survival_curve(data, ax, **kwargs):
    N = len(data)
    data = data[~data.date_of_death.isna()]
    date = pandas.to_datetime(data.date_of_death).sort_values()
    date_ = pandas.concat((pandas.Series([start_date]), date))
    ax.step(date_,
            (1 - numpy.concatenate(([0], numpy.arange(len(data))))/N)*100,
            where='post',
            **kwargs)

RA_quintiles = pandas.cut(data.acceleration_RA,
                          data.acceleration_RA.quantile([0,0.2,0.4,0.6,0.8,1.0]))
quintile_labels = ["First", "Second", "Third", "Fourth", "Fifth"]

def quintile_survival_plot(data, var, var_label=None):
    if var_label is None:
        var_label = var
    quintiles = pandas.qcut(data[var], 5)
    fig, ax = pylab.subplots(figsize=(8,6))
    for quintile, label in list(zip(quintiles.cat.categories, quintile_labels))[::-1]:
        survival_curve(data[quintiles == quintile], ax, label= label + " Quintile")
    fig.legend(loc=(0.15,0.15))
    ax.set_title(f"Survival by {var_label}")
    ax.set_ylabel("Survival Probability")
    ax.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter())
    ax.xaxis.set_major_formatter(matplotlib.dates.DateFormatter("%Y"))
    #ax2 = ax.twinx() # The right-hand side axis label
    #scale = len(data)/5
    #ax2.set_ylim(ax.get_ylim()[0]*scale, ax.get_ylim()[1]*scale)
    #ax2.set_ylabel("Participants")
    fig.tight_layout()
    return fig

def categorical_survival_plot(data, var, var_label=None, min_N=None):
    if var_label is None:
        var_label = var
    fig, ax = pylab.subplots(figsize=(8,6))
    value = data[var].astype("category")
    for cat in value.cat.categories:
        d = data[value == cat]
        if min_N is not None and len(d) < min_N:
            continue # Skip this category
        survival_curve(d, ax, label= cat)
    fig.legend(loc=(0.15,0.15))
    ax.set_ylabel("Survival Probability")
    ax.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter())
    fig.tight_layout()
    return fig

# Survival by RA
fig = quintile_survival_plot(data, "acceleration_RA", "RA")
fig.savefig(OUTDIR+"survival.RA.png")

# Survival by main_sleep_offset_avg
fig = quintile_survival_plot(data, "main_sleep_offset_mean", "Sleep Offset")
fig.savefig(OUTDIR+"survival.main_sleep_offset_mean.png")

# Survival by MVPA_overall_avg
fig = quintile_survival_plot(data, "MVPA_overall", "MVPA Mean")
fig.savefig(OUTDIR+"survival.MVPA_overall.png")

# Survival by MVPA_overall_avg
fig = quintile_survival_plot(data, "MVPA_hourly_SD", "MVPA hourly SD")
fig.savefig(OUTDIR+"survival.MVPA_hourly_SD.png")

# Survival by acceleration_hourly_SD
fig = quintile_survival_plot(data, "acceleration_hourly_SD", "Acceleration Hourly SD")
fig.savefig(OUTDIR+"survival.acceleration_hourly_SD.png")

# Survival by phase
fig, ax = pylab.subplots()
data['phase_adjusted'] = (data.phase - 8) % 24 + 8
fig = quintile_survival_plot(data, "phase_adjusted", "phase")
fig.savefig(OUTDIR+"survival.phase.png")

# Survival by RA and sex
fig, axes = pylab.subplots(ncols=2, sharey=True, sharex=True, figsize=(10,5))
for ax, sex in zip(axes, ["Male", "Female"]):
    for quintile, label in list(zip(RA_quintiles.cat.categories, quintile_labels))[::-1]:
        survival_curve(data[(data.sex == sex) & (RA_quintiles == quintile)], ax,
                        label=("RA " + label + " Quintile" if sex == "Male" else None))
        ax.set_title(sex)
ax1, ax2 = axes
ax1.xaxis.set_major_locator(matplotlib.dates.YearLocator())
ax.xaxis.set_major_formatter(matplotlib.dates.DateFormatter('%Y'))
ax1.set_ylabel("Survival Probability")
ax1.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter())
fig.legend()
fig.savefig(OUTDIR+"survival.RA.by_sex.png")


### Tests Survival
# Cox proportional hazards model
data['date_of_death_censored'] = pandas.to_datetime(data.date_of_death)
data.date_of_death_censored.fillna(data.date_of_death_censored.max(), inplace=True)
data['date_of_death_censored_number'] = (data.date_of_death_censored - data.date_of_death_censored.min()).dt.total_seconds()
uncensored = (~data.date_of_death.isna()).astype(int)
data['age_at_death_censored'] = (pandas.to_datetime(data.date_of_death) - birth_year) / pandas.to_timedelta("1Y")
entry_age = (data.actigraphy_start_date - birth_year) / pandas.to_timedelta("1Y")
data.age_at_death_censored.fillna(data.age_at_death_censored.max(), inplace=True)

if RECOMPUTE:
    covariate_formula = ' + '.join(["BMI", "smoking_ever"])
    survival_tests_data = []
    for var in activity_variables:
        print(var)
        for method in ['newton', 'cg']:# Try two convergence methods - some work for only one method
            try:
                formula = f"age_at_death_censored ~ {var} + sex + {covariate_formula}"
                result = smf.phreg(formula=formula,
                                    data=data,
                                    status=uncensored,
                                    entry=entry_age,
                                    ).fit(method=method)
                print('.')
                interaction_formula = f"age_at_death_censored ~ {var} * sex + {covariate_formula}"
                interaction_result = smf.phreg(formula=interaction_formula,
                                    data=data,
                                    status=uncensored,
                                    entry=entry_age,
                                    ).fit(method=method)
            except numpy.linalg.LinAlgError:
                print(f"MLE fit method {method} failed, trying alternative")
                continue
            break
        pvalues = pandas.Series(result.pvalues, index=result.model.exog_names)
        params = pandas.Series(result.params, index=result.model.exog_names)
        interaction_pvalues = pandas.Series(interaction_result.pvalues, index=interaction_result.model.exog_names)
        interaction_params = pandas.Series(interaction_result.params, index=interaction_result.model.exog_names)
        survival_tests_data.append({
            "activity_var": var,
            "p": pvalues[var],
            "log Hazard Ratio": params[var],
            "standardized log Hazard Ratio": params[var] * data[var].std(),
            "sex_difference_p": interaction_pvalues[f"{var}:sex[T.Male]"],
            "male_logHR": interaction_params[f"{var}:sex[T.Male]"] + interaction_params[f"{var}"],
            "female_logHR": interaction_params[f"{var}"],
        })
    survival_tests = pandas.DataFrame(survival_tests_data)
    survival_tests['q'] = BH_FDR(survival_tests.p)
    survival_tests['sex_difference_q'] = BH_FDR(survival_tests.sex_difference_p)
    survival_tests = pandas.merge(survival_tests, activity_variable_descriptions[["Category", "Subcategory", "Units"]], left_on="activity_var", right_index=True)
    survival_tests.to_csv(OUTDIR+"survival.by_activity_variable.txt", sep="\t", index=False)
else:
    survival_tests = pandas.read_csv(OUTDIR+"survival.by_activity_variable.txt", sep="\t")
survival_tests_raw = survival_tests.copy()

###  Connecting actigraphy and quantitative traits
Q_CUTOFF = 1e-9
significant_phecode_tests = phecode_tests[phecode_tests.q <= Q_CUTOFF]
significant_quantitative_tests = quantitative_tests[quantitative_tests.q <= Q_CUTOFF]
both_significant = significant_phecode_tests[["phecode", "activity_var"]].set_index("activity_var").join(significant_quantitative_tests[["phenotype", "activity_var"]].set_index("activity_var")).reset_index()
# drop bogus 'BMI' correlation, since included in the covariates, as well as 'nan'
both_significant = both_significant.query("phenotype != 'BMI' and phenotype == phenotype")

# Plot all the connected actigraphy <-> quantitative traits
import itertools
colormap = matplotlib.cm.get_cmap("Dark2").colors
def colormap_extended(colormap):
    i = 1
    def modify(c):
        return 1-(1-c)/i**0.7
    while True:
        yield from ((modify(c[0]), modify(c[1]), modify(c[2])) for c in colormap)
        i += 1
colormap = colormap_extended(colormap)
activity_var_colors = {var:color for var,color in zip(both_significant['activity_var'].unique(), colormap)}
max_N = both_significant.groupby(["phenotype", "phecode"]).count().max()
max_in_row = 4
def split_label(label, max_chars):
    #Split on whitespace then combine, inserting newlines whenever too many characters are reached
    words = label.split()
    parts = []
    line = []
    for word in words:
        line.append(word)
        if sum(len(x) for x in line) >= max_chars:
            # End line, too long
            parts.append(' '.join(line))
            line = []
    parts.append(' '.join(line))
    return '\n'.join(parts).strip()
def plot_interconnections(data):
    selected_phecodes = data.phecode.unique()
    selected_traits = data.phenotype.unique()
    fig_width = len(selected_phecodes)*0.6 + 3
    fig, ax = pylab.subplots(figsize=(fig_width,8))
    for i, phecode in enumerate(selected_phecodes):
        for j, trait_var in enumerate(selected_traits):
            d = data[(data.phecode == phecode)
                                    & (data.phenotype == trait_var)]
            for k, activity_var in enumerate(d.activity_var):
                x = i + (k % max_in_row  - 0.5 * min(len(d)-1, max_in_row-1)) / (max_in_row + 1)
                y = j + (k // max_in_row) / (max_N // max_in_row + 1)
                ax.plot([x], [y], "o", c=activity_var_colors[activity_var], label=activity_var)
    ax.set_xticks(numpy.arange(len(selected_phecodes)))
    ax.set_xticklabels(phecode_info.loc[selected_phecodes].phenotype.apply(lambda x: split_label(x,15)),
                            rotation=90, ha="center")
    ax.set_yticks(numpy.arange(len(selected_traits)))
    ax.set_yticklabels(selected_traits)
    fig.tight_layout()
    return fig, ax
mental_health_phecodes = [327, 296, 300]
fig, ax = plot_interconnections(both_significant[~both_significant.phecode.isin(mental_health_phecodes)])
fig.savefig(OUTDIR+"interconnections.general.png")
fig, ax = plot_interconnections(both_significant[both_significant.phecode.isin(mental_health_phecodes)])
fig.savefig(OUTDIR+"interconnections.mental_health.png")


# Draw a legend
fig, ax = pylab.subplots()
ax.set_visible(False)
fig.legend([matplotlib.lines.Line2D([0],[0], linestyle='', marker="o", color=activity_var_colors[var]) for var in activity_var_colors.keys()],
    activity_var_colors.keys(),
    loc="center",
    #bbox_to_anchor=(1.1,1),
    ncol=2)
fig.tight_layout()
fig.savefig(OUTDIR+"interconnections.legend.png")



### NETWORK ANALYSIS

def weight(p):
    if p == 0:
        return 100
    else:
        return -numpy.log10(p)
P_VALUE_CUTOFF = 1e-10
g_all = networkx.Graph()
g_all.add_nodes_from(activity_variables, type="activity_var")
g_all.add_nodes_from(phecode_tests.phecode_meaning.unique(), type="phecode")
g_all.add_nodes_from(quantitative_tests.phenotype.unique(), type="trait")
g_all.add_edges_from((row['activity_var'], row.phecode_meaning, {"p": row.p, "weight": weight(row.p)})
                    for _, row in phecode_tests.iterrows() if row.p < P_VALUE_CUTOFF)
g_all.add_edges_from((row.activity_var, row.phenotype, {"p": row.p, "weight": weight(row.p)})
                    for _, row in quantitative_tests.iterrows() if row.p < P_VALUE_CUTOFF)
if 'BMI' in g_all:
    g_all.remove_node('BMI') # Bogus
# Remove nodes without any edges
g = g_all.edge_subgraph(g_all.edges)

node_color_dict = {'phecode': '#ef476f', 'activity_var': '#118ab2', 'trait': '#06d6a0'}
node_colors = [node_color_dict[node['type']] for node in g.nodes.values()]
fig, ax = pylab.subplots(figsize=(12,10))
networkx.draw(g, with_labels=True,
                  node_color=node_colors,
                  edge_color="#aaaaaa")
fig.savefig(OUTDIR+"network.png")

def network_to_table(g, output_prefix):
    edge_table = pandas.concat([
            pandas.DataFrame(g.edges).rename(columns={0:"node1", 1:"node2"}),
            pandas.DataFrame(g.edges.values()),
        ],
        axis = 1)
    node_table = pandas.concat([
            pandas.DataFrame(g.nodes).rename(columns={0:"node"}),
            pandas.DataFrame(g.nodes.values()),
        ],
        axis = 1)
    node_table = pandas.merge(
        node_table,
        activity_variable_descriptions[["Category", "Subcategory", "Units"]],
        left_on="node",
        right_index=True,
        how="left")
    node_table = pandas.merge(
        node_table,
        numpy.log10(survival_tests.set_index('activity_var')[["p"]]).rename(columns=lambda x: "survival_log10" + x),
        left_on="node",
        right_on="activity_var",
        how="left")
    edge_table.to_csv(OUTDIR+output_prefix+".edges.txt", index=False, sep="\t")
    node_table.to_csv(OUTDIR+output_prefix+".nodes.txt", index=False, sep="\t")
network_to_table(g, "network")


g_min = networkx.empty_graph([n for n in g.nodes if n in activity_variables])
def weights_between(n1, n2):
    common_neighbors = set(g.adj[n1]).intersection(g.adj[n2])
    return sum(g.adj[n1][nbr]['weight'] * g.adj[n2][nbr]['weight']
                for nbr in common_neighbors)
g_min.add_edges_from((n1, n2, {'weight': weights_between(n1, n2)})
                        for n1 in g_min.nodes
                        for n2 in g_min.nodes)
g_min.remove_edges_from(edge for edge, data in g_min.edges.items() if data['weight'] < 100)
fig, ax = pylab.subplots(figsize=(12,10))
networkx.draw(g_min, with_labels=True, edge_color="#aaaaaa")
#fig.savefig(OUTDIR+"network.just_activity_vars.png")

# graphs by just the activity variable type
fig, axes = pylab.subplots(nrows=2,ncols=3, figsize=(15,15))
for ax, cat in zip(axes.flatten(), activity_variable_descriptions.Category.unique()):
    category = activity_variable_descriptions.index[activity_variable_descriptions.Category == cat]
    g_cat = g.edge_subgraph([e for e in g.edges
                                if (e[0] in category or e[1] in category)])
    networkx.draw(g_cat,
                    ax=ax,
                    with_labels=True,
                    edge_color="#aaaaaa",
                    node_color=[node_color_dict[g_cat.nodes[n]['type']] for n in g_cat.nodes])
    ax.set_title(cat)



### Plot survival assocations versus inter/intra personal variance for validation
fig, ax = pylab.subplots(figsize=(8,6))
colorby = survival_tests.activity_var.map(activity_variable_descriptions.Category)
color = colorby.map(color_by_actigraphy_cat)
def get_variance_ratio(var):
    if var.endswith("_abs_dev"):
        var = var[:-8]
    try:
        return activity_variance.corrected_intra_personal_normalized.loc[var]
    except KeyError:
        print(var)
        return float("NaN")
variance_ratio = survival_tests.activity_var.apply(get_variance_ratio)
variance_ratio.index = survival_tests.activity_var
ax.scatter(#-numpy.log10(survival_tests.p),
            survival_tests['standardized log Hazard Ratio'],
            variance_ratio,
            s=1-numpy.log10(survival_tests.p)*3,
            c=color)
ax.set_xlabel("Standardized log Hazard Ratio")
ax.set_ylabel("Within-person variation / Between-person variation")
ax.set_ylim(0,1)
ax.axvline(0, c='k')
for indx, row in survival_tests.sort_values(by="p").head(20).iterrows():
    # Label the top points
    ax.annotate(
        row.activity_var,
        (#-numpy.log10(row.p),
         row['standardized log Hazard Ratio'],
         variance_ratio.loc[row.activity_var]),
        xytext=(0,15),
        textcoords="offset pixels",
        arrowprops={'arrowstyle':"->"})
legend_elts = [matplotlib.lines.Line2D(
                        [0],[0],
                        marker="o", markerfacecolor=c, markersize=10,
                        label=cat if not pandas.isna(cat) else "NA",
                        c=c, lw=0)
                    for cat, c in color_by_actigraphy_cat.items()]
fig.legend(handles=legend_elts, ncol=2, fontsize="small")
#fig.tight_layout()
fig.savefig(OUTDIR+"survival_versus_variation.svg")

### Correlation plot of top survival variables
top_survival_vars = survival_tests.sort_values(by="p").activity_var.head(20)
def matrix_plot(data, **kwargs):
    fig, ax = pylab.subplots(figsize=(10,10))
    h = ax.imshow(data, **kwargs)
    ax.set_xticks(range(len(data.columns)))
    ax.set_xticklabels(data.columns, rotation=90)
    ax.set_xlim(-0.5, len(data.index)-0.5)
    ax.set_yticks(range(len(data.index)))
    ax.set_yticklabels(data.index)
    ax.set_ylim(-0.5, len(data.index)-0.5)
    c = fig.colorbar(h)
    fig.tight_layout()
    return fig, ax
fig, ax = matrix_plot(data[top_survival_vars].corr(), vmin=-1, vmax=1, cmap="bwr")
fig.savefig(OUTDIR+"correlation.top_survival_activity_vars.png")


### Assess what variables add to acceleration_RA the most
if RECOMPUTE:
    beyond_RA_tests_list = []
    for var in activity_variables:
        if var == "acceleration_RA":
            continue
        formula = f"date_of_death_censored_number ~ birth_year + sex + BMI + smoking + acceleration_RA + {var}"
        try:
            results = smf.phreg(formula=formula,
                                data=data,
                                status=uncensored,
                                ).fit()
        except numpy.linalg.LinAlgError:
            print(f"Failed regression on {var} - skipping")
            continue
        pvalues = pandas.Series(results.pvalues, index=results.model.exog_names)
        params = pandas.Series(results.params, index=results.model.exog_names)
        beyond_RA_tests_list.append({
            "activity_var": var,
            "p": pvalues[var],
            "p_RA": pvalues["acceleration_RA"],
            "standardized log Hazard Ratio": params[var] * data[var].std(),
            "standardized log Hazard Ratio RA": params['acceleration_RA'] * data['acceleration_RA'].std(),
        })
    beyond_RA_tests = pandas.DataFrame(beyond_RA_tests_list)
    beyond_RA_tests = pandas.merge(beyond_RA_tests, activity_variable_descriptions[["Category", "Subcategory"]],
                            left_on="activity_var",
                            right_index=True)
    beyond_RA_tests.to_csv(OUTDIR+"beyond_RA_tests.txt", sep="\t", index=False)
else:
    beyond_RA_tests = pandas.read_csv(OUTDIR+"beyond_RA_tests.txt", sep="\t")


## Plot the amount RA goes "beyond" other variables
fig, (ax1, ax2) = pylab.subplots(ncols=2, figsize=(10,6))
c = beyond_RA_tests.Subcategory.map(color_by_actigraphy_subcat)
ax1.scatter(
    beyond_RA_tests['standardized log Hazard Ratio'].abs(),
    beyond_RA_tests['standardized log Hazard Ratio RA'].abs(),
    c=c)
ax1.set_ylim(0, ax1.get_ylim()[1])
ax1.set_xlabel("log Hazard Ratio / SD of alternative variable")
ax1.set_ylabel("log Hazard Ratio / SD of RA")
ax1.axhline(survival_tests.loc[survival_tests.activity_var == "acceleration_RA", "standardized log Hazard Ratio"].abs().values,
            linestyle="--", c="k")
ax2.scatter(
    -numpy.log10(beyond_RA_tests.p),
    -numpy.log10(beyond_RA_tests.p_RA),
    c=c)
ax2.set_ylim(0, ax2.get_ylim()[1])
ax2.set_xlabel("-log10 p-value of alternative variable")
ax2.set_ylabel("-log10 p-value of RA")
ax2.axhline(-numpy.log10(survival_tests.loc[survival_tests.activity_var == "acceleration_RA", "p"].values),
            linestyle="--", c="k")
legend_from_colormap(fig, color_by_actigraphy_subcat, loc="upper left", fontsize="small", ncol=2)
fig.savefig(OUTDIR+"additive_benefit_RA.png")

## Plot comparison of circadian versus other variables for the phecodes where circadian do the best
# gather the top circadian phecodes
#circadian_does_best = phecode_tests.sort_values(by="p").groupby("phecode").apply(lambda x: x.iloc[0].activity_var_category == "Circadianness")
#circadian_phecodes = circadian_does_best.index[circadian_does_best]
#circadian_best_tests = phecode_tests[phecode_tests.phecode.isin(circadian_phecodes)].copy()
circadian_best_tests = phecode_tests.copy()
circadian_best_tests = circadian_best_tests[~circadian_best_tests.activity_var.str.startswith("self_report")]
circadian_best_tests['ordering'] = circadian_best_tests.phecode.map(circadian_best_tests.groupby("phecode").p.min().rank())
fig, (ax1, ax2) = pylab.subplots(figsize=(8,9), ncols=2, sharey=True)
yticks = {}
pvalues = []
ranks = []
for  phecode, row in circadian_best_tests.groupby('phecode'):
    rank = row.ordering.iloc[0]
    phenotype = row.phecode_meaning.iloc[0]
    if rank > 20:
        continue # Skip all but the most significant
    color = row.activity_var_category.map(color_by_actigraphy_cat)
    top_circ = row[row.activity_var_category == "Circadianness"].sort_values(by="p").iloc[0].activity_var
    top_physical = row[row.activity_var_category == "Physical activity"].sort_values(by="p").iloc[0].activity_var
    top_sleep = row[row.activity_var_category == "Sleep"].sort_values(by="p").iloc[0].activity_var
    #top_circ = "amplitude"
    #top_sleep = "main_sleep_ratio_mean"
    #top_physical = "acceleration_overall"
    _, circ_fit = compute_phecode_test(top_circ, phecode, data)
    _, physical_fit = compute_phecode_test(top_physical, phecode, data)
    _, sleep_fit = compute_phecode_test(top_sleep, phecode, data)
    circ_conf_int = circ_fit.conf_int().loc[f"Q({phecode})"].abs() / data[top_circ].std()
    physical_conf_int = physical_fit.conf_int().loc[f"Q({phecode})"].abs() / data[top_physical].std()
    sleep_conf_int = sleep_fit.conf_int().loc[f"Q({phecode})"].abs() / data[top_sleep].std()
    #ax1.scatter(-numpy.log10(row.p), row.ordering, c = color)
    ##ax2.scatter(numpy.abs(row.std_effect), row.ordering, c=color)
    #ax2.plot(circ_conf_int, [rank, rank], c = "r")
    #ax2.plot(physical_conf_int, [rank+0.1, rank+0.1], c = "b")
    #ax2.plot(sleep_conf_int, [rank+0.2, rank+0.2], c = "g")

    # Logistic model
    corr = data[[top_circ, top_physical]].corr().values[0,1]
    data['orthog'] = (data[top_circ] - data[top_circ].mean()) /data[top_circ].std() - corr * (data[top_physical] - data[top_physical].mean()) / data[top_physical].std()
    results = smf.logit(f"Q({phecode}) ~ {top_circ} + {top_sleep} + {top_physical} + {covariate_formula}", data=data).fit()
    marginals = results.get_margeff()
    ps = pandas.Series(results.pvalues, index=results.model.exog_names)[[top_circ, top_physical, top_sleep]]
    margeffs = pandas.Series(marginals.margeff, index=results.model.exog_names[1:])[[top_circ, top_physical, top_sleep]].abs()
    margeffs *= data[margeffs.index].std() # Standardize by the actigraphy variables used
    margeffs /= data[phecode].mean() # Standardize by the overall prevalence
    ses = pandas.Series(marginals.margeff_se, index=results.model.exog_names[1:])[[top_circ, top_physical, top_sleep]]
    ses *= data[margeffs.index].std() # Standardized SEs too
    ses /= data[phecode].mean()

    colors = [color_by_actigraphy_cat[c] for c in ["Circadianness", "Physical activity", "Sleep"]]
    ys = numpy.linspace(rank-0.15, rank+0.15, 3)
    #ax1.scatter(-numpy.log10(ps), [rank]*len(ps), c=colors)
    ax2.scatter(margeffs, ys, c=colors)
    for eff, se, y, c, p in zip(margeffs, ses, ys, colors, ps):
        ax1.barh([y], height=0.15, width=[-numpy.log10(p)], color=c)
        ax2.plot([eff - 1.96*se, eff + 1.96*se], [y, y], c=c)

    yticks[rank] = wrap(phenotype, 30)
    pvalues.append(ps[top_circ])
    ranks.append(rank)
    ax2.set_xlim(0, 0.8)
# Compute an FDR = 0.05 cutoff - but we don't actually compute a p for every phenotype
# so we will assume worst case, all others ps are 1
#pvalues = numpy.array(pvalues + [1]*(phecode_tests.phecode.nunique() - len(pvalues)))
pvalues = numpy.concatenate([pvalues, numpy.random.uniform(size=(phecode_tests.phecode.nunique() - len(pvalues)))])
qvalues = BH_FDR(pvalues)
qvalue_dict = {rank: q for rank, q in zip(ranks, qvalues[:len(ranks)])}
# Cut off half way between pvalues of worst gene passing FDR < 0.05 and best gene not passing
pvalue_cutoff = 0.5*pvalues[qvalues < 0.05].max() + 0.5*pvalues[qvalues >= 0.05].min()
ax1.axvline(-numpy.log10(pvalue_cutoff), linestyle="--", color="k")
ax1.set_yticks(list(yticks.keys()))
ax1.set_yticklabels([name if qvalue_dict[rank] >= 0.05 else name + " (*)"
                        for rank, name in yticks.items()])
ax1.set_xlabel("-log10 p-value")
ax1.set_xlim(left=0)
ax2.set_xlabel("Standardized Effect Size")
ax2.axvline(0, linestyle="-", color="k", linewidth=1)
legend_from_colormap(fig, {cat:color_by_actigraphy_cat[cat] for cat in ["Sleep", "Physical activity", "Circadianness"]})
ax1.margins(0.5, 0.02)
fig.tight_layout()
fig.savefig(OUTDIR+"circadian_vs_other_vars.png")


### Assess the correlations of the various measures
base_vars = ["acceleration", "moderate", "walking", "sleep", "sedentary", "tasks_light", "MET", "MVPA"]
additions = ["_overall", "_hourly_SD", "_within_day_SD", "_between_day_SD", "_peak_value_mean", "_RA", "_IV"]
results = {}
for base_var in base_vars:
    def corr(v1,v2):
        try:
            #return numpy.corrcoef(data[v1], data[v2])[0,1]
            return scipy.stats.spearmanr(data[v1], data[v2], nan_policy="omit")[0]
        except Exception as e:
            print(e)
            return float("NaN")
    row = {(a1,a2): corr(base_var + a1, base_var + a2)
            for a1 in additions
            for a2 in additions
            if a1 != a2}
    results[base_var] = row
correlations = pandas.DataFrame(results)

# Investigate using a subset of the activity variables for clarity
# by removing highly correlated variables (> 0.9)
selected_activity_variables = []
correlations = data[activity_variables].corr()
ordered_activity_variables = survival_tests.sort_values(by="p").activity_var
for var in ordered_activity_variables:
    corrs = correlations.loc[var][selected_activity_variables]
    if (corrs.abs() < 0.9).all():
        selected_activity_variables.append(var)
    else:
        print(f"Dropping {var} due to {corrs.abs().idxmax()}")

### Overall disease burden (number of phecodes) versus RA
fig, ax = pylab.subplots()
num_phecodes = data[phecode_groups].sum(axis=1)
phecode_ranges = pandas.cut(num_phecodes, [0,1,2,4,8,16,32,num_phecodes.max()+1])
xticklabels = []
for i, (phecode_range, group) in enumerate(data.groupby(phecode_ranges)):
    ax.boxplot(group.acceleration_RA.values, positions=[i], showfliers=False, widths=0.8)
    if phecode_range.right == phecode_range.left+1:
        xticklabels.append(f"{int(phecode_range.left)}")
    else:
        xticklabels.append(f"{int(phecode_range.left)}-{int(phecode_range.right-1)}")
ax.set_xticks(range(len(xticklabels)))
ax.set_xticklabels(xticklabels)
ax.set_xlabel("Number of unique diagnoses")
ax.set_ylabel("RA")
fig.savefig(OUTDIR+"num_phecodes.RA.png")

### Relative Risks
# Create interpretable risks from the phecode associations
if RECOMPUTE:
    d = phecode_tests[(phecode_tests.q < 0.01)].copy()
    covariate_formula = 'sex + age_at_actigraphy + BMI + smoking_ever + overall_health_good + ethnicity_white + education_College_or_University_degree + high_income'
    #covariate_formula = ' + '.join(covariates)
    risk_quantification_data = []
    for _, row in d.iterrows():
        try:
            results = smf.logit(f"Q({row.phecode}) ~ {row.activity_var} + {covariate_formula}", data=data).fit()
            marginal_effect = results.get_margeff().summary_frame().loc[row.activity_var, 'dy/dx']
            p = results.pvalues[row.activity_var]
        except numpy.linalg.LinAlgError:
            marginal_effect = float("NaN")
            p = float("NaN")
        incidence = data[row.phecode].mean()
        risk_quantification_data.append({
            "activity_var": row.activity_var,
            "phecode": row.phecode,
            "phecode_meaning": row.phecode_meaning,
            "phecode_category": row.phecode_category,
            "incidence": incidence,
            "marginal_effect": marginal_effect,
            "std_marginal_effect": marginal_effect * data[row.activity_var].std(),
            "p": p,
        })
    risk_quantification = pandas.DataFrame(risk_quantification_data)
    risk_quantification.to_csv(OUTDIR+"relative_risks.txt", sep="\t", index=False)
else:
    risk_quantification = pandas.read_csv(OUTDIR+"relative_risks.txt", sep="\t")

## Plot relative risks
fig, ax = pylab.subplots(figsize=(9,9))
colormap = color_by_phecode_cat
color = risk_quantification.phecode_category.map(colormap)
ax.scatter(
    risk_quantification.std_marginal_effect / risk_quantification.incidence,
    -numpy.log10(risk_quantification.p),
    c = color,
    marker="+")
ax.set_xlabel("Standardized Marginal Effect / Prevalence")
ax.set_ylabel("-log10 p-value")
legend_elts = [matplotlib.lines.Line2D(
                        [0],[0],
                        marker="o", markerfacecolor=c, markersize=10,
                        label=cat if not pandas.isna(cat) else "NA",
                        c=c, lw=0)
                    for cat, c in colormap.items()]
fig.legend(handles=legend_elts, ncol=2, fontsize="small")
fig.savefig(OUTDIR+"marginal_effect.volcano_plot.png")

#### Investigate medications
medications = pandas.read_csv("../processed/ukbb_medications.txt", sep="\t")
metformin_code = 1140884600
metformin = (medications.medication_code == metformin_code).groupby(medications.ID).any()
data['metformin'] = (data.index.map(metformin) == True)
print("Metformin analysis:")
print( data.groupby("metformin").phase.describe())
#TODO: metformin analysis in more detail

### Sex-difference survival plot
def sex_difference_survival_plot(d, color_by="activity_var_category"):
    colormap = {cat:color for cat, color in
                        zip(d[color_by].unique(),
                            [pylab.get_cmap("Set3")(i) for i in range(20)])}
    color = [colormap[c] for c in d[color_by]]
    fig, ax = pylab.subplots(figsize=(9,9))
    # The points
    ax.scatter(
        d.std_male_logHR,
        d.std_female_logHR,
        label="phenotypes",
        #s=-numpy.log10(d.p)*10,
        c=color)
    ax.set_title("Survival associations by sex")
    ax.spines['bottom'].set_color(None)
    ax.spines['top'].set_color(None)
    ax.spines['left'].set_color(None)
    ax.spines['right'].set_color(None)
    ax.axvline(c="k", lw=1)
    ax.axhline(c="k", lw=1)
    ax.set_xlabel("log HR in males")
    ax.set_ylabel("log HR in females")
    ax.set_aspect("equal")
    #ax.set_xlim(-0.5,0.5)
    #ax.set_ylim(-0.5,0.5)
    # Diagonal y=x line
    bound = max(abs(numpy.min([ax.get_xlim(), ax.get_ylim()])),
                numpy.max([ax.get_xlim(), ax.get_ylim()]))
    diag = numpy.array([-bound, bound])
    ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
    ax.plot(diag, -diag, linestyle="--", c='k', zorder=-1, label="diagonal")
    bbox = {'facecolor': (1,1,1,0.8), 'edgecolor':(0,0,0,0)}
    #ax.annotate("Male Effect Larger", xy=(0.4,0), ha="center", bbox=bbox, zorder=3)
    #ax.annotate("Male Effect Larger", xy=(-0.4,0), ha="center", bbox=bbox, zorder=3)
    #ax.annotate("Female Effect Larger", xy=(0,0.4), ha="center", bbox=bbox, zorder=3)
    #ax.annotate("Female Effect Larger", xy=(0,-0.25), ha="center", bbox=bbox, zorder=3)
    legend_elts = [matplotlib.lines.Line2D(
                            [0],[0],
                            marker="o", markerfacecolor=c, markersize=10,
                            label=cat if not pandas.isna(cat) else "NA",
                            c=c, lw=0)
                        for cat, c in colormap.items()]
    ax.legend(handles=legend_elts, ncol=2, fontsize="small")
    return fig, ax
d = survival_tests.copy()
activity_var_stds = data[activity_variables].std() #TODO: should we separate male/female stds?
d['std_male_logHR'] = d.activity_var.map(activity_var_stds) * d['male_logHR']
d['std_female_logHR'] = d.activity_var.map(activity_var_stds) * d['female_logHR']
d['activity_var_category'] = d.activity_var.map(activity_variable_descriptions.Subcategory)
fig, ax = sex_difference_survival_plot(d)
fig.savefig(OUTDIR+"survival.by_sex.png")


### Comparisons of self-reported versus objectively derived variables
# Some self-reported variables are closely matched by objectively derived variables
# so which one is best associated with phecodes?

# self_report_sleep_duration versus main_sleep_duration_mean or total_sleep_mean
# Sleep durations have a fairly U-shaped curve so we use the abs_dev versions here
# and the total_sleep_mean variable includes napping time as well as sleep time
# so we don't use that here since napping is investigated in the self_report_nap_during_day variable
variable_pairs = { "sleep duration": ("sleep_duration", "main_sleep_duration_mean"),
                   "sleep duration abs. dev.": ("sleep_duration_abs_dev", "main_sleep_duration_mean_abs_dev"),
                   "phase": ("morning_evening_person", "phase"),
                   #("morning_evening_person", "phase_abs_dev"),
                   "napping": ("nap_during_day", "other_sleep_mean"),
                   "sleeplessness": ("sleeplessness", "WASO_mean")}
#fig, (ax1, ax2) = pylab.subplots(ncols=2, sharey=True)
fig, ax2 = pylab.subplots(figsize=(4,4))
for i, (name, (self_report, activity_var)) in enumerate(variable_pairs.items()):
    if self_report == "sleep_duration_abs_dev":
        self_report_var = "self_report_sleep_duration_abs_dev"
    else:
        self_report_var = "self_report_" + self_report_circadian_variables[self_report]['name']
    print(f"Comparing: {self_report} to {activity_var}")
    self_report_survival = survival_tests.query(f"activity_var == '{self_report_var}'").iloc[0]
    self_report_objective = survival_tests.query(f"activity_var == '{activity_var}'").iloc[0]
    # Downsample the objective ot have the same population as the self_report
    downsampled_data = data[~data[self_report_var].isna()]
    downsampled_uncensored = uncensored[~data[self_report_var].isna()]
    downsampled_entry_age = entry_age[~data[self_report_var].isna()]
    covariate_formula = "BMI + smoking_ever"
    survival_test = smf.phreg(formula = f"age_at_death_censored ~ {activity_var} + sex + {covariate_formula}",
                                    data=downsampled_data,
                                    status=downsampled_uncensored,
                                    entry=downsampled_entry_age,
                                    ).fit()
    pvalues = pandas.Series(survival_test.pvalues, index=survival_test.model.exog_names)
    params = pandas.Series(survival_test.params, index=survival_test.model.exog_names)
    if self_report in data and data[self_report].dtype.name == "category":
        cat_counts = data[self_report].value_counts()
        most_common = cat_counts.idxmax()
        dropped_cats = cat_counts.index[cat_counts < 400].to_list() # Want at least this many in the category
        cats = '"' + '", "'.join(cat_counts.index[cat_counts >= 400].to_list()) + '"'
        subset = ~data[self_report].isin(dropped_cats)
        self_report_survival_test = smf.phreg(
            formula = f'age_at_death_censored ~ C({self_report}, levels=[{cats}]) + sex + {covariate_formula}',
            #formula = f'age_at_death_censored ~ {self_report} + sex + {covariate_formula}',
            data=data[subset],
            status=uncensored[subset],
            entry=entry_age[subset],
        ).fit()
        self_report_pvalue = self_report_survival_test.f_test(
            [[1 if j == i else 0 for j in range(len(self_report_survival_test.model.exog_names))]
                for i, var in enumerate(self_report_survival_test.model.exog_names)
                if self_report in var]
        ).pvalue
        self_report_pvalue = self_report_pvalue.flatten()[0]
    else:
        # Non-categorical values don't have an alternative, use the same as before
        self_report_pvalue = self_report_survival.p
    df = pandas.DataFrame.from_dict({
        "self_report": {
            "survival_p": self_report_survival.p,
            "survival_logHR": self_report_survival['standardized log Hazard Ratio'],
        },
        "objective": {
            "survival_p": self_report_objective.p,
            "survival_logHR": self_report_objective['standardized log Hazard Ratio'],
        },
        "downsampled_objective": {
            "survival_p": pvalues[activity_var],
            "survival_logHR": params[activity_var] * downsampled_data[activity_var].std(),
        },
        "self_report_cat": {
            "survival_p": self_report_pvalue,
        },
    })
    print(df.T)
    # Plot the points
    #ax1.scatter([i], [-numpy.log10(df.loc["survival_p", "self_report"])], c="k")
    #ax1.scatter([i], [-numpy.log10(df.loc["survival_p", "downsampled_objective"])], c="r")
    #ax2.scatter([i], [-numpy.log10(df.loc["survival_p", "self_report_cat"])], c="k")
    #ax2.scatter([i], [-numpy.log10(df.loc["survival_p", "objective"])], c="r")
    BAR_WIDTH = 0.6
    ax2.bar(2*i-BAR_WIDTH/2, -numpy.log10(df.loc['survival_p', 'self_report_cat']), color="k", width=BAR_WIDTH)
    ax2.bar(2*i+BAR_WIDTH/2, -numpy.log10(df.loc['survival_p', 'objective']), color="r", width=BAR_WIDTH)
#ax1.set_ylabel("-log10 p-value")
#ax1.set_xticks(range(len(variable_pairs)))
#ax1.set_xticklabels([f"{self_report}\n{activity_var}" for self_report, activity_var in variable_pairs])
#ax2.set_xticks(range(len(variable_pairs)))
#ax2.set_xticklabels([f"{self_report}\n{activity_var}" for self_report, activity_var in variable_pairs])
#ax1.xaxis.set_tick_params(rotation=90)
ax2.set_xticks(2*numpy.arange(len(variable_pairs)))
ax2.set_xticklabels([name for name in variable_pairs.keys()])
ax2.set_ylabel("-log10 p-value")
ax2.xaxis.set_tick_params(rotation=90)
legend_elts = [matplotlib.lines.Line2D(
                        [0],[0],
                        marker="o", markerfacecolor=c, markersize=10,
                        label=l,
                        c=c, lw=0)
                    for c, l in zip(["k", "r"], ["Subjective", "Objective"])]
ax2.legend(handles=legend_elts)
fig.tight_layout()
fig.savefig(OUTDIR+"subjective_objective_comparison.survival.png")


# Compare the objective/subjective variables again on the phecodes
fig, axes = pylab.subplots(ncols=len(variable_pairs), figsize=(15,7), sharex=True, sharey=True)
Q_CUTOFF = 0.05
for ax, (name, (self_report, activity_var)) in zip(axes.flatten(), variable_pairs.items()):
    if self_report is not "sleep_duration_abs_dev":
        self_report_var = "self_report_" + self_report_circadian_variables[self_report]['name']
    else:
        self_report_var = "self_report_sleep_duration_abs_dev"
    either_significant = phecode_tests[
        ((phecode_tests.activity_var == self_report_var) |
         (phecode_tests.activity_var == activity_var)) &
        (phecode_tests.q < Q_CUTOFF)].phecode.unique()
    downsampled_data = data[~data[self_report_var].isna()]
    objective_tests_list = []
    for phecode in either_significant:
        covariate_formula = ' + '.join(c for c in covariates if c != 'sex')
        N = data[phecode].sum()
        fit = OLS(f"{activity_var} ~ Q({phecode}) + sex * ({covariate_formula})",
                     data=downsampled_data)
        p = fit.pvalues[f"Q({phecode})"]
        coeff = fit.params[f"Q({phecode})"]
        std_effect = coeff / data[activity_var].std()
        N_cases = data.loc[~data[activity_var].isna(), phecode].sum()
        objective_tests_list.append({"phecode": phecode,
                                "activity_var": activity_var,
                                "p": p,
                                "coeff": coeff,
                                "std_effect": std_effect,
                                "N_cases": N_cases,
                               })
    objective_tests = pandas.DataFrame(objective_tests_list)
    subjective_ps = objective_tests.phecode.map(phecode_tests[(phecode_tests.activity_var == self_report_var)].set_index("phecode").p)
    colors = objective_tests.phecode.map(phecode_info.category).map(color_by_phecode_cat)
    ax.scatter(-numpy.log10(objective_tests.p), -numpy.log10(subjective_ps), c=colors)
    ax.set_aspect("equal")
    ax.set_xlabel(f"-log10 p objective variable")
    if ax == axes.flatten()[0]:
        ax.set_ylabel(f"-log10 p subjective variable")
    ax.set_title(name)
for ax in axes.flatten():
    bound = max(abs(numpy.min([ax.get_xlim(), ax.get_ylim()])),
                numpy.max([ax.get_xlim(), ax.get_ylim()]))
    diag = numpy.array([0, bound])
    ax.plot(diag, diag, linestyle="--", c='k', zorder=-1, label="diagonal")
    ax.set_xlim(0,20)
    ax.set_ylim(0,20)
fig.tight_layout()
fig.savefig(OUTDIR+"objective_subjective_comparison.png")

## Investigate phenotypes by diagnosis age
def plot_by_diagnosis_date(ICD10_codes, phecode, phenotype_name):
    # gather icd10 entries by the date of first diagnosis
    Q_CUTOFF = 0.01
    NUM_GROUPS = 20 # number of equal-sized groupsgroups to break the diagnosis dates into
    GROUPS_PER_POINT = 5 # Number of those groups to use at once: use 1 for non-overlapping, higher numbers to get overlap of adjacent
    in_icd10_codes = numpy.any([icd10_entries.ICD10.str.startswith(code) for code in ICD10_codes], axis=0)
    diag_dates = icd10_entries[in_icd10_codes].groupby("ID").first_date.first()
    activity_vars_to_test = phecode_tests[(phecode_tests.phecode == phecode) & (phecode_tests.q < Q_CUTOFF)].activity_var.unique()

    date_data = data.copy()
    date_data['diag_date'] = pandas.to_datetime(pandas.Series(data.index.map(diag_dates), data.index))
    date_data['diag_date_minus_actigraphy'] = (date_data.diag_date - date_data.actigraphy_start_date) / pandas.to_timedelta("1Y")
    date_data['diag_date_group'] = pandas.qcut(date_data.diag_date_minus_actigraphy, NUM_GROUPS)
    by_date_list = []
    for i, date_group in enumerate(date_data.diag_date_group.cat.categories):
        date_groups = date_data.diag_date_group.cat.categories[i:i+GROUPS_PER_POINT]
        if len(date_groups) < GROUPS_PER_POINT:
            continue # Insufficient groups, reached end of the study
        mid_point = date_groups[GROUPS_PER_POINT//2].mid
        date_data['in_group'] = (date_data.diag_date_group.isin(date_groups)).astype(int)
        # Compare only to people who do not have the phecode from any source
        date_data['use'] = date_data.in_group | (date_data[phecode] == 0)
        for activity_var in activity_vars_to_test:
            if activity_var.startswith("self_report"):
                continue # Not based off actigraphy, skip
            covariate_formula = ' + '.join(c for c in covariates if c != 'sex')
            N = data[phecode].sum()
            fit = OLS(f"{activity_var} ~ in_group + sex * ({covariate_formula})",
                         data=date_data,
                         subset=date_data.use)
            p = fit.pvalues["in_group"]
            coeff = fit.params["in_group"]
            std_effect = coeff / data[activity_var].std()
            N_cases = data.loc[~data[activity_var].isna(), phecode].sum()
            by_date_list.append({
                                    "group": date_group,
                                    "group_mid": mid_point,
                                    "activity_var": activity_var,
                                    "p": p,
                                    "coeff": coeff,
                                    "std_effect": std_effect,
                                    "N_cases": N_cases,
                                   })
    by_date = pandas.DataFrame(by_date_list)

    fig, ax = pylab.subplots(figsize=(10,7))
    def plot_var(data):
        data = data.sort_values(by="group_mid")
        cat = activity_variable_descriptions.Subcategory[data.activity_var.iloc[0]]
        color = color_by_actigraphy_subcat[cat]
        #cat = activity_variable_descriptions.Category[data.activity_var.iloc[0]]
        #color = color_by_actigraphy_cat[cat]
        ax.plot(-data.group_mid,
                data.std_effect.abs(),
                #-numpy.log10(data.p),
                c = color, label=cat,
                linewidth=3)
    ax.set_xlabel("Years since diagnosis")
    ax.set_ylabel("Effect size")
    ax.set_title(phenotype_name)
    by_date.groupby('activity_var').apply(plot_var)
    legend_from_colormap(fig, color_by_actigraphy_subcat, fontsize="small", loc="upper left")
    return fig, ax, by_date

fig, ax, hypertension_by_date = plot_by_diagnosis_date(["I10"], 401, "Hypertension")
fig.savefig(OUTDIR+"by_date.hypertension.png")

fig, ax, diabetes_by_date = plot_by_diagnosis_date(["E09", "E10", "E11", "E13"], 250, "Diabetes")
fig.savefig(OUTDIR+"by_date.diabetes.png")

fig, ax, IHD_by_date = plot_by_diagnosis_date(["I20", "I21", "I22", "I23", "I25"], 411, "Ischemic Heart Disease")
fig.savefig(OUTDIR+"by_date.IHD.png")

fig, ax, pneumonia_by_date = plot_by_diagnosis_date(["J09", "J10", "J11", "J12", "J13", "J14", "J15", "J16", "J17", "J18"], 480, "Pneumonia")
fig.savefig(OUTDIR+"by_date.pneumonia.png")

fig, ax, mood_disorders_by_date = plot_by_diagnosis_date(["F30", "F31", "F32", "F33", "F34", "F39"], 296, "Mood Disorders")
fig.savefig(OUTDIR+"by_date.mood_disorders.png")

fig, ax, anxiety_by_date = plot_by_diagnosis_date(["F40", "F41"], 300, "Anxiety Disorders")
fig.savefig(OUTDIR+"by_date.anxiety_disorders.png")

fig, ax, renal_failure_by_date = plot_by_diagnosis_date(["N17", "N18", "N19", "Y60", "Y84", "Z49"], 585, "Renal Failure")
fig.savefig(OUTDIR+"by_date.renal_failure.png")

#### Combine all tables into the summary with the header file
print(f"Writing out the complete results table to {OUTDIR+'results.xlsx'}")
import openpyxl
workbook = openpyxl.load_workbook("../table_header.xlsx")
descriptions = workbook['Variables']
start_column = len(list(descriptions.tables.values())[0].tableColumns) + 1
var_stats = data[activity_variable_descriptions.index].describe(percentiles=[0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]).T
for i, col in enumerate(var_stats.columns):
    descriptions.cell(1, start_column + i, col)
    for j, value in enumerate(var_stats[col].values):
        descriptions.cell(j+2, start_column + i, value)
workbook.save(OUTDIR+"results.xlsx")
with pandas.ExcelWriter(OUTDIR+"results.xlsx", mode="a") as writer:
    survival_tests_raw.sort_values(by="p").to_excel(writer, sheet_name="Survival Associations", index=False)
    phecode_tests_raw.sort_values(by="p").to_excel(writer, sheet_name="Phecode Associations", index=False)
    quantitative_tests_raw.sort_values(by="p").to_excel(writer, sheet_name="Quantitative Associations", index=False)
    phecode_tests_by_sex_raw.sort_values(by="p_diff").to_excel(writer, sheet_name="Sex-specific Associations", index=False)
    age_tests.sort_values(by="p").to_excel(writer, sheet_name="Age-dependence", index=False)
